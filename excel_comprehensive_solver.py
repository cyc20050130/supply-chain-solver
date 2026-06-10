from __future__ import annotations

import argparse
import json
import math
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import Workbook

import solve


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


DAYS = 30
DEFAULT_XLS = Path(r"C:\Users\cyc20\Desktop\excel\电视综合运营★★★☆-标准版个人练习（15_18）场.xls")
OUT_DIR = Path(r"C:\Users\cyc20\Desktop\excel")
SHORTAGE_PENALTY = 100_000_000.0


def log(message: str) -> None:
    print(f"[excel-comprehensive] {message}")


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def cost_score(actual: float, target: float, points: float) -> float:
    if target <= 0:
        return 0.0
    return clamp((1 - (actual - target) / (target * 0.2)) * points, 0.0, points)


def deviation_score(actual: float, target: float, points: float) -> float:
    if target <= 0:
        return 0.0
    return clamp((1 - actual / target) * points, 0.0, points)


def satisfaction_score(actual_ratio: float, points: float) -> float:
    return clamp(actual_ratio * points, 0.0, points)


def spread_integer(total: float, days: int = DAYS) -> list[int]:
    total_int = int(math.ceil(max(0.0, total)))
    base = total_int // days
    remainder = total_int - base * days
    return [base + (1 if idx < remainder else 0) for idx in range(days)]


def read_sections_xlrd(xls_path: Path) -> dict[str, list[tuple[Any, ...]]]:
    """Read old BIFF .xls safely through an ASCII temp copy, preserving solve.py section shape."""
    xls_path = xls_path.resolve()
    if not xls_path.exists():
        raise FileNotFoundError(xls_path)

    tmp_dir = Path(tempfile.mkdtemp(prefix="sc_excel_xls_"))
    try:
        ascii_xls = tmp_dir / "case.xls"
        shutil.copy2(xls_path, ascii_xls)
        book = xlrd.open_workbook(str(ascii_xls))
        sheet = book.sheet_by_index(0)
        raw_rows = [tuple(sheet.cell_value(row, col) for col in range(sheet.ncols)) for row in range(sheet.nrows)]
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

    title_keywords = [
        "产品及原料清单",
        "货币汇率表",
        "销售网点历史销售数据",
        "销售网点销售数据",
        "销售网点历史数据",
        "销售网点",
        "供应商产能数据",
        "工厂原料消耗及期初库存",
        "工厂产能及期初库存",
        "工厂产能",
        "承运商信息",
        "地点信息",
        "运输路线",
    ]
    title_idx: dict[str, int] = {}
    for idx, row in enumerate(raw_rows):
        row_text = " ".join(solve.sv(cell) for cell in row if solve.sv(cell))
        for keyword in title_keywords:
            if keyword in row_text and keyword not in title_idx:
                title_idx[keyword] = idx

    sections: dict[str, list[tuple[Any, ...]]] = {}
    ordered = sorted(title_idx.items(), key=lambda item: item[1])
    for pos, (title, start) in enumerate(ordered):
        end = ordered[pos + 1][1] if pos + 1 < len(ordered) else len(raw_rows)
        rows = [raw_rows[i] for i in range(start + 1, end) if any(solve.sv(cell) for cell in raw_rows[i])]
        if rows:
            sections[title] = rows
    return sections


def parse_case(xls_path: Path) -> dict[str, Any]:
    sections = read_sections_xlrd(xls_path)
    qtype = solve.detect_type(sections, xls_path)
    if qtype != "综合":
        raise ValueError(f"当前脚本只处理综合案例，识别到: {qtype}")
    products = solve.parse_products(sections)
    factories = solve.parse_factories(sections)
    factory_materials = solve.parse_factory_materials(sections)
    suppliers = solve.parse_suppliers(sections)
    routes = solve.parse_routes(sections)
    rates = solve.parse_rates(sections)
    sales = solve.parse_sales(sections)
    carriers = solve.parse_carriers_safe(sections, rates)
    forecasts = [solve.forecast_node(node) for node in sales]
    return {
        "xls_path": xls_path,
        "sections": sections,
        "products": products,
        "factories": factories,
        "factory_materials": factory_materials,
        "suppliers": suppliers,
        "routes": routes,
        "rates": rates,
        "sales": sales,
        "carriers": carriers,
        "forecasts": forecasts,
    }


def route_cost_expr(route: solve.Route, amount_var: Any, used_var: Any, ratio: float, pulp: Any) -> Any:
    billable_rate = route.rate * max(ratio, 0.001)
    freight = pulp.LpVariable(f"freight_{abs(hash((route.route, str(amount_var))))}", lowBound=0)
    freight_problem = (freight, [
        freight >= billable_rate * amount_var,
        freight >= route.min_freight * used_var,
    ])
    return freight_problem


def solve_transport_milp(
    *,
    name: str,
    sources: dict[str, dict[str, Any]],
    destinations: dict[str, dict[str, Any]],
    routes: list[solve.Route],
    products: list[solve.Product],
    cargo: str,
    carriers: list[Any] | None,
) -> dict[str, Any]:
    import pulp

    ratio = solve.charge_ratio(products, cargo)
    lane_options = [(idx, route) for idx, route in enumerate(routes) if route.src in sources and route.dst in destinations]
    if not lane_options:
        return {
            "status": "NoRoute",
            "cargo": cargo,
            "shipments": [],
            "freight_cost": 0.0,
            "source_cost": 0.0,
            "shortage": sum(sum(info.get("demand", [])) for info in destinations.values()),
            "failures": [f"{cargo} 没有可用运输路线"],
        }

    max_amount = max(
        sum(float(src.get("initial", 0.0)) + sum(src.get("supply", [])) for src in sources.values()),
        sum(sum(dst.get("demand", [])) for dst in destinations.values()),
        1.0,
    )

    model = pulp.LpProblem(name, pulp.LpMinimize)
    route_by_lane = {idx: route for idx, route in lane_options}
    x: dict[tuple[int, int], Any] = {}
    y: dict[tuple[int, int], Any] = {}
    fcost: dict[tuple[int, int], Any] = {}
    src_inv: dict[tuple[str, int], Any] = {}
    dst_inv: dict[tuple[str, int], Any] = {}
    shortage: dict[tuple[str, int], Any] = {}

    for lane_id, route in lane_options:
        for day in range(1, DAYS + 1):
            key = (lane_id, day)
            x[key] = pulp.LpVariable(f"x_{lane_id}_{day}", lowBound=0, cat="Integer")
            y[key] = pulp.LpVariable(f"y_{lane_id}_{day}", lowBound=0, upBound=1, cat="Binary")
            fcost[key] = pulp.LpVariable(f"fcost_{lane_id}_{day}", lowBound=0)
            model += x[key] <= max_amount * y[key]
            if route.min_qty > 0:
                model += x[key] >= route.min_qty * y[key]
            model += fcost[key] >= route.rate * ratio * x[key]
            model += fcost[key] >= route.min_freight * y[key]

    for source, info in sources.items():
        supply = list(info.get("supply", []))
        limit = info.get("limit")
        for day in range(1, DAYS + 1):
            src_inv[(source, day)] = pulp.LpVariable(
                f"src_inv_{len(src_inv)}",
                lowBound=0,
                upBound=limit if isinstance(limit, (int, float)) and limit > 0 else None,
            )
            outbound = [x[(lane_id, day)] for lane_id, route in lane_options if route.src == source]
            prev = float(info.get("initial", 0.0)) if day == 1 else src_inv[(source, day - 1)]
            today_supply = supply[day - 1] if day - 1 < len(supply) else 0.0
            model += src_inv[(source, day)] == prev + today_supply - pulp.lpSum(outbound)

    for destination, info in destinations.items():
        demand = list(info.get("demand", []))
        limit = info.get("limit")
        for day in range(1, DAYS + 1):
            dst_inv[(destination, day)] = pulp.LpVariable(
                f"dst_inv_{len(dst_inv)}",
                lowBound=0,
                upBound=limit if isinstance(limit, (int, float)) and limit > 0 else None,
            )
            shortage[(destination, day)] = pulp.LpVariable(f"short_{len(shortage)}", lowBound=0)
            arrivals = []
            for lane_id, route in lane_options:
                if route.dst != destination:
                    continue
                ship_day = day - route.lead
                if 1 <= ship_day <= DAYS:
                    arrivals.append(x[(lane_id, ship_day)])
            prev = float(info.get("initial", 0.0)) if day == 1 else dst_inv[(destination, day - 1)]
            today_demand = demand[day - 1] if day - 1 < len(demand) else 0.0
            model += dst_inv[(destination, day)] == prev + pulp.lpSum(arrivals) - today_demand + shortage[(destination, day)]

    source_cost_terms = []
    for (lane_id, day), var in x.items():
        unit_cost = float(sources.get(route_by_lane[lane_id].src, {}).get("unit_cost", 0.0) or 0.0)
        if unit_cost:
            source_cost_terms.append(unit_cost * var)

    model += pulp.lpSum(fcost.values()) + pulp.lpSum(source_cost_terms) + SHORTAGE_PENALTY * pulp.lpSum(shortage.values())
    status_code = model.solve(pulp.PULP_CBC_CMD(msg=False, gapRel=0.001, timeLimit=600))
    status = pulp.LpStatus[status_code]

    shipments: list[dict[str, Any]] = []
    source_cost = 0.0
    for lane_id, route in lane_options:
        for day in range(1, DAYS + 1):
            amount = int(round(pulp.value(x[(lane_id, day)]) or 0))
            if amount <= 0:
                continue
            unit_cost = float(sources.get(route.src, {}).get("unit_cost", 0.0) or 0.0)
            row = {
                "cargo": cargo,
                "source": route.src,
                "destination": route.dst,
                "amount": amount,
                "ship_day": day,
                "arrival_day": day + route.lead,
                "route": route.route,
                "carrier": solve.route_carrier(route, carriers),
                "mode": solve.route_mode(route),
                "lead": route.lead,
                "unit_cost": unit_cost,
                "source_cost": amount * unit_cost,
                "freight_cost": float(pulp.value(fcost[(lane_id, day)]) or solve.route_cost(route, amount, ratio)),
            }
            source_cost += row["source_cost"]
            shipments.append(row)

    total_shortage = 0.0
    failures: list[str] = []
    daily_shortage: list[dict[str, Any]] = []
    for (destination, day), var in shortage.items():
        short = float(pulp.value(var) or 0.0)
        total_shortage += short
        if short > 1e-6:
            daily_shortage.append({"destination": destination, "day": day, "shortage": short})
            failures.append(f"{cargo}-{destination} 第{day}天缺口 {short:.0f}")

    return {
        "status": status,
        "cargo": cargo,
        "shipments": sorted(shipments, key=lambda row: (row["ship_day"], row["route"], row["amount"])),
        "freight_cost": sum(row["freight_cost"] for row in shipments),
        "source_cost": source_cost,
        "shortage": total_shortage,
        "daily_shortage": daily_shortage,
        "failures": failures,
    }


def build_product_transport(case: dict[str, Any], production_daily: list[int]) -> dict[str, Any]:
    factory = case["factories"][0]
    product_name = factory.product or next((p.name for p in case["products"] if p.kind == "产品"), "")
    destinations = {
        row["node"]: {
            "initial": row["init"],
            "demand": spread_integer(row["forecast"]),
            "limit": row["limit"],
        }
        for row in case["forecasts"]
    }
    sources = {
        factory.name: {
            "initial": factory.init,
            "supply": production_daily,
            "limit": factory.limit,
        }
    }
    routes = [route for route in case["routes"] if route.src == factory.name and route.dst in destinations]
    return solve_transport_milp(
        name="product_transport",
        sources=sources,
        destinations=destinations,
        routes=routes,
        products=case["products"],
        cargo=product_name,
        carriers=case["carriers"],
    )


def build_material_transport(case: dict[str, Any], material: solve.FactoryMaterial, production_daily: list[int]) -> dict[str, Any]:
    demand = [qty * solve.material_bom(case["products"], material.material) for qty in production_daily]
    sources: dict[str, dict[str, Any]] = {}
    for supplier in case["suppliers"]:
        if supplier.material != material.material:
            continue
        sources[supplier.name] = {
            "initial": supplier.init,
            "supply": [supplier.daily] * DAYS,
            "limit": supplier.available if supplier.available > 0 else None,
            "unit_cost": solve.currency_to_cny(supplier.price, supplier.currency, case["rates"]),
        }
    destinations = {
        material.factory: {
            "initial": material.init,
            "demand": demand,
            "limit": material.limit,
        }
    }
    routes = [route for route in case["routes"] if route.dst == material.factory and route.src in sources]
    transport = solve_transport_milp(
        name=f"material_{material.material}",
        sources=sources,
        destinations=destinations,
        routes=routes,
        products=case["products"],
        cargo=material.material,
        carriers=case["carriers"],
    )
    for row in transport["shipments"]:
        row["material"] = material.material
        row["supplier"] = row["source"]
        row["factory"] = material.factory
        row["purchase_cost"] = row["source_cost"]
    return transport


def score_candidate(case: dict[str, Any], production_target: int, product_transport: dict[str, Any], material_transports: list[dict[str, Any]]) -> dict[str, Any]:
    forecast_total = sum(row["forecast"] for row in case["forecasts"])
    product_shortage = float(product_transport.get("shortage", 0.0) or 0.0)
    material_shortage_units = []
    for material, transport in zip(case["factory_materials"], material_transports):
        bom = solve.material_bom(case["products"], material.material)
        material_shortage_units.append((float(transport.get("shortage", 0.0) or 0.0)) / max(bom, 0.001))
    production_shortage = max(material_shortage_units) if material_shortage_units else 0.0
    actual_production = max(0.0, production_target - production_shortage)
    actual_sales = max(0.0, forecast_total - product_shortage)
    total_freight = product_transport["freight_cost"] + sum(t["freight_cost"] for t in material_transports)
    total_purchase = sum(t["source_cost"] for t in material_transports)
    unit_logistics = total_freight / max(actual_sales, 0.001)
    unit_procurement = total_purchase / max(actual_production, 0.001)
    production_satisfaction = actual_production / max(production_target, 0.001) if production_target else 1.0
    market_satisfaction = actual_sales / max(forecast_total, 0.001) if forecast_total else 1.0

    targets = {
        "prediction_deviation": 0.05,
        "unit_logistics_cost": 110.0,
        "unit_procurement_cost": 1600.0,
        "production_satisfaction": 1.0,
        "market_satisfaction": 1.0,
    }
    prediction_deviation = 0.0
    score_rows = [
        {"item": "预测偏差率", "target": targets["prediction_deviation"], "actual": prediction_deviation, "points": deviation_score(prediction_deviation, targets["prediction_deviation"], 10), "max": 10},
        {"item": "单位物流成本", "target": targets["unit_logistics_cost"], "actual": unit_logistics, "points": cost_score(unit_logistics, targets["unit_logistics_cost"], 20), "max": 20},
        {"item": "单位采购成本", "target": targets["unit_procurement_cost"], "actual": unit_procurement, "points": cost_score(unit_procurement, targets["unit_procurement_cost"], 20), "max": 20},
        {"item": "生产满足率", "target": targets["production_satisfaction"], "actual": production_satisfaction, "points": satisfaction_score(production_satisfaction, 15), "max": 15},
        {"item": "市场满足率", "target": targets["market_satisfaction"], "actual": market_satisfaction, "points": satisfaction_score(market_satisfaction, 35), "max": 35},
    ]
    return {
        "production_target": production_target,
        "actual_production": actual_production,
        "actual_sales": actual_sales,
        "unit_logistics": unit_logistics,
        "unit_procurement": unit_procurement,
        "production_satisfaction": production_satisfaction,
        "market_satisfaction": market_satisfaction,
        "total_freight": total_freight,
        "total_purchase": total_purchase,
        "score_rows": score_rows,
        "score": sum(row["points"] for row in score_rows),
    }


def solve_comprehensive_case(xls_path: Path) -> dict[str, Any]:
    case = parse_case(xls_path)
    factory = case["factories"][0]
    forecast_total = sum(row["forecast"] for row in case["forecasts"])
    sales_init = sum(row["init"] for row in case["forecasts"])
    capacity = int(factory.daily * DAYS)
    minimum_for_market = int(math.ceil(max(0.0, forecast_total - sales_init - factory.init)))
    base = int(math.ceil(max(minimum_for_market, min(capacity, forecast_total - sales_init))))
    candidates = {
        minimum_for_market,
        base,
        capacity,
        int(capacity * 0.85),
        int(capacity * 0.90),
        int(capacity * 0.95),
        int(capacity * 0.98),
    }
    candidates = sorted({max(0, min(capacity, target)) for target in candidates if target > 0})
    if not candidates:
        candidates = [capacity]

    solved: list[dict[str, Any]] = []
    for target in candidates:
        log(f"求解候选生产量 {target}")
        production_daily = spread_integer(target)
        product_transport = build_product_transport(case, production_daily)
        material_transports = [
            build_material_transport(case, material, production_daily)
            for material in case["factory_materials"]
        ]
        score = score_candidate(case, target, product_transport, material_transports)
        failures = product_transport.get("failures", [])
        for transport in material_transports:
            failures.extend(transport.get("failures", []))
        solved.append(
            {
                **score,
                "production_daily": production_daily,
                "product_transport": product_transport,
                "material_transports": material_transports,
                "failures": failures,
                "statuses": {
                    "product": product_transport["status"],
                    "materials": {transport["cargo"]: transport["status"] for transport in material_transports},
                },
            }
        )

    best = max(solved, key=lambda row: (row["score"], row["market_satisfaction"], -row["unit_logistics"]))
    return {
        "case": xls_path.name,
        "xls_path": str(xls_path),
        "solver": "Excel comprehensive MILP functions",
        "candidate_targets": candidates,
        "sales_forecast": case["forecasts"],
        "production": {
            "factory": factory.name,
            "product": factory.product,
            "target": best["production_target"],
            "actual": best["actual_production"],
            "capacity": capacity,
            "daily": best["production_daily"],
        },
        "score": best["score"],
        "score_rows": best["score_rows"],
        "metrics": {
            "unit_logistics": best["unit_logistics"],
            "unit_procurement": best["unit_procurement"],
            "production_satisfaction": best["production_satisfaction"],
            "market_satisfaction": best["market_satisfaction"],
            "total_freight": best["total_freight"],
            "total_purchase": best["total_purchase"],
        },
        "product_transport": best["product_transport"],
        "material_transports": best["material_transports"],
        "failures": best["failures"],
        "statuses": best["statuses"],
        "all_candidates": [
            {
                "production_target": row["production_target"],
                "score": row["score"],
                "unit_logistics": row["unit_logistics"],
                "unit_procurement": row["unit_procurement"],
                "production_satisfaction": row["production_satisfaction"],
                "market_satisfaction": row["market_satisfaction"],
                "statuses": row["statuses"],
                "failure_count": len(row["failures"]),
            }
            for row in solved
        ],
    }


def compact_shipments(shipments: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str, int], list[dict[str, Any]]] = {}
    for row in shipments:
        key = (row["cargo"], row["route"], row.get("carrier", ""), int(row["amount"]))
        grouped.setdefault(key, []).append(row)
    compact: list[dict[str, Any]] = []
    for (cargo, route, carrier, amount), rows in sorted(grouped.items(), key=lambda item: (item[0][0], item[0][1], item[0][3])):
        days = sorted(int(row["ship_day"]) for row in rows)
        gaps = [b - a for a, b in zip(days, days[1:])]
        interval = int(round(sum(gaps) / len(gaps))) if gaps else 1
        compact.append({
            "cargo": cargo,
            "route": route,
            "carrier": carrier,
            "amount_per_trip": amount,
            "start_day": days[0],
            "trips": len(rows),
            "interval_days": max(1, interval),
            "total_amount": amount * len(rows),
        })
    return compact


def export_outputs(result: dict[str, Any], out_dir: Path = OUT_DIR) -> dict[str, str]:
    out_dir.mkdir(parents=True, exist_ok=True)
    stem = Path(result["xls_path"]).stem
    json_path = out_dir / f"{stem}_excel规划求解.json"
    xlsx_path = out_dir / f"{stem}_excel规划求解结果.xlsx"
    json_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = "销售预测"
    ws.append(["销售网点", "预测销量", "日均", "安全库存", "期初库存", "库存上限", "方法"])
    for row in result["sales_forecast"]:
        ws.append([row["node"], row["forecast"], row["daily_avg"], row["safety_stock"], row["init"], row["limit"], row["method"]])

    ws = wb.create_sheet("生产计划")
    ws.append(["工厂", "产品", "计划生产量", "实际可生产", "30天产能"])
    p = result["production"]
    ws.append([p["factory"], p["product"], p["target"], p["actual"], p["capacity"]])
    ws.append([])
    ws.append(["天", "生产量"])
    for day, amount in enumerate(p["daily"], start=1):
        ws.append([day, amount])

    ws = wb.create_sheet("采购计划")
    ws.append(["原料", "供应商", "工厂", "采购量", "采购成本", "运费"])
    procurement: dict[tuple[str, str, str], dict[str, Any]] = {}
    for transport in result["material_transports"]:
        for row in transport["shipments"]:
            key = (row["cargo"], row["source"], row["destination"])
            current = procurement.setdefault(key, {"amount": 0, "purchase": 0.0, "freight": 0.0})
            current["amount"] += row["amount"]
            current["purchase"] += row["source_cost"]
            current["freight"] += row["freight_cost"]
    for (cargo, source, destination), values in sorted(procurement.items()):
        ws.append([cargo, source, destination, values["amount"], values["purchase"], values["freight"]])

    ws = wb.create_sheet("运输填报")
    ws.append(["货物", "路线", "承运商", "每趟运输数量", "起运日期", "承运趟数", "几天一趟", "合计"])
    shipments = list(result["product_transport"]["shipments"])
    for transport in result["material_transports"]:
        shipments.extend(transport["shipments"])
    for row in compact_shipments(shipments):
        ws.append([row["cargo"], row["route"], row["carrier"], row["amount_per_trip"], f"1-{row['start_day']:02d}", row["trips"], row["interval_days"], row["total_amount"]])

    ws = wb.create_sheet("评分")
    ws.append(["评分项", "指标值", "实际值", "得分", "满分"])
    for row in result["score_rows"]:
        target = row["target"] * 100 if row["target"] <= 1 else row["target"]
        actual = row["actual"] * 100 if row["target"] <= 1 else row["actual"]
        ws.append([row["item"], target, actual, row["points"], row["max"]])
    ws.append([])
    ws.append(["候选总分", result["score"]])
    ws.append(["说明", "预测偏差率在没有平台未来真实销量前按0%候选；提交平台后应回放真实成绩。"])

    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            width = min(42, max(10, max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells) + 2))
            sheet.column_dimensions[column_cells[0].column_letter].width = width
    wb.save(xlsx_path)
    return {"json": str(json_path), "xlsx": str(xlsx_path)}


def self_test() -> None:
    assert abs(cost_score(423.71, 415, 60) - 53.70) < 0.05
    assert abs(satisfaction_score(0.9842, 40) - 39.37) < 0.05
    assert deviation_score(0.1612, 0.05, 10) == 0.0
    assert abs(cost_score(1809.29, 1600, 20) - 6.92) < 0.05
    assert sum(spread_integer(1234)) == 1234


def main() -> None:
    parser = argparse.ArgumentParser(description="基于桌面 excel 综合案例 .xls 的函数化规划求解器")
    parser.add_argument("xls", nargs="?", type=Path, default=DEFAULT_XLS)
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args()
    if args.self_test:
        self_test()
        print("self-test passed")
        return

    result = solve_comprehensive_case(args.xls)
    paths = export_outputs(result)
    print(json.dumps({
        "case": result["case"],
        "score": round(result["score"], 4),
        "production": result["production"],
        "metrics": result["metrics"],
        "statuses": result["statuses"],
        "failures": result["failures"][:10],
        "outputs": paths,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
