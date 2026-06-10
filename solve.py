"""
供应链表格启发式工作流

用法:
  python -X utf8 solve.py "采购\\白砂糖采购★★-标准版个人练习（17_38）场.xls"
  python -X utf8 solve.py --all
  python -X utf8 solve.py --self-test
  python -X utf8 solve.py --check-env

输出:
  与输入 .xls 同目录生成 xxx_求解方案.html，并删除同名旧方案页。
"""

from __future__ import annotations

import argparse
import html
import json
import math
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
from collections import defaultdict
from dataclasses import dataclass
from itertools import combinations as iter_combinations
from itertools import permutations as iter_permutations
from itertools import product as iter_product
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook


PLAN_DAYS = 30
ACTIVE_SOLVER_MODE = "high"
ACTIVE_HAR_CONTEXT: dict[str, Any] | None = None


def set_solver_mode(mode: str) -> None:
    global ACTIVE_SOLVER_MODE
    ACTIVE_SOLVER_MODE = "extreme" if str(mode).lower() in {"extreme", "limit", "极限版", "极限"} else "high"


def solver_mode_label() -> str:
    return "极限版" if ACTIVE_SOLVER_MODE == "extreme" else "高分版"


def is_extreme_mode() -> bool:
    return ACTIVE_SOLVER_MODE == "extreme"


def heatwater_pattern_audit_enabled() -> bool:
    value = os.environ.get("SUPPLY_CHAIN_HEATWATER_PATTERN_AUDIT", "")
    return str(value).strip().lower() in {"1", "true", "yes", "on", "audit"}


def heatwater_global_audit_enabled() -> bool:
    value = os.environ.get("SUPPLY_CHAIN_HEATWATER_GLOBAL_AUDIT", "")
    return str(value).strip().lower() in {"1", "true", "yes", "on", "audit"}


def heatwater_full_global_audit_enabled() -> bool:
    value = os.environ.get("SUPPLY_CHAIN_HEATWATER_FULL_GLOBAL_AUDIT", "")
    return str(value).strip().lower() in {"1", "true", "yes", "on", "audit"}


def heatwater_counterexample_audit_enabled() -> bool:
    value = os.environ.get("SUPPLY_CHAIN_HEATWATER_COUNTEREXAMPLE_AUDIT", "")
    return str(value).strip().lower() in {"1", "true", "yes", "on", "audit", "counterexample"}


def heatwater_subset_audit_enabled() -> bool:
    value = os.environ.get("SUPPLY_CHAIN_HEATWATER_SUBSET_AUDIT", "")
    return str(value).strip().lower() in {"1", "true", "yes", "on", "audit", "subset"}


def env_int(name: str, default: int) -> int:
    try:
        return int(str(os.environ.get(name, "")).strip() or default)
    except ValueError:
        return default


def safe_stem(value: str) -> str:
    stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value or "model")).strip("._")
    return stem[:80] or "model"

SOFFICE_ENV_VARS = ("SUPPLY_CHAIN_SOFFICE", "LIBREOFFICE_PATH", "SOFFICE")
SOFFICE_CANDIDATES = (
    r"C:\Program Files\LibreOffice\program\soffice.com",
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.com",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    "/usr/bin/soffice",
    "/usr/local/bin/soffice",
    "/opt/libreoffice/program/soffice",
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
)


DEFAULT_SCORE_CONFIG: dict[str, dict[str, Any]] = {
    "采购": {
        "plan_days": 30,
        "targets": {"production_satisfaction": 1.0},
        "points": {"单位采购成本": 60.0, "生产满足率": 40.0},
    },
    "生产": {
        "plan_days": 30,
        "targets": {"unit_logistics_cost": 0.0, "market_satisfaction": 1.0},
        "points": {"单位物流成本": 60.0, "市场满足率": 40.0},
    },
    "销售": {
        "plan_days": 30,
        "targets": {"prediction_deviation": 0.05, "unit_logistics_cost": 50.0, "market_satisfaction": 1.0},
        "points": {"预测偏差率": 20.0, "单位物流成本": 50.0, "市场满足率": 30.0},
    },
    "综合": {
        "plan_days": 30,
        "targets": {
            "prediction_deviation": 0.05,
            "unit_logistics_cost": 110.0,
            "unit_procurement_cost": 1600.0,
            "production_satisfaction": 1.0,
            "market_satisfaction": 1.0,
        },
        "points": {"预测偏差率": 10.0, "单位物流成本": 20.0, "单位采购成本": 20.0, "生产满足率": 15.0, "市场满足率": 35.0},
    },
}


CASE_SCORE_CONFIG: dict[str, dict[str, Any]] = {
    "白砂糖": {"plan_days": 30, "targets": {"unit_procurement_cost": 415.0}},
    "牙膏": {"plan_days": 30, "targets": {"unit_procurement_cost": 2370.0}},
    "硫磺": {"plan_days": 90, "targets": {"unit_procurement_cost": 630.0}},
    "电脑": {"plan_days": 30, "targets": {"unit_logistics_cost": 190.0}},
    "毛衣": {"plan_days": 30, "targets": {"unit_logistics_cost": 110.0}},
    "汉堡": {"plan_days": 30, "targets": {"unit_logistics_cost": 40.0}},
    "节能灯": {"plan_days": 30, "targets": {"unit_logistics_cost": 50.0}},
    "热水器": {"plan_days": 60, "targets": {"unit_logistics_cost": 130.0}},
    "速冻水饺": {"plan_days": 30, "targets": {"unit_logistics_cost": 25.0}},
    "电视": {
        "plan_days": 30,
        "forecast_bias": False,
        "targets": {"prediction_deviation": 0.05, "unit_logistics_cost": 100.0, "unit_procurement_cost": 1550.0, "production_satisfaction": 1.0, "market_satisfaction": 1.0},
        "points": {"预测偏差率": 10.0, "单位物流成本": 25.0, "单位采购成本": 25.0, "生产满足率": 15.0, "市场满足率": 25.0},
    },
    "羽绒服": {
        "plan_days": 60,
        "targets": {"prediction_deviation": 0.05, "unit_logistics_cost": 25.0, "unit_procurement_cost": 125.0, "production_satisfaction": 1.0, "market_satisfaction": 1.0},
        "points": {"预测偏差率": 10.0, "单位物流成本": 20.0, "单位采购成本": 20.0, "生产满足率": 15.0, "市场满足率": 35.0},
    },
    "蓄电池": {
        "plan_days": 60,
        "targets": {"prediction_deviation": 0.05, "unit_logistics_cost": 70.0, "unit_procurement_cost": 310.0, "production_satisfaction": 1.0, "market_satisfaction": 1.0},
        "points": {"预测偏差率": 10.0, "单位物流成本": 20.0, "单位采购成本": 20.0, "生产满足率": 20.0, "市场满足率": 30.0},
    },
}


VERIFIED_SCORE_OVERRIDES: dict[str, dict[str, Any]] = {
    "硫磺国际采购★★☆-标准版供应链设计规划竞赛1": {
        "qtype": "采购",
        "score": 98.75,
        "unit_procurement": 632.62,
        "production_satisfaction": 1.0,
        "note": "平台回执校准：实际单位成本 632.62，最终得分 98.75",
    },
}


VERIFIED_SEGMENT_DISTANCES: dict[str, float] = {
    # 人工从平台路线弹窗确认的单段距离。用于按本趟数量选择承运商；
    # 没有证据的单段不硬填距离，继续使用 xls 里的整条路线成本口径。
    "苏州工厂-->苏州火车站": 22.0,
    "武汉火车站-->武汉总代": 38.0,
}


VERIFIED_SEGMENT_DURATIONS: dict[str, int] = {
    # 人工确认的单段运输时间。只对这些路线自动补中转段起运日；
    # 未确认的中转段不从图像识别或通用规则推断日期。
    "沙特石油-->达曼港码头": 1,
    "达曼港码头-->防城港码头": 22,
    "防城港码头-->防城火车站": 1,
    "防城火车站-->贵阳火车站": 2,
    "贵阳火车站-->贵州瓮福": 1,
    "防城港码头-->贵州瓮福": 1,
    "江西铜业-->贵溪火车站": 1,
    "贵溪火车站-->贵阳火车站": 4,

    # 白砂糖采购，来自“物流实训/1-白砂糖-96.36分.docx”运输计划截图。
    "东亚糖业公司-->南宁火车站": 1,
    "南宁糖业公司-->南宁火车站": 1,
    "南宁火车站-->广州火车站": 2,
    "南宁火车站-->惠州火车站": 3,
    "南宁火车站-->湛江火车站": 1,
    "广州火车站-->广州可乐工厂": 1,
    "惠州火车站-->惠州可乐工厂": 1,
    "湛江火车站-->湛江可乐工厂": 1,

    # 牙膏甘油跨境采购，来自“物流实训/3-牙膏-94分.docx”运输计划截图。
    "马来甘油厂-->吉隆坡港": 1,
    "吉隆坡港-->盐田港": 10,
    "盐田港-->西安生产基地": 3,

    # 节能灯销售，来自“物流实训/7-节能灯-98.docx”运输计划截图。
    "中山欧普公司-->中山火车站": 1,
    "中山火车站-->武汉火车站": 4,
    "武汉火车站-->天门门店": 1,
    "武汉火车站-->武汉门店": 1,

    # 电脑生产，来自“物流实训/4-电脑-100.docx”运输计划截图。
    "厦门工厂-->厦门火车站": 1,
    "厦门火车站-->武汉火车站": 3,
    "厦门火车站-->石家庄火车站": 6,
    "武汉火车站-->天门代理": 1,
    "武汉火车站-->黄冈代理": 1,
    "石家庄火车站-->保定代理": 1,
    "石家庄火车站-->衡水代理": 1,

    # 电视综合，来自“物流实训/10-电视机-100.docx”运输计划截图。
    "广东芯片厂-->广州火车站": 1,
    "武汉液晶面板厂-->武汉火车站": 1,
    "济南液晶面板厂-->济南火车站": 1,
    "福建机壳厂-->福州火车站": 1,
    "西安外壳厂-->西安火车站": 1,
    "郑州芯片厂-->郑州火车站": 1,
    "广州火车站-->青岛火车站": 7,
    "武汉火车站-->青岛火车站": 4,
    "济南火车站-->青岛火车站": 2,
    "福州火车站-->青岛火车站": 5,
    "西安火车站-->青岛火车站": 5,
    "郑州火车站-->青岛火车站": 3,
    "青岛火车站-->电视工厂": 1,
    "电视工厂-->青岛火车站": 1,
    "青岛火车站-->北京火车站": 2,
    "青岛火车站-->南京火车站": 2,
    "北京火车站-->北京总经销": 1,
    "南京火车站-->南京总经销": 1,

    # 毛衣生产，来自“物流实训/6-毛衣.docx”运输计划截图。
    "东莞工厂-->东莞火车站": 1,
    "东莞火车站-->广州火车站": 1,
    "泉州工厂-->泉州火车站": 1,
    "泉州火车站-->广州火车站": 2,
    "苏州工厂-->苏州火车站": 1,
    "苏州火车站-->广州火车站": 4,
    "广州火车站-->毛织品外贸公司": 1,

    # 热水器销售，来自“物流实训/9-热水器-100.docx”运输计划截图。
    "东莞工厂-->惠州火车站": 1,
    "东莞工厂-->武汉总代": 3,
    "苏州工厂-->洋山码头": 1,
    "苏州工厂-->苏州火车站": 1,
    "苏州工厂-->北方总代": 3,
    "洋山码头-->横滨码头": 7,
    "横滨码头-->日本总代": 1,
    "武汉火车站-->武汉总代": 1,
    "石家庄火车站-->北方总代": 1,

    # 羽绒服跨境运营，来自“物流实训/11-羽绒服-100.docx”运输计划截图。
    "宁波工厂-->北方销售中心": 3,
    "宁波工厂-->栎社机场": 1,
    "宁波工厂-->北仑码头": 1,
    "宁波工厂-->宁波火车站": 1,
    "北京火车站-->北方销售中心": 1,
    "南京火车站-->南方销售中心": 1,
    "东京机场-->日本销售中心": 1,
    "东京港-->日本销售中心": 1,
    "宁波火车站-->南京火车站": 2,
    "宁波火车站-->北京火车站": 5,
    "北仑码头-->东京港": 9,
    "栎社机场-->东京机场": 1,
    "濮阳羽绒供应商-->郑州火车站": 1,
    "六安羽绒供应商-->合肥火车站": 1,
    "郑州火车站-->宁波火车站": 4,
    "合肥火车站-->宁波火车站": 2,
    "宁波火车站-->宁波工厂": 1,
    "金华拉链供应商-->宁波工厂": 1,
    "苏州面料供应商-->宁波工厂": 1,
    "广州面料供应商-->广州火车站": 1,
    "广州火车站-->宁波火车站": 5,

    # 汽车蓄电池综合，来自“物流实训/12-汽车蓄电池.docx”运输计划截图。
    "四川化工厂-->达州火车站": 1,
    "四川化工厂-->广州蓄电池厂": 2,
    "四川化工厂-->苏州蓄电池厂": 3,
    "贵阳化工厂-->广州蓄电池厂": 2,
    "贵阳化工厂-->苏州蓄电池厂": 4,
    "广州火车站-->广州蓄电池厂": 1,
    "苏州火车站-->苏州蓄电池厂": 1,
    "达州火车站-->苏州火车站": 5,
    "达州火车站-->广州火车站": 4,
    "常州塑料厂-->苏州蓄电池厂": 1,
    "东莞塑料厂-->广州蓄电池厂": 1,
    "广州蓄电池厂-->广州火车站": 1,
    "苏州蓄电池厂-->洋山港": 1,
    "苏州蓄电池厂-->华北总代": 2,
    "苏州蓄电池厂-->华中总代": 2,
    "苏州蓄电池厂-->苏州火车站": 1,
    "广州火车站-->西安火车站": 6,
    "广州火车站-->武汉火车站": 4,
    "武汉火车站-->华中总代": 1,
    "石家庄火车站-->华北总代": 1,
    "苏州火车站-->西安火车站": 5,
    "洋山港-->横滨码头": 5,
    "西安火车站-->西北总代": 1,
    "石家庄极板厂-->苏州蓄电池厂": 2,
    "石家庄极板厂-->石家庄火车站": 1,
    "宝鸡极板厂-->广州蓄电池厂": 3,
    "宝鸡极板厂-->西安火车站": 1,
    "石家庄火车站-->广州火车站": 6,
    "石家庄火车站-->苏州火车站": 4,
    "西安火车站-->苏州火车站": 4,
}


VERIFIED_SEGMENT_DURATIONS_BY_CARRIER: dict[tuple[str, str], int] = {
    # 同一铁路单段在不同承运商下可能运输时间不同，优先按承运商取值。
    ("武汉火车站-->青岛火车站", "西铁货运"): 4,
    ("武汉火车站-->青岛火车站", "中原铁路"): 3,
    ("福州火车站-->青岛火车站", "西铁货运"): 5,
    ("福州火车站-->青岛火车站", "中原铁路"): 4,
    ("西安火车站-->青岛火车站", "西铁货运"): 5,
    ("西安火车站-->青岛火车站", "中原铁路"): 4,
    ("惠州火车站-->武汉火车站", "南方铁路"): 3,
    ("惠州火车站-->武汉火车站", "中铁快运"): 4,
    ("苏州火车站-->石家庄火车站", "南方铁路"): 4,
    ("苏州火车站-->石家庄火车站", "中铁快运"): 5,
    ("苏州火车站-->石家庄火车站", "中原铁路"): 4,
    ("东莞拉链供应商-->宁波工厂", "易达快运"): 3,
    ("东莞拉链供应商-->宁波工厂", "开源陆运"): 2,
    ("苏州火车站-->武汉火车站", "南方铁运"): 2,
    ("苏州火车站-->武汉火车站", "中原铁路"): 3,
    ("西安火车站-->广州火车站", "南方铁运"): 5,
    ("西安火车站-->广州火车站", "中原铁路"): 6,
}


def verified_segment_duration(segment: str, carrier: str) -> int | None:
    context = ACTIVE_HAR_CONTEXT or {}
    by_carrier = context.get("segment_days_by_carrier") or {}
    if isinstance(by_carrier, dict):
        direct_key = f"{segment}|||{carrier}"
        if direct_key in by_carrier:
            return int(by_carrier[direct_key])
        for key, duration in by_carrier.items():
            if not isinstance(key, str) or "|||" not in key:
                continue
            known_segment, known_carrier = key.split("|||", 1)
            if known_segment == segment and known_carrier and known_carrier in str(carrier):
                return int(duration)
    carrier_text = str(carrier)
    for (known_segment, known_carrier), duration in VERIFIED_SEGMENT_DURATIONS_BY_CARRIER.items():
        if known_segment == segment and known_carrier in carrier_text:
            return duration
    return VERIFIED_SEGMENT_DURATIONS.get(segment)


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


def log(msg: str) -> None:
    print(f"[solve] {msg}")


def sv(value: Any) -> str:
    return "" if value is None else str(value).strip()


def nv(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    if cleaned in ("", ".", "-", "-."):
        return default
    try:
        return float(cleaned)
    except ValueError:
        return default


def ceil_int(value: float) -> int:
    return int(math.ceil(max(0.0, value)))


def money(value: float) -> str:
    return f"{value:,.2f}"


def qty(value: float) -> str:
    return f"{ceil_int(value):,}"


def find_soffice() -> Path | None:
    """Find a LibreOffice executable without assuming a Windows-only path."""
    for var_name in SOFFICE_ENV_VARS:
        raw = os.environ.get(var_name)
        if raw:
            candidate = Path(raw).expanduser()
            if candidate.exists():
                return candidate

    for executable in ("soffice.com", "soffice", "libreoffice"):
        found = shutil.which(executable)
        if found:
            return Path(found)

    for raw in SOFFICE_CANDIDATES:
        candidate = Path(raw).expanduser()
        if candidate.exists():
            return candidate

    return None


def check_environment() -> bool:
    required_modules = ("numpy", "openpyxl", "pandas")
    optional_modules = ("pulp",)
    ok = True
    print("Python:", sys.version.split()[0])
    for module_name in required_modules:
        try:
            __import__(module_name)
            print(f"[OK] Python package: {module_name}")
        except Exception as exc:
            ok = False
            print(f"[MISSING] Python package: {module_name} ({exc})")
    for module_name in optional_modules:
        try:
            __import__(module_name)
            print(f"[OK] Python package: {module_name}")
        except Exception as exc:
            ok = False
            print(f"[MISSING] Python package: {module_name} ({exc})")

    soffice = find_soffice()
    if soffice:
        print(f"[OK] LibreOffice: {soffice}")
    else:
        ok = False
        env_list = ", ".join(SOFFICE_ENV_VARS)
        print(f"[MISSING] LibreOffice soffice executable. Set one of: {env_list}")
    return ok


def convert_xls_to_xlsx(xls_path: Path) -> Path:
    """把中文/特殊文件名 .xls 先复制为 ASCII 临时名，再交给 LibreOffice 转换。"""
    xls_path = xls_path.resolve()
    if not xls_path.exists():
        raise FileNotFoundError(f"文件不存在: {xls_path}")
    soffice = find_soffice()
    if not soffice:
        env_list = ", ".join(SOFFICE_ENV_VARS)
        raise FileNotFoundError(f"找不到 LibreOffice soffice，请安装 LibreOffice 或设置环境变量: {env_list}")

    tmp_dir = Path(tempfile.mkdtemp(prefix="sc_xls_"))
    ascii_xls = tmp_dir / "input.xls"
    shutil.copy2(xls_path, ascii_xls)

    profile_dir = tmp_dir / "lo_profile"
    profile_dir.mkdir(parents=True, exist_ok=True)
    last_error: subprocess.CalledProcessError | None = None
    for _ in range(3):
        try:
            subprocess.run(
                [
                    str(soffice),
                    "--headless",
                    "--norestore",
                    "--nodefault",
                    "--nolockcheck",
                    "--nofirststartwizard",
                    f"-env:UserInstallation={profile_dir.as_uri()}",
                    "--convert-to",
                    "xlsx",
                    "--outdir",
                    str(tmp_dir),
                    str(ascii_xls),
                ],
                check=True,
                capture_output=True,
                timeout=90,
            )
            break
        except subprocess.CalledProcessError as exc:
            last_error = exc
    else:
        assert last_error is not None
        raise last_error

    xlsx_path = tmp_dir / "input.xlsx"
    if not xlsx_path.exists():
        matches = list(tmp_dir.glob("*.xlsx"))
        if not matches:
            raise RuntimeError(f"转换失败: {tmp_dir} 中没有 .xlsx")
        xlsx_path = matches[0]
    return xlsx_path


def read_workbook(xls_path: Path) -> dict[str, list[tuple[Any, ...]]]:
    xlsx_path = convert_xls_to_xlsx(xls_path)
    wb = load_workbook(xlsx_path, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        raw_rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
        shutil.rmtree(xlsx_path.parent, ignore_errors=True)

    title_keywords = [
        "产品及原料清单",
        "货币汇率表",
        "销售网点历史销售数据",
        "销售网点销售数据",
        "销售网点历史数据",
        "销售网点",
        "供应商产能数据",
        "工厂原料消耗及期初库存",
        "工厂产能及期初库存",
        "工厂产能",
        "承运商信息",
        "地点信息",
        "运输路线",
    ]

    title_idx: dict[str, int] = {}
    for idx, row in enumerate(raw_rows):
        row_text = " ".join(sv(cell) for cell in row if sv(cell))
        for keyword in title_keywords:
            if keyword in row_text and keyword not in title_idx:
                title_idx[keyword] = idx

    sections: dict[str, list[tuple[Any, ...]]] = {}
    ordered = sorted(title_idx.items(), key=lambda item: item[1])
    for pos, (title, start) in enumerate(ordered):
        end = ordered[pos + 1][1] if pos + 1 < len(ordered) else len(raw_rows)
        rows = [
            raw_rows[i]
            for i in range(start + 1, end)
            if any(sv(cell) for cell in raw_rows[i])
        ]
        if rows:
            sections[title] = rows
    return sections


def find_section(sections: dict[str, list[tuple[Any, ...]]], *keywords: str) -> list[tuple[Any, ...]]:
    for title, rows in sections.items():
        if all(keyword in title for keyword in keywords):
            return rows
    for title, rows in sections.items():
        if any(keyword in title for keyword in keywords):
            return rows
    return []


@dataclass
class Product:
    kind: str
    name: str
    unit: str
    bom: float
    charge_ratio: float


@dataclass
class Factory:
    name: str
    product: str
    unit: str
    init: float
    daily: float
    limit: float
    excess_fee: float


@dataclass
class FactoryMaterial:
    factory: str
    material: str
    unit: str
    init: float
    daily: float
    limit: float
    excess_fee: float


@dataclass
class Supplier:
    name: str
    material: str
    unit: str
    init: float
    daily: float
    available: float
    price: float
    currency: str


@dataclass
class SalesNode:
    node: str
    product: str
    unit: str
    init: float
    monthly: list[float]
    daily: list[float]
    limit: float
    excess_fee: float


@dataclass
class Route:
    route: str
    src: str
    dst: str
    distance: float
    rate: float
    min_qty: float
    min_freight: float
    lead: int
    currency: str
    segment_distances: tuple[float | None, ...] = ()


@dataclass(frozen=True)
class TransportOption:
    route: Route
    carrier: str
    segment_costs: tuple[tuple[float, float], ...]
    lead: int


@dataclass(frozen=True)
class LaneOption:
    lane_id: int
    route: Route
    option: TransportOption


@dataclass(frozen=True)
class HarCarrier:
    name: str
    mode: str
    efficiency_km_per_day: float
    unit_rate_cny: float
    start_fee_cny: float


def parse_products(sections: dict[str, list[tuple[Any, ...]]]) -> list[Product]:
    products: list[Product] = []
    last_kind = ""
    for row in find_section(sections, "产品及原料清单"):
        kind = sv(row[1] if len(row) > 1 else "")
        name = sv(row[2] if len(row) > 2 else "")
        if kind in ("产品", "原料"):
            last_kind = kind
        elif not kind and last_kind == "原料" and name:
            kind = "原料"
        if kind not in ("产品", "原料") or not name or name == "名称":
            continue
        products.append(
            Product(
                kind=kind,
                name=name,
                unit=sv(row[3] if len(row) > 3 else ""),
                bom=nv(row[4] if len(row) > 4 else 1.0, 1.0),
                charge_ratio=nv(row[5] if len(row) > 5 else 1.0, 1.0) or 1.0,
            )
        )
    return products


def parse_factories(sections: dict[str, list[tuple[Any, ...]]]) -> list[Factory]:
    factories: list[Factory] = []
    rows = find_section(sections, "工厂产能及期初库存") or find_section(sections, "工厂产能")
    for row in rows:
        name = sv(row[1] if len(row) > 1 else "")
        product = sv(row[2] if len(row) > 2 else "")
        daily = nv(row[5] if len(row) > 5 else 0)
        if not name or name in ("工厂", "货物", "仓库信息", "仓库名称"):
            continue
        if "仓库" in name and daily <= 0:
            continue
        if product in ("", "货物", "产品名称") and daily <= 0:
            continue
        factories.append(
            Factory(
                name=name,
                product=product,
                unit=sv(row[3] if len(row) > 3 else ""),
                init=nv(row[4] if len(row) > 4 else 0),
                daily=daily,
                limit=nv(row[6] if len(row) > 6 else 0, 999999),
                excess_fee=nv(row[7] if len(row) > 7 else 0),
            )
        )
    return factories


def parse_factory_materials(sections: dict[str, list[tuple[Any, ...]]]) -> list[FactoryMaterial]:
    materials: list[FactoryMaterial] = []
    for row in find_section(sections, "工厂原料消耗及期初库存"):
        factory = sv(row[1] if len(row) > 1 else "")
        material = sv(row[2] if len(row) > 2 else "")
        if not factory or not material or factory == "工厂":
            continue
        materials.append(
            FactoryMaterial(
                factory=factory,
                material=material,
                unit=sv(row[3] if len(row) > 3 else ""),
                init=nv(row[4] if len(row) > 4 else 0),
                daily=nv(row[5] if len(row) > 5 else 0),
                limit=nv(row[6] if len(row) > 6 else 0, 999999),
                excess_fee=nv(row[7] if len(row) > 7 else 0),
            )
        )
    return materials


def parse_suppliers(sections: dict[str, list[tuple[Any, ...]]]) -> list[Supplier]:
    suppliers: list[Supplier] = []
    for row in find_section(sections, "供应商产能"):
        name = sv(row[1] if len(row) > 1 else "")
        if not name or name == "供应商":
            continue
        suppliers.append(
            Supplier(
                name=name,
                material=sv(row[2] if len(row) > 2 else ""),
                unit=sv(row[3] if len(row) > 3 else ""),
                init=nv(row[4] if len(row) > 4 else 0),
                daily=nv(row[5] if len(row) > 5 else 0),
                available=nv(row[8] if len(row) > 8 else 0),
                price=nv(row[9] if len(row) > 9 else 0),
                currency=sv(row[10] if len(row) > 10 else "CNY") or "CNY",
            )
        )
    return suppliers


def parse_rates(sections: dict[str, list[tuple[Any, ...]]]) -> dict[tuple[str, str], float]:
    rates: dict[tuple[str, str], float] = {("CNY", "CNY"): 1.0}
    for row in find_section(sections, "货币汇率表"):
        src = sv(row[1] if len(row) > 1 else "")
        dst = sv(row[2] if len(row) > 2 else "")
        rate = nv(row[3] if len(row) > 3 else 0)
        if src and dst and rate:
            rates[(src, dst)] = rate
    return rates


def _parse_js_number(text: str, default: float = 0.0) -> float:
    raw = str(text or "").strip()
    if raw.startswith("."):
        raw = "0" + raw
    try:
        return float(raw)
    except ValueError:
        return default


def set_har_context(context: dict[str, Any] | None) -> None:
    """Switch platform context and drop route-option cache tied to old data."""
    global ACTIVE_HAR_CONTEXT
    ACTIVE_HAR_CONTEXT = context
    ROUTE_TRANSPORT_OPTIONS_CACHE.clear()


def merge_har_contexts(contexts: list[dict[str, Any] | None]) -> dict[str, Any] | None:
    clean_contexts = [context for context in contexts if context]
    if not clean_contexts:
        return None
    merged: dict[str, Any] = {
        "path": ";".join(sv(context.get("path")) for context in clean_contexts if context.get("path")),
        "carriers": [],
        "segment_distances": {},
        "segment_days_by_carrier": {},
        "business_entries": [],
        "source_kind": "multi" if len(clean_contexts) > 1 else clean_contexts[0].get("source_kind", ""),
    }
    carriers_by_name: dict[str, dict[str, Any]] = {}
    for context in clean_contexts:
        for row in context.get("carriers") or []:
            if not isinstance(row, dict):
                continue
            name = sv(row.get("name"))
            if name:
                carriers_by_name[name] = dict(row)
        merged["segment_distances"].update(context.get("segment_distances") or {})
        merged["segment_days_by_carrier"].update(context.get("segment_days_by_carrier") or {})
        merged["business_entries"].extend(context.get("business_entries") or [])
    merged["carriers"] = list(carriers_by_name.values())
    return merged


def load_frontend_contexts(paths: list[str | Path] | None) -> dict[str, Any] | None:
    if not paths:
        return None
    return merge_har_contexts([load_har_context(Path(path)) for path in paths])


def _collect_carry_plans(payload: Any, plans: list[dict[str, Any]]) -> None:
    if isinstance(payload, dict):
        raw_plans = payload.get("carryPlans")
        if isinstance(raw_plans, list):
            plans.extend(plan for plan in raw_plans if isinstance(plan, dict))
        for value in payload.values():
            if isinstance(value, (dict, list)):
                _collect_carry_plans(value, plans)
    elif isinstance(payload, list):
        for item in payload:
            _collect_carry_plans(item, plans)


def _json_arrays_after_key(text: str, key: str) -> list[Any]:
    decoder = json.JSONDecoder()
    arrays: list[Any] = []
    for match in re.finditer(rf'"{re.escape(key)}"\s*:\s*', text):
        idx = match.end()
        while idx < len(text) and text[idx].isspace():
            idx += 1
        if idx >= len(text) or text[idx] != "[":
            continue
        try:
            value, _end = decoder.raw_decode(text[idx:])
        except json.JSONDecodeError:
            continue
        arrays.append(value)
    return arrays


def load_har_context(har_path: Path | None) -> dict[str, Any] | None:
    if not har_path:
        return None
    path = Path(har_path)
    if not path.exists():
        raise FileNotFoundError(f"前端数据文件不存在: {path}")
    raw_file_text = path.read_text(encoding="utf-8-sig")
    try:
        data = json.loads(raw_file_text)
    except json.JSONDecodeError:
        data = {
            "log": {
                "entries": [
                    {
                        "request": {"url": str(path)},
                        "response": {"status": 200, "content": {"mimeType": "text/plain", "text": raw_file_text}},
                    }
                ]
            }
        }
    entries = data.get("log", {}).get("entries", []) if isinstance(data, dict) else []
    texts: list[str] = []
    business_entries: list[dict[str, Any]] = []
    for entry in entries:
        request = entry.get("request", {})
        response = entry.get("response", {})
        url = request.get("url", "")
        content = response.get("content", {})
        text = content.get("text", "") or ""
        if text:
            texts.append(text)
        if "/execute/" in url or "/view/execute/" in url:
            business_entries.append(
                {
                    "url": url,
                    "status": response.get("status"),
                    "mime": content.get("mimeType", ""),
                    "length": len(text),
                    "postData": (request.get("postData") or {}).get("text", ""),
                }
            )
    carriers: list[dict[str, Any]] = []
    segment_distances: dict[str, float] = {}
    segment_days_by_carrier: dict[tuple[str, str], int] = {}
    joined = "\n".join(texts)
    all_carry_plans: list[dict[str, Any]] = []
    for text in texts:
        raw = str(text or "").strip()
        if not raw or raw[0] not in "[{":
            continue
        try:
            payload = json.loads(raw)
        except Exception:
            continue
        _collect_carry_plans(payload, all_carry_plans)
    for value in _json_arrays_after_key(joined, "carryPlans"):
        if isinstance(value, list):
            all_carry_plans.extend(plan for plan in value if isinstance(plan, dict))
    for plan in all_carry_plans:
        src = sv(plan.get("polName"))
        dst = sv(plan.get("podName"))
        if not src or not dst:
            continue
        segment = f"{src}-->{dst}"
        distance = nv(plan.get("carryDistance"), 0.0)
        if distance > 0:
            segment_distances[segment] = float(distance)
        carrier = sv(plan.get("podCarrierName"))
        days = int(nv(plan.get("carryDays"), 0.0) or 0)
        if carrier and days > 0:
            segment_days_by_carrier[(segment, carrier)] = days
    for body in re.findall(r"DATA_CARRIERS\.push\(\{(.*?)\}\);", joined, flags=re.S):
        def str_field(name: str) -> str:
            match = re.search(rf"{name}\s*:\s*'([^']*)'", body)
            return match.group(1).strip() if match else ""

        def num_field(name: str) -> float:
            match = re.search(rf"{name}\s*:\s*([+-]?(?:\d+(?:\.\d*)?|\.\d+))", body)
            return _parse_js_number(match.group(1)) if match else 0.0

        name = str_field("carrierName")
        if not name:
            continue
        carriers.append(
            {
                "name": name,
                "mode": str_field("carrierType"),
                "efficiency": num_field("efficiency"),
                "unit_price": num_field("unitPrice"),
                "lowest_charge": num_field("lowestCharge"),
                "currency": str_field("currencyCode") or "CNY",
                "country": str_field("countryCode"),
            }
        )
    return {
        "path": str(path),
        "carriers": carriers,
        "segment_distances": segment_distances,
        "segment_days_by_carrier": {f"{segment}|||{carrier}": days for (segment, carrier), days in segment_days_by_carrier.items()},
        "business_entries": business_entries,
        "source_kind": "har" if path.suffix.lower() == ".har" else "frontend",
    }


def har_carriers_for_rates(rates: dict[tuple[str, str], float]) -> list[HarCarrier] | None:
    context = ACTIVE_HAR_CONTEXT
    if not context:
        return None
    rows = context.get("carriers") or []
    carriers: list[HarCarrier] = []
    for row in rows:
        mode_code = sv(row.get("mode")).upper()
        mode = {"P": "汽运", "R": "铁路", "O": "海运", "A": "空运"}.get(mode_code, mode_code)
        currency = sv(row.get("currency")) or "CNY"
        carriers.append(
            HarCarrier(
                name=sv(row.get("name")),
                mode=mode,
                efficiency_km_per_day=float(row.get("efficiency") or 0.0),
                unit_rate_cny=currency_to_cny(float(row.get("unit_price") or 0.0), currency, rates),
                start_fee_cny=currency_to_cny(float(row.get("lowest_charge") or 0.0), currency, rates),
            )
        )
    return carriers or None


def parse_sales(sections: dict[str, list[tuple[Any, ...]]], days: int = PLAN_DAYS) -> list[SalesNode]:
    rows = (
        find_section(sections, "销售网点历史销售数据")
        or find_section(sections, "销售网点销售数据")
        or find_section(sections, "销售网点历史数据")
        or find_section(sections, "销售网点")
    )
    sales: list[SalesNode] = []
    for row in rows:
        node = sv(row[1] if len(row) > 1 else "")
        if not node or node == "销售网点":
            continue
        init = nv(row[4] if len(row) > 4 else 0)
        values = [nv(row[i] if len(row) > i else 0) for i in range(5, len(row))]
        non_zero = [value for value in values if value > 0]
        monthly: list[float] = []
        daily: list[float] = []
        if len(non_zero) >= min(20, days) and len(row) > 5 + days:
            daily = values[:days]
            limit = nv(row[5 + days] if len(row) > 5 + days else 0, 999999)
            excess = nv(row[6 + days] if len(row) > 6 + days else 0)
        else:
            monthly = values[:6]
            limit = nv(row[11] if len(row) > 11 else 0, 999999)
            excess = nv(row[12] if len(row) > 12 else 0)
        sales.append(
            SalesNode(
                node=node,
                product=sv(row[2] if len(row) > 2 else ""),
                unit=sv(row[3] if len(row) > 3 else ""),
                init=init,
                monthly=monthly,
                daily=daily,
                limit=limit,
                excess_fee=excess,
            )
        )
    return sales


def parse_routes(sections: dict[str, list[tuple[Any, ...]]]) -> list[Route]:
    routes: list[Route] = []
    for row in find_section(sections, "运输路线"):
        route_name = sv(row[1] if len(row) > 1 else "")
        if not route_name or route_name == "运输路线" or "-->" not in route_name:
            continue
        points = [part.strip() for part in route_name.split("-->") if part.strip()]
        routes.append(
            Route(
                route=route_name,
                src=points[0],
                dst=points[-1],
                distance=nv(row[6] if len(row) > 6 else 0),
                rate=nv(row[7] if len(row) > 7 else 0),
                min_qty=nv(row[8] if len(row) > 8 else 0),
                min_freight=nv(row[9] if len(row) > 9 else 0),
                lead=ceil_int(nv(row[10] if len(row) > 10 else 1, 1)),
                currency=sv(row[11] if len(row) > 11 else "CNY") or "CNY",
            )
        )
    single_segment_distances = {
        route.route: float(route.distance)
        for route in routes
        if len(route_segments(route.route)) == 1 and route.distance > 0
    }
    for route in routes:
        route.segment_distances = tuple(single_segment_distances.get(segment) for _src, _dst, segment in route_segments(route.route))
    return routes


def detect_type(sections: dict[str, list[tuple[Any, ...]]], xls_path: Path | None = None) -> str:
    titles = list(sections)
    has_sales = any("销售网点" in title for title in titles)
    has_supplier = any("供应商产能" in title for title in titles)
    has_factory = any("工厂产能" in title for title in titles)
    has_factory_material = any("工厂原料" in title for title in titles)
    file_hint = xls_path.name if xls_path else ""

    if has_sales and has_supplier and has_factory and has_factory_material:
        return "综合"
    if "销售" in file_hint and has_sales:
        return "销售"
    if "生产" in file_hint and has_sales and has_factory:
        return "生产"
    if has_sales and has_factory and not has_supplier and not has_factory_material:
        return "销售"
    if has_sales and has_factory:
        return "生产"
    if has_supplier or has_factory_material:
        return "采购"
    return "采购"


def moving_weighted(monthly: list[float], use_bias: bool = True) -> float:
    values = [value for value in monthly if value > 0]
    if len(values) >= 6:
        values = values[-6:]
        weights = [1, 2, 3, 4, 5, 6]
        weighted = sum(value * weight for value, weight in zip(values, weights)) / sum(weights)
        if not use_bias:
            return weighted
        first3 = sum(values[:3])
        last3 = sum(values[3:6])
        bias = (last3 - first3) / 9
        if abs(bias) > 150:
            bias = (last3 - first3) / 12
        return weighted + bias
    return sum(values) / len(values) if values else 0.0


def spread_integer(total: float, days: int = PLAN_DAYS) -> list[int]:
    total_int = ceil_int(total)
    base = total_int // days
    remainder = total_int - base * days
    return [base + (1 if idx < remainder else 0) for idx in range(days)]


def production_daily_schedule(
    production_rows: list[dict[str, Any]],
    factories: list[Factory],
    days: int = PLAN_DAYS,
) -> dict[str, list[int]]:
    """Expand fill-table production rows into daily finished-goods output."""
    factory_by_name = {factory.name: factory for factory in factories}
    schedule = {factory.name: [0] * days for factory in factories}
    for row in production_rows:
        factory_name = sv(row.get("factory"))
        if factory_name not in schedule:
            continue
        explicit_daily = row.get("daily_schedule")
        if explicit_daily:
            for day_idx, value in enumerate(list(explicit_daily)[:days]):
                amount = int(round(float(value or 0.0)))
                if amount > 0:
                    schedule[factory_name][day_idx] += amount
            continue
        amount = ceil_int(row.get("amount", 0))
        if amount <= 0:
            continue
        factory = factory_by_name.get(factory_name)
        daily_capacity = int(math.floor(float(factory.daily or 0.0))) if factory else 0
        if daily_capacity <= 0:
            capacity = float(row.get("capacity") or 0.0)
            daily_capacity = int(math.floor(capacity / max(days, 1))) if capacity > 0 else 0
        if daily_capacity <= 0:
            daily_capacity = max(1, ceil_int(amount / max(days, 1)))

        remaining = amount
        for day_idx in range(days):
            if remaining <= 0:
                break
            produced = min(daily_capacity, remaining)
            schedule[factory_name][day_idx] += int(produced)
            remaining -= int(produced)
        if remaining > 0:
            schedule[factory_name][-1] += int(remaining)
    return schedule


def production_amounts_from_transport(
    factories: list[Factory],
    product_transport: list[dict[str, Any]],
) -> dict[str, int]:
    shipped_by_factory: dict[str, int] = defaultdict(int)
    for row in product_transport:
        factory_name = sv(row.get("factory") or row.get("source"))
        if factory_name:
            shipped_by_factory[factory_name] += int(row.get("amount") or 0)
    return {
        factory.name: max(0, shipped_by_factory.get(factory.name, 0) - int(factory.init or 0))
        for factory in factories
    }


def align_production_to_transport(
    *,
    forecasts: list[dict[str, Any]],
    product_transport: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows = [dict(row) for row in product_transport]
    production_rows: list[dict[str, Any]] = []
    last_signature: tuple[Any, ...] | None = None
    for _ in range(4):
        amounts = production_amounts_from_transport(factories, rows)
        production_rows = production_rows_from_amounts(factories, cargo, amounts, days)
        rows = effective_shipments(rows, days)
        signature = tuple(
            (sv(row.get("factory") or row.get("source")), sv(row.get("destination")), sv(row.get("route")), int(row.get("ship_day") or 1), int(row.get("amount") or 0))
            for row in rows
        )
        production_signature = tuple((row["factory"], int(row.get("amount") or 0)) for row in production_rows)
        combined_signature = (signature, production_signature)
        if combined_signature == last_signature:
            break
        last_signature = combined_signature
    return production_rows, rows


def factory_output_schedule(
    factories: list[Factory],
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
) -> dict[str, list[int]]:
    if production_rows is not None:
        return production_daily_schedule(production_rows, factories, days)
    return {
        factory.name: [int(round(float(factory.daily or 0.0)))] * days
        for factory in factories
    }


def production_rows_from_amounts(
    factories: list[Factory],
    product_name: str,
    amounts: dict[str, float],
    days: int,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for factory in factories:
        amount = ceil_int(max(0.0, float(amounts.get(factory.name, 0.0) or 0.0)))
        rows.append(
            {
                "factory": factory.name,
                "product": factory.product or product_name,
                "amount": amount,
                "capacity": float(factory.daily or 0.0) * days,
                "init": factory.init,
            }
        )
    return rows


def feasible_production_amount(
    factory: Factory,
    factory_materials: list[FactoryMaterial],
    products: list[Product],
    days: int,
) -> int:
    cap = int(math.floor(float(factory.daily or 0.0) * days))
    if cap <= 0:
        return 0
    material_caps: list[int] = []
    for item in factory_materials:
        if item.factory != factory.name:
            continue
        bom = material_bom(products, item.material)
        if bom > 0:
            material_caps.append(int(math.floor(float(item.init or 0.0) / bom)))
    if material_caps:
        cap = min(cap, min(material_caps))
    return max(0, cap)


def finished_goods_capacity_sources(factories: list[Factory], days: int) -> dict[str, dict[str, Any]]:
    return {
        factory.name: {
            "initial": factory.init,
            "supply": [int(round(factory.daily))] * days,
            "supply_is_capacity": True,
            "limit": 0.0,
        }
        for factory in factories
    }


def procurement_summary_rows(material_transport: list[dict[str, Any]], note: str) -> list[dict[str, Any]]:
    procurement_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in material_transport:
        if row.get("supplier") == "缺口":
            continue
        key = (row["material"], row["supplier"], row["factory"])
        current = procurement_by_key.setdefault(
            key,
            {
                "material": row["material"],
                "supplier": row["supplier"],
                "factory": row["factory"],
                "amount": 0,
                "unit_price": row.get("unit_price", 0.0),
                "purchase_cost": 0.0,
                "freight_cost": 0.0,
                "route": row.get("route", ""),
                "mode": row.get("mode", ""),
                "carrier": row.get("carrier", ""),
                "lead": row.get("lead", 0),
                "ship_day": row.get("ship_day", 1),
                "arrival_day": row.get("arrival_day", 1),
                "note": note,
            },
        )
        current["amount"] += int(row.get("amount") or 0)
        current["purchase_cost"] += float(row.get("purchase_cost") or 0.0)
        current["freight_cost"] += float(row.get("freight_cost") or 0.0)
    return sorted(procurement_by_key.values(), key=lambda row: (row["material"], row["supplier"], row["factory"]))


def material_procurement_for_production(
    *,
    production_rows: list[dict[str, Any]],
    factories: list[Factory],
    factory_materials: list[FactoryMaterial],
    products: list[Product],
    suppliers: list[Supplier],
    routes: list[Route],
    rates: dict[tuple[str, str], float],
    carriers: list[Any] | None,
    days: int,
    name_prefix: str,
    note: str,
) -> dict[str, Any]:
    production_daily_by_factory = production_daily_schedule(production_rows, factories, days)
    material_shipments: list[dict[str, Any]] = []
    material_failures: list[str] = []
    material_status: dict[str, str] = {}
    product_loss_by_factory: dict[str, float] = defaultdict(float)
    transports: list[dict[str, Any]] = []

    for idx, item in enumerate(factory_materials):
        production_daily = production_daily_by_factory.get(item.factory, [0] * days)
        if not production_daily:
            production_daily = [0] * days
        transport = _solve_material_procurement_milp(
            material=item,
            production_daily=production_daily,
            products=products,
            suppliers=suppliers,
            routes=routes,
            rates=rates,
            carriers=carriers,
            days=days,
            name=f"{name_prefix}_{idx}",
            note=note,
        )
        transports.append(transport)
        material_status[f"{item.factory}-{item.material}"] = str(transport.get("status", "Unknown"))
        material_failures.extend(str(failure) for failure in transport.get("failures", []))
        material_shipments.extend(transport.get("shipments", []))
        shortage = float(transport.get("shortage", 0.0) or 0.0)
        bom = material_bom(products, item.material)
        if shortage > 0 and bom > 0:
            product_loss_by_factory[item.factory] = max(product_loss_by_factory[item.factory], shortage / bom)

    planned_by_factory = {
        sv(row.get("factory")): float(row.get("amount") or 0.0)
        for row in production_rows
    }
    planned_production = sum(planned_by_factory.values())
    actual_production = sum(
        max(0.0, planned - product_loss_by_factory.get(factory_name, 0.0))
        for factory_name, planned in planned_by_factory.items()
    )
    production_satisfaction = min(1.0, actual_production / max(planned_production, 0.001)) if planned_production else 1.0
    procurement = procurement_summary_rows(material_shipments, note)
    return {
        "transports": transports,
        "shipments": material_shipments,
        "procurement": procurement,
        "purchase_cost": sum(float(row.get("purchase_cost") or 0.0) for row in procurement),
        "freight_cost": sum(float(row.get("freight_cost") or 0.0) for row in procurement),
        "planned_production": planned_production,
        "actual_production": actual_production,
        "production_satisfaction": production_satisfaction,
        "failures": material_failures,
        "status": material_status,
        "loss_by_factory": dict(product_loss_by_factory),
    }


def forecast_node(node: SalesNode, days: int = PLAN_DAYS, use_bias: bool = True) -> dict[str, Any]:
    if node.daily:
        daily_values = [value for value in node.daily[:days] if value > 0]
        forecast = sum(node.daily[:days])
        daily_avg = forecast / max(1, len(node.daily[:days]))
        daily_std = float(np.std(daily_values)) if daily_values else 0.0
        method = "逐日数据汇总"
    else:
        forecast = moving_weighted(node.monthly, use_bias=use_bias) * days / PLAN_DAYS
        daily_avg = forecast / days
        daily_std = float(np.std([value / PLAN_DAYS for value in node.monthly if value > 0]))
        method = "移动加权平均+趋势偏差" if use_bias else "移动加权平均"

    forecast_amount = max(0, int(round(forecast)))
    cv = daily_std / max(daily_avg, 0.001)
    safety_days = 1 if cv < 0.10 else 2 if cv < 0.25 else 3
    safety_stock = min(max(0, node.limit - node.init), math.ceil(daily_avg * safety_days))
    return {
        "node": node.node,
        "product": node.product,
        "forecast": forecast_amount,
        "daily_avg": daily_avg,
        "daily_std": daily_std,
        "safety_days": safety_days,
        "safety_stock": ceil_int(safety_stock),
        "init": node.init,
        "limit": node.limit,
        "excess_fee": node.excess_fee,
        "method": method,
        "plan_days": days,
        "daily_demand": [max(0, int(round(value))) for value in node.daily[:days]] if node.daily else spread_integer(forecast_amount, days),
    }


def currency_to_cny(value: float, currency: str, rates: dict[tuple[str, str], float]) -> float:
    return value * rates.get((currency, "CNY"), 1.0)


def route_mode(route: Route | None) -> str:
    if not route:
        return "无可用路线"
    if "码头" in route.route:
        return "海运干线+汽运接驳"
    if "火车站" in route.route:
        return "铁路干线+汽运接驳"
    return "汽运直达"


def route_points(route_name: str) -> list[str]:
    return [part.strip() for part in str(route_name).split("-->") if part.strip()]


def segment_mode(src: str, dst: str) -> str:
    if "火车站" in src and "火车站" in dst:
        return "铁路"
    if ("码头" in src or "港" in src) and ("码头" in dst or "港" in dst):
        return "海运"
    return "汽运"


def carrier_value(carrier: Any, attr: str, default: Any = None) -> Any:
    return getattr(carrier, attr, default)


def carrier_name(carrier: Any) -> str:
    return sv(carrier_value(carrier, "name", ""))


def carrier_mode_value(carrier: Any) -> str:
    mode = sv(carrier_value(carrier, "mode", ""))
    if mode:
        return mode
    type_code = sv(carrier_value(carrier, "type", "")).upper()
    return {"P": "汽运", "R": "铁路", "S": "海运", "W": "海运"}.get(type_code, type_code)


def is_international_road_segment(src: str, dst: str) -> bool:
    text = f"{src}{dst}"
    foreign_keywords = ("沙特", "达曼", "石油")
    domestic_keywords = ("防城", "贵阳", "贵州", "贵溪", "江西", "达州", "瓮福", "火车站")
    return any(keyword in text for keyword in foreign_keywords) and not all(keyword in text for keyword in domestic_keywords)


def segment_carrier_candidates(mode: str, src: str, dst: str, carriers: list[Any] | None) -> list[Any]:
    if not carriers:
        return []
    candidates = [carrier for carrier in carriers if carrier_mode_value(carrier) == mode]
    if mode == "汽运":
        international = is_international_road_segment(src, dst)
        filtered = [carrier for carrier in candidates if ("国际" in carrier_name(carrier)) == international]
        if filtered:
            candidates = filtered
        elif not international:
            domestic = [carrier for carrier in candidates if "国际" not in carrier_name(carrier)]
            if domestic:
                candidates = domestic
    return candidates


def route_segments(route_name: str) -> list[tuple[str, str, str]]:
    points = route_points(route_name)
    return [(src, dst, f"{src}-->{dst}") for src, dst in zip(points, points[1:])]


def segment_distance(segment: str, route: Route | None = None) -> float | None:
    context = ACTIVE_HAR_CONTEXT or {}
    har_distances = context.get("segment_distances") or {}
    if isinstance(har_distances, dict) and segment in har_distances:
        distance = float(har_distances.get(segment) or 0.0)
        if distance > 0:
            return distance
    if segment in VERIFIED_SEGMENT_DISTANCES:
        return float(VERIFIED_SEGMENT_DISTANCES[segment])
    if route and route.route == segment and route.distance > 0:
        return float(route.distance)
    if route and route.segment_distances:
        for (_src, _dst, item), distance in zip(route_segments(route.route), route.segment_distances):
            if item == segment and distance is not None and distance > 0:
                return float(distance)
    return None


def carrier_segment_days(carrier: Any, distance: float) -> int | None:
    efficiency = float(carrier_value(carrier, "efficiency_km_per_day", 0.0) or 0.0)
    if efficiency <= 0 or distance <= 0:
        return None
    return max(1, ceil_int(distance / efficiency))


def carrier_segment_cost(carrier: Any, amount: float, charge_ratio: float, distance: float) -> float:
    billable_distance = max(0, math.floor(float(distance) + 0.5))
    billable = max(amount, 0.0) * max(charge_ratio, 0.001) * billable_distance
    variable = billable * float(carrier_value(carrier, "unit_rate_cny", 0.0) or 0.0)
    minimum = float(carrier_value(carrier, "start_fee_cny", 0.0) or 0.0)
    return max(variable, minimum)


def carrier_segment_slope(carrier: Any, charge_ratio: float, distance: float) -> float:
    billable_distance = max(0, math.floor(float(distance) + 0.5))
    return float(carrier_value(carrier, "unit_rate_cny", 0.0) or 0.0) * max(charge_ratio, 0.001) * billable_distance


def best_carrier_for_segment(
    *,
    src: str,
    dst: str,
    segment: str,
    amount: float,
    charge_ratio: float,
    distance: float,
    carriers: list[Any] | None,
    expected_days: int | None = None,
) -> Any | None:
    candidates = segment_carrier_candidates(segment_mode(src, dst), src, dst, carriers)
    if not candidates:
        return None
    return min(
        candidates,
        key=lambda carrier: (
            carrier_segment_cost(carrier, amount, charge_ratio, distance),
            carrier_segment_days(carrier, distance) or 999,
            carrier_name(carrier),
        ),
    )


def fallback_carrier_name_for_route(route: Route | None) -> str:
    if not route:
        return "无承运商"
    if "火车站" in route.route:
        return "铁路承运商"
    if "码头" in route.route or "港" in route.route:
        return "海运承运商"
    return "汽运承运商"


def fallback_transport_option(route: Route, charge_ratio: float, carriers: list[Any] | None = None) -> TransportOption:
    return TransportOption(
        route=route,
        carrier=inferred_route_carrier_fallback(route, carriers),
        segment_costs=((float(route.rate or 0.0) * max(charge_ratio, 0.001), float(route.min_freight or 0.0)),),
        lead=max(0, int(route.lead or 0)),
    )


def fallback_segment_carrier_names(route: Route, carriers: list[Any] | None) -> list[str]:
    segments = route_segments(route.route)
    inferred = inferred_route_carrier_fallback(route, carriers)
    inferred_parts = [part.strip() for part in inferred.split("+") if part.strip()]
    if len(inferred_parts) == len(segments):
        return inferred_parts

    names: list[str] = []
    for src, dst, _segment in segments:
        candidates = segment_carrier_candidates(segment_mode(src, dst), src, dst, carriers)
        if len(candidates) == 1:
            names.append(carrier_name(candidates[0]))
        elif candidates:
            best = min(
                candidates,
                key=lambda c: (
                    float(carrier_value(c, "start_fee_cny", 0.0) or 0.0),
                    float(carrier_value(c, "unit_rate_cny", 0.0) or 0.0),
                    carrier_name(c),
                ),
            )
            names.append(carrier_name(best))
        else:
            names.append(fallback_carrier_name_for_route(route))
    return names


def option_slope_sum(option: TransportOption) -> float:
    return sum(float(slope) for slope, _minimum in option.segment_costs)


def option_minimum_sum(option: TransportOption) -> float:
    return sum(float(minimum) for _slope, minimum in option.segment_costs)


def option_breakpoints(option: TransportOption, max_amount: float) -> list[float]:
    points = {1.0, max(1.0, float(max_amount or 1.0))}
    for slope, minimum in option.segment_costs:
        slope = float(slope)
        minimum = float(minimum)
        if slope > 0 and minimum > 0:
            qty = minimum / slope
            for candidate in (qty - 1, qty, qty + 1):
                if 1 <= candidate <= max_amount:
                    points.add(float(candidate))
    return sorted(points)


def prune_dominated_transport_options(options: list[TransportOption], max_amount: float | None = None) -> list[TransportOption]:
    if len(options) <= 1 or max_amount is None or float(max_amount or 0.0) <= 1.0:
        return options
    reference_amount = max(1.0, float(max_amount or 1.0))
    sample_points = {1.0, 10.0, 100.0, 1000.0, reference_amount}
    for option in options:
        sample_points.update(option_breakpoints(option, reference_amount))
    sample_points = {point for point in sample_points if 1 <= point <= reference_amount}
    kept: list[TransportOption] = []
    for option in options:
        dominated = False
        for other in options:
            if other is option:
                continue
            if int(other.lead or 0) > int(option.lead or 0):
                continue
            if other.carrier == option.carrier and other.route.route == option.route.route:
                continue
            if all(transport_option_cost(other, point) <= transport_option_cost(option, point) + 1e-6 for point in sample_points):
                if (
                    int(other.lead or 0) < int(option.lead or 0)
                    or option_slope_sum(other) < option_slope_sum(option) - 1e-9
                    or option_minimum_sum(other) < option_minimum_sum(option) - 1e-6
                ):
                    dominated = True
                    break
        if not dominated:
            kept.append(option)
    return kept or options


ROUTE_TRANSPORT_OPTIONS_CACHE: dict[tuple[Any, ...], tuple[TransportOption, ...]] = {}


def carriers_cache_signature(carriers: list[Any] | None) -> tuple[Any, ...]:
    if not carriers:
        return ()
    return tuple(
        (
            carrier_name(carrier),
            carrier_mode_value(carrier),
            round(float(carrier_value(carrier, "efficiency_km_per_day", 0.0) or 0.0), 6),
            round(float(carrier_value(carrier, "unit_rate_cny", 0.0) or 0.0), 8),
            round(float(carrier_value(carrier, "start_fee_cny", 0.0) or 0.0), 4),
        )
        for carrier in carriers
    )


def har_route_context_signature() -> tuple[Any, ...]:
    context = ACTIVE_HAR_CONTEXT or {}
    segment_days = context.get("segment_days_by_carrier") or {}
    if not isinstance(segment_days, dict):
        return ()
    return tuple(sorted((str(key), int(value or 0)) for key, value in segment_days.items()))


def route_options_cache_key(
    route: Route | None,
    charge_ratio: float,
    carriers: list[Any] | None,
    max_amount: float | None,
    prune_dominated: bool,
) -> tuple[Any, ...]:
    if route is None:
        return ("none",)
    segments = route_segments(route.route)
    segment_distances = []
    for _src, _dst, segment in segments:
        distance = segment_distance(segment, route)
        segment_distances.append((segment, None if distance is None else round(float(distance), 6)))
    return (
        route.route,
        route.src,
        route.dst,
        round(float(route.distance or 0.0), 6),
        round(float(route.rate or 0.0), 8),
        round(float(route.min_freight or 0.0), 4),
        int(route.lead or 0),
        route.currency,
        tuple(None if item is None else round(float(item), 6) for item in (route.segment_distances or ())),
        tuple(segment_distances),
        har_route_context_signature(),
        round(float(charge_ratio or 0.0), 8),
        carriers_cache_signature(carriers),
        None if max_amount is None else round(float(max_amount), 6),
        bool(prune_dominated),
    )


def route_transport_options(
    route: Route | None,
    charge_ratio: float = 1.0,
    carriers: list[Any] | None = None,
    max_amount: float | None = None,
    prune_dominated: bool = True,
) -> list[TransportOption]:
    key = route_options_cache_key(route, charge_ratio, carriers, max_amount, prune_dominated)
    cached = ROUTE_TRANSPORT_OPTIONS_CACHE.get(key)
    if cached is not None:
        return list(cached)
    if len(ROUTE_TRANSPORT_OPTIONS_CACHE) > 200_000:
        ROUTE_TRANSPORT_OPTIONS_CACHE.clear()
    if not route:
        options: list[TransportOption] = []
        ROUTE_TRANSPORT_OPTIONS_CACHE[key] = tuple(options)
        return options
    segments = route_segments(route.route)
    if not carriers or not segments:
        options = [fallback_transport_option(route, charge_ratio, carriers)]
        ROUTE_TRANSPORT_OPTIONS_CACHE[key] = tuple(options)
        return options

    fallback_names = fallback_segment_carrier_names(route, carriers)
    known_choices: list[tuple[int, list[tuple[str, float, float, int]]]] = []
    known_segment_count = 0
    for segment_idx, (src, dst, segment) in enumerate(segments):
        distance = segment_distance(segment, route)
        if distance is None:
            continue
        candidates = segment_carrier_candidates(segment_mode(src, dst), src, dst, carriers)
        if not candidates:
            continue
        choices = []
        for carrier in candidates:
            slope = carrier_segment_slope(carrier, charge_ratio, distance)
            minimum = float(carrier_value(carrier, "start_fee_cny", 0.0) or 0.0)
            lead = carrier_segment_days(carrier, distance) or verified_segment_duration(segment, carrier_name(carrier)) or verified_segment_duration(segment, "") or 0
            choices.append((carrier_name(carrier), slope, minimum, int(lead)))
        known_choices.append((segment_idx, choices))
        known_segment_count += 1

    if known_segment_count <= 0:
        return [fallback_transport_option(route, charge_ratio, carriers)]

    options: list[TransportOption] = []
    choice_lists = [choices for _segment_idx, choices in known_choices]
    baseline_known_slope = sum(min(float(item[1]) for item in choices) for _segment_idx, choices in known_choices)
    baseline_known_minimum = sum(min(float(item[2]) for item in choices) for _segment_idx, choices in known_choices)
    for combo in iter_product(*choice_lists):
        carrier_parts = list(fallback_names)
        known_slope = 0.0
        known_minimum = 0.0
        known_lead = 0
        segment_costs: list[tuple[float, float]] = []
        for (segment_idx, _choices), item in zip(known_choices, combo):
            carrier_parts[segment_idx] = item[0]
            slope = float(item[1])
            minimum = float(item[2])
            known_slope += slope
            known_minimum += minimum
            known_lead += max(0, int(item[3]))
            segment_costs.append((slope, minimum))
        if known_segment_count < len(segments):
            total_slope = float(route.rate or 0.0) * max(charge_ratio, 0.001)
            residual_slope = max(0.0, total_slope - baseline_known_slope)
            residual_minimum = max(0.0, float(route.min_freight or 0.0) - baseline_known_minimum)
            if residual_slope > 0 or residual_minimum > 0:
                segment_costs.append((residual_slope, residual_minimum))
        lead = known_lead if known_segment_count == len(segments) else max(max(0, int(route.lead or 0)), known_lead)
        options.append(
            TransportOption(
                route=route,
                carrier="+".join(carrier_parts),
                segment_costs=tuple(segment_costs),
                lead=lead if lead > 0 else max(0, int(route.lead or 0)),
            )
        )
    options = options or [fallback_transport_option(route, charge_ratio, carriers)]
    if prune_dominated and max_amount is not None:
        options = prune_dominated_transport_options(options, max_amount)
    ROUTE_TRANSPORT_OPTIONS_CACHE[key] = tuple(options)
    return options


def transport_option_cost(option: TransportOption, amount: float) -> float:
    if amount <= 0:
        return 0.0
    return sum(max(float(amount) * slope, minimum) for slope, minimum in option.segment_costs)


def transport_option_economic_lot(option: TransportOption | None) -> int:
    if option is None:
        return 1
    lot = 1
    for slope, minimum in option.segment_costs:
        if slope > 0 and minimum > 0:
            lot = max(lot, ceil_int(minimum / slope))
    return lot


def best_transport_option(route: Route | None, amount: float, charge_ratio: float = 1.0, carriers: list[Any] | None = None) -> TransportOption | None:
    options = route_transport_options(route, charge_ratio, carriers)
    if not options:
        return None
    return min(
        options,
        key=lambda option: (
            transport_option_cost(option, amount),
            int(option.lead or 0),
            option.carrier,
        ),
    )


def route_effective_lead(route: Route | None, amount: float, charge_ratio: float = 1.0, carriers: list[Any] | None = None) -> int:
    option = best_transport_option(route, amount, charge_ratio, carriers)
    if option is not None:
        return int(option.lead or 0)
    return int(route.lead or 0) if route else 0


def route_cost(route: Route | None, amount: float, charge_ratio: float = 1.0, carriers: list[Any] | None = None) -> float:
    if not route or amount <= 0:
        return 0.0
    option = best_transport_option(route, amount, charge_ratio, carriers)
    return transport_option_cost(option, amount) if option is not None else 0.0


def route_unit_cost(route: Route | None, amount: float, charge_ratio: float = 1.0, carriers: list[Any] | None = None) -> float:
    return route_cost(route, amount, charge_ratio, carriers) / max(amount, 0.001)


def route_score(route: Route, amount: float, charge_ratio: float, urgent: bool = False, carriers: list[Any] | None = None) -> float:
    unit = route_unit_cost(route, amount, charge_ratio, carriers)
    lead_penalty = route_effective_lead(route, amount, charge_ratio, carriers) * (200.0 if urgent else 0.25)
    return unit + lead_penalty


def pick_best_route(
    routes: list[Route],
    src: str,
    dst: str,
    amount: float,
    charge_ratio: float = 1.0,
    urgent: bool = False,
    carriers: list[Any] | None = None,
) -> Route | None:
    exact = [route for route in routes if route.src == src and route.dst == dst]
    loose = [route for route in routes if src in route.route and dst in route.route]
    candidates = exact or loose
    if not candidates:
        return None
    return min(candidates, key=lambda route: route_score(route, amount, charge_ratio, urgent, carriers))


def material_bom(products: list[Product], material: str) -> float:
    match = next((product for product in products if product.kind == "原料" and product.name == material), None)
    return match.bom if match and match.bom > 0 else 1.0


def charge_ratio(products: list[Product], cargo: str) -> float:
    match = next((product for product in products if product.name == cargo), None)
    return match.charge_ratio if match else 1.0


def case_keyword(xls_path: Path) -> str:
    stem = xls_path.stem
    stem = re.sub(r"[★☆\s].*$", "", stem)
    for suffix in ("综合运营", "国际采购", "原料采购", "采购", "生产", "销售"):
        if stem.endswith(suffix):
            return stem[: -len(suffix)] or stem
    match = re.match(r"([\u4e00-\u9fffA-Za-z0-9]+?)(综合|采购|生产|销售|运营)", stem)
    return match.group(1) if match else stem


def score_config(xls_path: Path, qtype: str) -> dict[str, Any]:
    keyword = case_keyword(xls_path)
    base = DEFAULT_SCORE_CONFIG.get(qtype, DEFAULT_SCORE_CONFIG["采购"])
    config = {
        "plan_days": base.get("plan_days", PLAN_DAYS),
        "targets": dict(base.get("targets", {})),
        "points": dict(base.get("points", {})),
        "forecast_bias": base.get("forecast_bias", True),
    }
    for key, override in CASE_SCORE_CONFIG.items():
        if key in keyword or key in xls_path.name:
            config["plan_days"] = override.get("plan_days", config["plan_days"])
            config["targets"].update(override.get("targets", {}))
            config["points"].update(override.get("points", {}))
            if "forecast_bias" in override:
                config["forecast_bias"] = override["forecast_bias"]
            break
    return config


def plan_days_for_case(xls_path: Path, qtype: str) -> int:
    return int(score_config(xls_path, qtype).get("plan_days", PLAN_DAYS) or PLAN_DAYS)


def forecast_bias_for_case(xls_path: Path, qtype: str) -> bool:
    return bool(score_config(xls_path, qtype).get("forecast_bias", True))


def score_targets(xls_path: Path, qtype: str) -> dict[str, float]:
    return dict(score_config(xls_path, qtype).get("targets", {}))


def score_points(xls_path: Path, qtype: str) -> dict[str, float]:
    return dict(score_config(xls_path, qtype).get("points", {}))


def verified_score_override_for(xls_path: Path, qtype: str) -> dict[str, Any] | None:
    for key, override in VERIFIED_SCORE_OVERRIDES.items():
        if key not in xls_path.stem:
            continue
        if override.get("qtype") and override["qtype"] != qtype:
            continue
        return override
    return None


def apply_verified_score_override(result: dict[str, Any], xls_path: Path) -> None:
    override = verified_score_override_for(xls_path, str(result.get("qtype", "")))
    if not override:
        return
    if result.get("qtype") == "采购" and override.get("unit_procurement"):
        actual_cost = float(override["unit_procurement"])
        production_satisfaction = float(override.get("production_satisfaction", result.get("production_satisfaction", 1.0)))
        targets = score_targets(xls_path, "采购")
        target_cost = float(targets.get("unit_procurement_cost") or 0.0)
        procurement_points = cost_score(actual_cost, target_cost, 60.0)
        satisfaction_points = satisfaction_score(production_satisfaction, 40.0)
        result["unit_procurement"] = actual_cost
        result["production_satisfaction"] = production_satisfaction
        result["score_rows"] = [
            {"item": "单位采购成本", "target": target_cost, "actual": actual_cost, "points": procurement_points, "max": 60},
            {"item": "生产满足率", "target": 1.0, "actual": production_satisfaction, "points": satisfaction_points, "max": 40},
        ]
    if override.get("score") is not None:
        result["score"] = float(override["score"])
    result["score_note"] = str(override.get("note") or "平台回执校准")
    result.setdefault("assumptions", []).append(str(override.get("note") or "平台回执校准"))


def score_note_with_level(result: dict[str, Any], xls_path: Path) -> str:
    qtype = str(result.get("qtype", ""))
    if verified_score_override_for(xls_path, qtype):
        return str(result.get("score_note") or "平台回执校准")
    raw_score = result.get("score")
    if isinstance(raw_score, (int, float)):
        return f"公式估算，待平台回执校准：{result.get('score_note', '')}".strip()
    return str(result.get("score_note") or "未配置正式评分指标")



def inferred_route_carrier_fallback(route: Route | None, carriers: list[Any] | None) -> str:
    if not route:
        return "无承运商"
    if carriers is not None:
        try:
            from carrier_infer import infer_route_carrier

            return infer_route_carrier(route, carriers)
        except Exception:
            pass
    return fallback_carrier_name_for_route(route)


def route_carrier(route: Route | None, carriers: list[Any] | None = None, amount: float | None = None, charge_ratio: float = 1.0) -> str:
    if not route:
        return "无承运商"
    if not carriers or amount is None or amount <= 0:
        return inferred_route_carrier_fallback(route, carriers)
    option = best_transport_option(route, amount, charge_ratio, carriers)
    return option.carrier if option is not None else inferred_route_carrier_fallback(route, carriers)


def parse_carriers_safe(sections: dict[str, list[tuple[Any, ...]]], rates: dict[tuple[str, str], float]) -> list[Any] | None:
    har_carriers = har_carriers_for_rates(rates)
    if har_carriers:
        return har_carriers
    try:
        from carrier_infer import parse_carriers

        return parse_carriers(sections, rates)
    except Exception:
        return None


def cost_score(actual: float, target: float, points: float) -> float:
    """平台成本类评分: clamp((1 - (actual - target) / (target * 20%)) * points, 0, points)."""
    if target <= 0:
        return 0.0
    return max(0.0, min(points, (1 - (actual - target) / (target * 0.2)) * points))


def satisfaction_score(actual_ratio: float, points: float) -> float:
    """平台满足率评分: target=100% 时等价于 actual_ratio * points，上限 points。"""
    return max(0.0, min(points, actual_ratio * points))


def deviation_score(actual_deviation: float, target_deviation: float, points: float) -> float:
    """平台预测偏差评分: clamp((1 - actual_deviation / target_deviation) * points, 0, points)."""
    if target_deviation <= 0:
        return 0.0
    return max(0.0, min(points, (1 - actual_deviation / target_deviation) * points))


FULL_SATISFACTION_EPS = 1e-6


def is_full_satisfaction(value: float, shortage: float = 0.0) -> bool:
    return float(value or 0.0) >= 1.0 - FULL_SATISFACTION_EPS and float(shortage or 0.0) <= FULL_SATISFACTION_EPS


def product_transport_priority(metrics: dict[str, float], shipments_count: int = 0) -> tuple[float, float, float, float, float]:
    """Mode-aware objective for product transport candidates."""
    market = float(metrics.get("market_satisfaction", 0.0) or 0.0)
    shortage = float(metrics.get("shortage", 0.0) or 0.0)
    unit_logistics = float(metrics.get("unit_logistics", 0.0) or 0.0)
    score = float(metrics.get("score", 0.0) or 0.0)
    full = 1.0 if is_full_satisfaction(market, shortage) else 0.0
    if is_extreme_mode():
        return (
            full,
            market,
            -unit_logistics,
            score,
            -float(shipments_count or 0),
        )
    return (
        full,
        market,
        score,
        -float(shipments_count or 0),
        -unit_logistics,
        )


def is_time_limited_status(status: dict[str, Any]) -> bool:
    text = " ".join(str(status.get(key, "")) for key in ("status", "raw_status", "method"))
    return "TimeLimit" in text


def normalized_cost_value(actual: float, target: float) -> float:
    actual_value = float(actual or 0.0)
    target_value = float(target or 0.0)
    return actual_value / target_value if target_value > 0 else actual_value


def result_satisfaction_cost_priority(result: dict[str, Any], targets: dict[str, float], rows_count: int = 0) -> tuple[float, float, float, float, float, float, float]:
    """Mode-aware objective for full-result candidates."""
    production = float(result.get("production_satisfaction", 1.0) or 0.0)
    market = float(result.get("market_satisfaction", 1.0) or 0.0)
    unit_procurement = float(result.get("unit_procurement", 0.0) or 0.0)
    unit_logistics = float(result.get("unit_logistics", 0.0) or 0.0)
    normalized_cost = (
        normalized_cost_value(unit_procurement, float(targets.get("unit_procurement_cost", 0.0) or 0.0))
        + normalized_cost_value(unit_logistics, float(targets.get("unit_logistics_cost", 0.0) or 0.0))
    )
    full = 1.0 if is_full_satisfaction(production) and is_full_satisfaction(market) else 0.0
    score = float(result.get("score", 0.0) or 0.0)
    if is_extreme_mode():
        return (full, min(production, market), production, market, -normalized_cost, score, -float(rows_count or 0))
    return (full, min(production, market), production, market, -float(rows_count or 0), score, -normalized_cost)


def allocate_procurement(
    material: str,
    factory: str,
    amount: float,
    suppliers: list[Supplier],
    routes: list[Route],
    products: list[Product],
    rates: dict[tuple[str, str], float],
    urgent: bool,
    days: int = PLAN_DAYS,
    carriers: list[Any] | None = None,
) -> tuple[list[dict[str, Any]], float, float]:
    related = [supplier for supplier in suppliers if supplier.material == material]
    ranked: list[tuple[float, Supplier, Route | None, float, float]] = []
    ratio = charge_ratio(products, material)
    for supplier in related:
        route = pick_best_route(routes, supplier.name, factory, amount, ratio, urgent, carriers)
        unit_purchase = currency_to_cny(supplier.price, supplier.currency, rates)
        unit_freight = route_unit_cost(route, max(amount, 1.0), ratio, carriers)
        lead_risk = route_effective_lead(route, max(amount, 1.0), ratio, carriers) * (200.0 if urgent else 10.0) if route else 99_999.0
        ranked.append((unit_purchase + unit_freight + lead_risk, supplier, route, unit_purchase, unit_freight))
    ranked.sort(key=lambda item: item[0])

    remaining = amount
    rows: list[dict[str, Any]] = []
    total_purchase = 0.0
    total_freight = 0.0
    for _, supplier, route, unit_purchase, _ in ranked:
        if remaining <= 0:
            break
        available = supplier.available or supplier.daily * days or remaining
        allocation = min(remaining, available)
        allocation = ceil_int(allocation)
        if allocation <= 0:
            continue
        freight = route_cost(route, allocation, ratio, carriers)
        rows.append(
            {
                "material": material,
                "supplier": supplier.name,
                "factory": factory,
                "amount": allocation,
                "unit_price": unit_purchase,
                "purchase_cost": allocation * unit_purchase,
                "freight_cost": freight,
                "route": route.route if route else "未找到可用路线",
                "mode": route_mode(route),
                "carrier": route_carrier(route, carriers, allocation, ratio),
                "lead": route_effective_lead(route, allocation, ratio, carriers) if route else 0,
                "ship_day": 1,
                "arrival_day": 1 + (route_effective_lead(route, allocation, ratio, carriers) if route else 0),
                "note": "按到岸成本排序" if route else "缺路线，需人工确认",
            }
        )
        total_purchase += allocation * unit_purchase
        total_freight += freight
        remaining -= allocation
    if remaining > 0:
        rows.append(
            {
                "material": material,
                "supplier": "缺口",
                "factory": factory,
                "amount": remaining,
                "unit_price": 0.0,
                "purchase_cost": 0.0,
                "freight_cost": 0.0,
                "route": "供应商可供量不足",
                "mode": "风险",
                "carrier": "无承运商",
                "lead": 0,
                "ship_day": 1,
                "arrival_day": 1,
                "note": "会影响生产满足率",
            }
        )
    return rows, total_purchase, total_freight


def daily_demands_for_forecast(forecast: dict[str, Any], days: int) -> list[int]:
    values = forecast.get("daily_demand")
    if values:
        cleaned = [ceil_int(value) for value in list(values)[:days]]
        if len(cleaned) < days:
            cleaned.extend([0] * (days - len(cleaned)))
        return cleaned
    return spread_integer(float(forecast.get("forecast", 0.0)), days)


def daily_market_replay(
    forecasts: list[dict[str, Any]],
    product_transport_rows: list[dict[str, Any]],
    days: int | None = None,
) -> dict[str, Any]:
    if days is None:
        days = max([int(row.get("plan_days") or 0) for row in forecasts] + [PLAN_DAYS])

    demand_by_node = {row["node"]: daily_demands_for_forecast(row, days) for row in forecasts}
    inventory = {row["node"]: float(row.get("init", 0.0) or 0.0) for row in forecasts}
    arrivals: dict[tuple[str, int], int] = {}
    for row in product_transport_rows:
        destination = sv(row.get("destination") or row.get("node") or row.get("dst"))
        if not destination:
            continue
        amount = ceil_int(row.get("amount", 0))
        ship_day = int(row.get("ship_day") or 1)
        if row.get("arrival_day"):
            arrival_day = int(row["arrival_day"])
        else:
            arrival_day = ship_day + int(row.get("lead") or 0)
        if 1 <= arrival_day <= days and amount > 0:
            arrivals[(destination, arrival_day)] = arrivals.get((destination, arrival_day), 0) + amount

    daily_rows: list[dict[str, Any]] = []
    total_demand = 0
    total_served = 0
    total_shortage = 0
    over_limit_days = 0
    forecast_by_node = {row["node"]: row for row in forecasts}
    for day in range(1, days + 1):
        for node, demands in demand_by_node.items():
            inventory[node] = inventory.get(node, 0.0) + arrivals.get((node, day), 0)
            demand = demands[day - 1] if day - 1 < len(demands) else 0
            served = min(inventory[node], demand)
            shortage = max(0.0, demand - served)
            inventory[node] -= served
            limit = float(forecast_by_node.get(node, {}).get("limit", 0.0) or 0.0)
            over_limit = max(0.0, inventory[node] - limit) if limit > 0 else 0.0
            if limit > 0 and inventory[node] > limit + 1e-6:
                over_limit_days += 1
            total_demand += int(round(demand))
            total_served += int(round(served))
            total_shortage += int(round(shortage))
            daily_rows.append(
                {
                    "day": day,
                    "node": node,
                    "arrivals": arrivals.get((node, day), 0),
                    "demand": int(round(demand)),
                    "served": int(round(served)),
                    "shortage": int(round(shortage)),
                    "ending_inventory": int(round(inventory[node])),
                    "limit": int(round(limit)) if limit > 0 else 0,
                    "over_limit": int(math.ceil(over_limit)) if over_limit > 1e-6 else 0,
                }
            )

    risks = []
    if total_shortage > 0:
        risks.append(f"销售网点逐日断货 {total_shortage:,}")
    if over_limit_days > 0:
        risks.append(f"销售网点库存超过上限 {over_limit_days} 个节点日")
    return {
        "days": days,
        "demand": total_demand,
        "served": total_served,
        "shortage": total_shortage,
        "market_satisfaction": min(1.0, total_served / max(total_demand, 1)),
        "ending_inventory": {node: int(round(value)) for node, value in inventory.items()},
        "daily_rows": daily_rows,
        "risks": risks,
    }


def repair_factory_ship_days(
    shipments: list[dict[str, Any]],
    factories: list[Factory],
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    """Delay finished-goods shipments that would make a factory's daily stock negative."""
    if not shipments or not factories:
        return shipments

    rows = [dict(row) for row in shipments]
    factory_by_name = {factory.name: factory for factory in factories}
    output_by_factory = factory_output_schedule(factories, days, production_rows)

    def row_factory(row: dict[str, Any]) -> str:
        return sv(row.get("factory") or row.get("source"))

    def delay_priority(row: dict[str, Any]) -> tuple[int, int, int]:
        route = sv(row.get("route"))
        is_bridge = "断货桥接" in sv(row.get("note"))
        is_direct = "火车站" not in route and "码头" not in route and "港" not in route
        if is_bridge:
            route_rank = 2
        elif is_direct:
            route_rank = 0
        else:
            route_rank = 1
        return (route_rank, int(row.get("lead") or 0), int(row.get("amount") or 0))

    for factory_name, factory in factory_by_name.items():
        related = [idx for idx, row in enumerate(rows) if row_factory(row) == factory_name]
        if not related:
            continue
        for _ in range(max(1, days * max(1, len(related)))):
            changed = False
            inventory = float(factory.init or 0.0)
            for day in range(1, days + 1):
                inventory += float(output_by_factory.get(factory_name, [0] * days)[day - 1])
                today = [idx for idx in related if int(rows[idx].get("ship_day") or 1) == day]
                outbound = sum(float(rows[idx].get("amount") or 0.0) for idx in today)
                if outbound <= inventory + 1e-6:
                    inventory -= outbound
                    continue

                excess = outbound - inventory
                moved = 0.0
                for idx in sorted(today, key=lambda item: delay_priority(rows[item])):
                    if moved >= excess - 1e-6:
                        break
                    if int(rows[idx].get("ship_day") or 1) >= days:
                        continue
                    rows[idx]["ship_day"] = int(rows[idx].get("ship_day") or 1) + 1
                    rows[idx]["arrival_day"] = int(rows[idx].get("arrival_day") or rows[idx]["ship_day"]) + 1
                    moved += float(rows[idx].get("amount") or 0.0)
                    changed = True
                break
            if not changed:
                break

    return sorted(rows, key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("route")), sv(row.get("destination"))))


def effective_shipments(shipments: list[dict[str, Any]], days: int) -> list[dict[str, Any]]:
    rows = []
    for row in shipments:
        arrival_day = int(row.get("arrival_day") or (int(row.get("ship_day") or 1) + int(row.get("lead") or 0)))
        amount = int(row.get("amount") or 0)
        if amount <= 0 or arrival_day > days:
            continue
        rows.append(row)
    return rows


def route_for_shipment(row: dict[str, Any], routes: list[Route]) -> Route | None:
    route_name = sv(row.get("route"))
    source = sv(row.get("source") or row.get("factory"))
    destination = sv(row.get("destination") or row.get("node") or row.get("dst"))
    for route in routes:
        if route.route == route_name and (not source or route.src == source) and (not destination or route.dst == destination):
            return route
    for route in routes:
        if route.route == route_name:
            return route
    return None


def refresh_shipment_route_fields(
    row: dict[str, Any],
    route: Route | None,
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None = None,
) -> None:
    if not route:
        return
    amount = int(row.get("amount") or 0)
    ratio = charge_ratio(products, cargo)
    row["lead"] = route_effective_lead(route, max(amount, 1), ratio, carriers)
    row["arrival_day"] = int(row.get("ship_day") or 1) + int(row["lead"])
    row["freight_cost"] = route_cost(route, amount, ratio, carriers)
    row["carrier"] = route_carrier(route, carriers, amount, ratio)


def refresh_transport_rows(
    rows: list[dict[str, Any]],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None = None,
) -> list[dict[str, Any]]:
    refreshed: list[dict[str, Any]] = []
    for row in rows:
        copy = dict(row)
        route = route_for_shipment(copy, routes)
        refresh_shipment_route_fields(copy, route, products, cargo, carriers)
        refreshed.append(copy)
    return refreshed


def reduce_shipment_amount(
    rows: list[dict[str, Any]],
    idx: int,
    reduction: int,
    *,
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None = None,
) -> int:
    if idx < 0 or idx >= len(rows) or reduction <= 0:
        return 0
    row = rows[idx]
    amount = int(row.get("amount") or 0)
    if amount <= 0:
        rows.pop(idx)
        return 0
    route = route_for_shipment(row, routes)
    remaining = amount - min(reduction, amount)
    if remaining <= 0:
        rows.pop(idx)
        return amount
    row["amount"] = int(remaining)
    if route:
        refresh_shipment_route_fields(row, route, products, cargo, carriers)
    return amount - remaining


def trim_factory_overdraw_shipments(
    shipments: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    days: int,
    carriers: list[Any] | None = None,
    production_rows: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    rows = effective_shipments([dict(row) for row in shipments], days)
    if not rows or not factories:
        return rows
    output_by_factory = factory_output_schedule(factories, days, production_rows)

    for _ in range(max(1, len(rows) * max(days, 1))):
        changed = False
        for factory in factories:
            inventory = float(factory.init or 0.0)
            for day in range(1, days + 1):
                inventory += float(output_by_factory.get(factory.name, [0] * days)[day - 1])
                today = [
                    idx for idx, row in enumerate(rows)
                    if sv(row.get("factory") or row.get("source")) == factory.name
                    and int(row.get("ship_day") or 1) == day
                ]
                outbound = sum(float(rows[idx].get("amount") or 0.0) for idx in today)
                if outbound <= inventory + 1e-6:
                    inventory -= outbound
                    continue
                excess = ceil_int(outbound - inventory)
                for idx in sorted(
                    today,
                    key=lambda item: (
                        float(rows[item].get("amount") or 0.0),
                        float(rows[item].get("freight_cost") or 0.0) / max(float(rows[item].get("amount") or 0.0), 1.0),
                    ),
                    reverse=True,
                ):
                    reduced = reduce_shipment_amount(rows, idx, excess, routes=routes, products=products, cargo=cargo, carriers=carriers)
                    excess -= reduced
                    changed = True
                    if excess <= 0:
                        break
                break
            if changed:
                break
        if not changed:
            break
    return sorted(effective_shipments(rows, days), key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("route")), sv(row.get("destination"))))


def trim_store_overstock_shipments(
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    days: int,
    carriers: list[Any] | None = None,
) -> list[dict[str, Any]]:
    rows = effective_shipments([dict(row) for row in shipments], days)
    if not rows or not forecasts:
        return rows

    for _ in range(max(1, len(rows) * max(days, 1))):
        replay = daily_market_replay(forecasts, rows, days)
        over_rows = [row for row in replay.get("daily_rows", []) if int(row.get("over_limit") or 0) > 0]
        if not over_rows:
            break
        issue = min(over_rows, key=lambda row: (int(row.get("day") or 1), sv(row.get("node"))))
        node = sv(issue.get("node"))
        day = int(issue.get("day") or 1)
        excess = int(issue.get("over_limit") or 0)
        candidates = []
        for idx, row in enumerate(rows):
            if sv(row.get("destination") or row.get("node") or row.get("dst")) != node:
                continue
            arrival_day = int(row.get("arrival_day") or (int(row.get("ship_day") or 1) + int(row.get("lead") or 0)))
            if arrival_day <= day:
                unit_cost = float(row.get("freight_cost") or 0.0) / max(float(row.get("amount") or 0.0), 1.0)
                candidates.append((arrival_day, float(row.get("amount") or 0.0), unit_cost, idx))
        if not candidates:
            break
        changed = False
        for _arrival_day, _amount, _unit_cost, idx in sorted(candidates, reverse=True):
            reduced = reduce_shipment_amount(rows, idx, excess, routes=routes, products=products, cargo=cargo, carriers=carriers)
            excess -= reduced
            changed = changed or reduced > 0
            if excess <= 0:
                break
        if not changed:
            break
    return sorted(effective_shipments(rows, days), key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("route")), sv(row.get("destination"))))


def sanitize_product_transport(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    days: int,
    carriers: list[Any] | None = None,
    production_rows: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    rows = effective_shipments(refresh_transport_rows([dict(row) for row in shipments], routes, products, cargo, carriers), days)
    for _ in range(4):
        signature = [(sv(row.get("route")), int(row.get("ship_day") or 1), int(row.get("amount") or 0)) for row in rows]
        rows = effective_shipments(repair_factory_ship_days(rows, factories, days, production_rows), days)
        rows = trim_factory_overdraw_shipments(rows, factories, routes, products, cargo, days, carriers, production_rows)
        rows = trim_store_overstock_shipments(forecasts, rows, routes, products, cargo, days, carriers)
        rows = effective_shipments(repair_factory_ship_days(rows, factories, days, production_rows), days)
        rows = effective_shipments(refresh_transport_rows(rows, routes, products, cargo, carriers), days)
        next_signature = [(sv(row.get("route")), int(row.get("ship_day") or 1), int(row.get("amount") or 0)) for row in rows]
        if next_signature == signature:
            break
    return sorted(rows, key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("route")), sv(row.get("destination"))))


def fill_market_shortages_with_fast_routes(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    assignment: dict[str, Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    factories: list[Factory],
    days: int,
    max_rounds: int = 1,
    production_rows: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    rows = effective_shipments(refresh_transport_rows([dict(row) for row in shipments], routes, products, cargo, carriers), days)
    forecast_by_node = {row["node"]: row for row in forecasts}
    ratio = charge_ratio(products, cargo)

    for _ in range(max_rounds):
        replay = daily_market_replay(forecasts, rows, days)
        shortage_rows = [row for row in replay.get("daily_rows", []) if float(row.get("shortage") or 0.0) > 0]
        if not shortage_rows:
            break

        additions: dict[tuple[str, str, str, int], dict[str, Any]] = {}
        for shortage in shortage_rows:
            node = sv(shortage.get("node"))
            factory = assignment.get(node)
            if not factory:
                continue
            candidates = [route for route in routes if route.src == factory.name and route.dst == node]
            if not candidates:
                continue
            amount = ceil_int(shortage.get("shortage") or 0)
            route = min(
                candidates,
                key=lambda item: (
                    route_effective_lead(item, max(amount, 1.0), ratio, carriers),
                    route_unit_cost(item, max(amount, 1.0), ratio, carriers),
                ),
            )
            demand_day = int(shortage.get("day") or 1)
            lead = route_effective_lead(route, max(amount, 1.0), ratio, carriers)
            ship_day = max(1, min(days, demand_day - lead))
            key = (factory.name, node, route.route, ship_day)
            if key not in additions:
                additions[key] = {
                    "destination": node,
                    "source": factory.name,
                    "factory": factory.name,
                    "cargo": cargo,
                    "amount": 0,
                    "route": route.route,
                    "mode": route_mode(route),
                    "carrier": route_carrier(route, carriers, amount, ratio),
                    "lead": lead,
                    "ship_day": ship_day,
                    "arrival_day": ship_day + lead,
                    "freight_cost": 0.0,
                    "note": "断货桥接补货",
                    "_route_obj": route,
                }
            additions[key]["amount"] += amount

        if not additions:
            break

        for row in additions.values():
            route = row.pop("_route_obj")
            amount = ceil_int(row["amount"])
            row["amount"] = amount
            refresh_shipment_route_fields(row, route, products, cargo, carriers)
            rows.append(row)

        rows = effective_shipments(refresh_transport_rows(repair_factory_ship_days(rows, factories, days, production_rows), routes, products, cargo, carriers), days)

    return rows


def fill_total_market_gap_fast(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    assignment: dict[str, Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    factories: list[Factory],
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    rows = effective_shipments(refresh_transport_rows([dict(row) for row in shipments], routes, products, cargo, carriers), days)
    replay = daily_market_replay(forecasts, rows, days)
    shortage_by_node = defaultdict(int)
    for day_row in replay.get("daily_rows", []):
        shortage = ceil_int(day_row.get("shortage") or 0)
        if shortage > 0:
            shortage_by_node[sv(day_row.get("node"))] += shortage
    if not shortage_by_node:
        return rows

    ratio = charge_ratio(products, cargo)
    shipped_by_factory = defaultdict(int)
    for row in rows:
        shipped_by_factory[sv(row.get("source") or row.get("factory"))] += int(row.get("amount") or 0)
    output_by_factory = factory_output_schedule(factories, days, production_rows)
    available_by_factory = {
        factory.name: max(0, int(math.floor(float(factory.init or 0.0) + sum(output_by_factory.get(factory.name, [0] * days)) - shipped_by_factory.get(factory.name, 0))))
        for factory in factories
    }

    additions: list[dict[str, Any]] = []
    for node, shortage in sorted(shortage_by_node.items(), key=lambda item: item[1], reverse=True):
        remaining = int(shortage)
        ranked: list[tuple[float, int, Factory, Route]] = []
        preferred = assignment.get(node)
        factory_order = ([preferred] if preferred else []) + [factory for factory in factories if factory is not preferred]
        for factory in factory_order:
            if not factory or available_by_factory.get(factory.name, 0) <= 0:
                continue
            for route in lane_routes(routes, factory.name, node):
                reference = max(1, min(remaining, available_by_factory.get(factory.name, 0)))
                ranked.append((route_score(route, reference, ratio, urgent=True, carriers=carriers), route_effective_lead(route, reference, ratio, carriers), factory, route))
        for _score, _lead_rank, factory, route in sorted(ranked, key=lambda item: (item[0], item[1], item[2].name)):
            if remaining <= 0:
                break
            available = available_by_factory.get(factory.name, 0)
            amount = int(min(remaining, available))
            if amount <= 0:
                continue
            lead = route_effective_lead(route, amount, ratio, carriers)
            ship_day = 1
            row = {
                "destination": node,
                "source": factory.name,
                "factory": factory.name,
                "cargo": cargo,
                "amount": amount,
                "route": route.route,
                "mode": route_mode(route),
                "carrier": route_carrier(route, carriers, amount, ratio),
                "lead": lead,
                "ship_day": ship_day,
                "arrival_day": ship_day + lead,
                "freight_cost": route_cost(route, amount, ratio, carriers),
                "note": "总缺口快速补齐",
            }
            additions.append(row)
            available_by_factory[factory.name] = available - amount
            remaining -= amount
    if not additions:
        return rows
    rows.extend(additions)
    return effective_shipments(refresh_transport_rows(repair_factory_ship_days(rows, factories, days, production_rows), routes, products, cargo, carriers), days)


def fill_early_shortages_from_any_factory(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    factories: list[Factory],
    days: int,
    max_rounds: int = 8,
    production_rows: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    rows = effective_shipments(refresh_transport_rows([dict(row) for row in shipments], routes, products, cargo, carriers), days)
    if not forecasts or not factories:
        return rows
    ratio = charge_ratio(products, cargo)
    output_by_factory = factory_output_schedule(factories, days, production_rows)

    def available(factory_name: str, ship_day: int, current_rows: list[dict[str, Any]]) -> int:
        factory = next((item for item in factories if item.name == factory_name), None)
        if not factory:
            return 0
        produced = float(factory.init or 0.0) + sum(output_by_factory.get(factory_name, [0] * days)[:ship_day])
        shipped = sum(
            float(row.get("amount") or 0.0)
            for row in current_rows
            if sv(row.get("factory") or row.get("source")) == factory_name
            and int(row.get("ship_day") or 1) <= ship_day
        )
        return max(0, int(math.floor(produced - shipped)))

    for _ in range(max_rounds):
        replay = daily_market_replay(forecasts, rows, days)
        shortage_rows = [row for row in replay.get("daily_rows", []) if int(row.get("shortage") or 0) > 0]
        if not shortage_rows:
            break
        changed = False
        for shortage_row in sorted(shortage_rows, key=lambda row: (int(row.get("day") or 1), -int(row.get("shortage") or 0))):
            node = sv(shortage_row.get("node"))
            shortage = int(shortage_row.get("shortage") or 0)
            shortage_day = int(shortage_row.get("day") or 1)
            if shortage <= 0:
                continue
            candidates: list[tuple[int, float, Factory, Route, int]] = []
            for factory in factories:
                for route in lane_routes(routes, factory.name, node):
                    lead = route_effective_lead(route, shortage, ratio, carriers)
                    latest_ship_day = shortage_day - lead
                    if latest_ship_day < 1:
                        continue
                    amount_avail = available(factory.name, latest_ship_day, rows)
                    if amount_avail <= 0:
                        continue
                    reference = max(1, min(shortage, amount_avail))
                    candidates.append((lead, route_unit_cost(route, reference, ratio, carriers), factory, route, latest_ship_day))
            if not candidates:
                continue
            lead, _unit, factory, route, ship_day = min(candidates, key=lambda item: (item[0], item[1], item[2].name))
            amount = min(shortage, available(factory.name, ship_day, rows))
            if amount <= 0:
                continue
            rows.append(
                {
                    "destination": node,
                    "source": factory.name,
                    "factory": factory.name,
                    "cargo": cargo,
                    "amount": int(amount),
                    "route": route.route,
                    "mode": route_mode(route),
                    "carrier": route_carrier(route, carriers, amount, ratio),
                    "lead": lead,
                    "ship_day": int(ship_day),
                    "arrival_day": int(ship_day) + lead,
                    "freight_cost": route_cost(route, amount, ratio, carriers),
                    "note": "早期断货任意工厂快线补货",
                }
            )
            rows = effective_shipments(refresh_transport_rows(repair_factory_ship_days(rows, factories, days, production_rows), routes, products, cargo, carriers), days)
            changed = True
            break
        if not changed:
            break
    return sorted(rows, key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("route")), sv(row.get("destination"))))


def trim_expensive_shipments_for_score(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    xls_path: Path,
    qtype: str,
    days: int,
) -> list[dict[str, Any]]:
    if qtype not in {"生产", "销售"} or not shipments or not forecasts:
        return shipments
    if qtype == "生产":
        return shipments
    if qtype == "销售" and days > 45:
        return shipments
    targets = score_targets(xls_path, qtype)
    points = score_points(xls_path, qtype)
    target_logistics = float(targets.get("unit_logistics_cost", 0.0) or 0.0)
    if target_logistics <= 0:
        return shipments

    def plan_score(rows: list[dict[str, Any]]) -> dict[str, float]:
        replay = daily_market_replay(forecasts, rows, days)
        freight = sum(float(row.get("freight_cost") or 0.0) for row in rows)
        if qtype == "销售":
            denominator = replay["served"]
            score = deviation_score(0.0, targets["prediction_deviation"], points["预测偏差率"])
            score += cost_score(freight / max(denominator, 0.001), target_logistics, points["单位物流成本"])
            score += satisfaction_score(replay["market_satisfaction"], points["市场满足率"])
        else:
            denominator = sum(float(row.get("amount") or 0.0) for row in rows)
            score = cost_score(freight / max(denominator, 0.001), target_logistics, points["单位物流成本"])
            score += satisfaction_score(replay["market_satisfaction"], points["市场满足率"])
        unit = freight / max(denominator, 0.001)
        return {
            "score": float(score),
            "unit_logistics": float(unit),
            "market_satisfaction": float(replay["market_satisfaction"]),
            "shortage": float(replay["shortage"]),
        }

    rows = [dict(row) for row in shipments]
    best_metrics = plan_score(rows)
    for _ in range(min(200, len(rows))):
        best_candidate: tuple[tuple[float, float, float, float], int, list[dict[str, Any]], dict[str, float]] | None = None
        for idx, row in enumerate(rows):
            amount = float(row.get("amount") or 0.0)
            if amount <= 0:
                continue
            candidate = rows[:idx] + rows[idx + 1 :]
            metrics = plan_score(candidate)
            priority = product_transport_priority(metrics, len(candidate))
            if priority > product_transport_priority(best_metrics, len(rows)):
                unit_cost = float(row.get("freight_cost") or 0.0) / max(amount, 0.001)
                item = (priority, int(unit_cost * 1000), candidate, metrics)
                if best_candidate is None or item[:2] > best_candidate[:2]:
                    best_candidate = item
        if best_candidate is None:
            break
        rows = best_candidate[2]
        best_metrics = best_candidate[3]
    return rows


def advance_shortage_shipments_by_faster_route(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    xls_path: Path,
    qtype: str,
    days: int,
    max_rounds: int = 12,
    production_rows: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    if qtype not in {"生产", "销售"} or not forecasts or not shipments:
        return shipments
    ratio = charge_ratio(products, cargo)
    rows = effective_shipments(refresh_transport_rows([dict(row) for row in shipments], routes, products, cargo, carriers), days)

    def candidate_metrics(candidate_rows: list[dict[str, Any]]) -> dict[str, float]:
        return score_product_transport_candidate(
            forecasts=forecasts,
            shipments=candidate_rows,
            xls_path=xls_path,
            qtype=qtype,
            days=days,
        )

    current = candidate_metrics(rows)
    for _ in range(max_rounds):
        replay = daily_market_replay(forecasts, rows, days)
        shortage_rows = [row for row in replay.get("daily_rows", []) if int(row.get("shortage") or 0) > 0]
        if not shortage_rows:
            break
        best: tuple[tuple[float, float, float], list[dict[str, Any]], dict[str, float]] | None = None
        for shortage in sorted(shortage_rows, key=lambda row: (int(row.get("day") or 1), -int(row.get("shortage") or 0))):
            node = sv(shortage.get("node"))
            shortage_day = int(shortage.get("day") or 1)
            shortage_amount = int(shortage.get("shortage") or 0)
            if shortage_amount <= 0:
                continue
            for idx, late_row in enumerate(rows):
                if sv(late_row.get("destination") or late_row.get("node") or late_row.get("dst")) != node:
                    continue
                original_amount = int(late_row.get("amount") or 0)
                if original_amount <= 0:
                    continue
                original_arrival = int(late_row.get("arrival_day") or (int(late_row.get("ship_day") or 1) + int(late_row.get("lead") or 0)))
                if original_arrival <= shortage_day:
                    continue
                source = sv(late_row.get("source") or late_row.get("factory"))
                ship_day = int(late_row.get("ship_day") or 1)
                if not source:
                    continue
                for alt_route in lane_routes(routes, source, node):
                    if alt_route.route == sv(late_row.get("route")):
                        continue
                    reference = max(1, min(original_amount, shortage_amount))
                    alt_lead = route_effective_lead(alt_route, reference, ratio, carriers)
                    latest_alt_ship_day = shortage_day - alt_lead
                    if latest_alt_ship_day < 1:
                        continue
                    ship_day_options = sorted(
                        {
                            max(1, min(days, latest_alt_ship_day)),
                            max(1, min(days, ship_day)),
                        },
                        reverse=True,
                    )
                    qty_options = sorted(
                        {
                            min(original_amount, shortage_amount),
                            min(original_amount, max(1, shortage_amount // 2)),
                            min(original_amount, max(1, ceil_int(shortage_amount * 1.1))),
                        },
                        reverse=True,
                    )
                    for alt_ship_day in ship_day_options:
                        if alt_ship_day + alt_lead > shortage_day:
                            continue
                        for qty in qty_options:
                            if qty <= 0:
                                continue
                            candidate_rows = [dict(row) for row in rows]
                            original_route = route_for_shipment(candidate_rows[idx], routes)
                            remaining = original_amount - qty
                            if remaining > 0:
                                candidate_rows[idx]["amount"] = remaining
                                if original_route:
                                    refresh_shipment_route_fields(candidate_rows[idx], original_route, products, cargo, carriers)
                            else:
                                candidate_rows.pop(idx)
                            new_row = {
                                "destination": node,
                                "source": source,
                                "factory": source,
                                "cargo": cargo,
                                "amount": int(qty),
                                "route": alt_route.route,
                                "mode": route_mode(alt_route),
                                "carrier": route_carrier(alt_route, carriers, qty, ratio),
                                "lead": alt_lead,
                                "ship_day": int(alt_ship_day),
                                "arrival_day": int(alt_ship_day) + alt_lead,
                                "freight_cost": route_cost(alt_route, qty, ratio, carriers),
                                "note": "早期断货提前改走更快路线",
                            }
                            candidate_rows.append(new_row)
                            candidate_rows = sanitize_product_transport(
                                forecasts=forecasts,
                                shipments=candidate_rows,
                                factories=factories,
                                routes=routes,
                                products=products,
                                cargo=cargo,
                                carriers=carriers,
                                days=days,
                                production_rows=production_rows,
                            )
                            hard_risks = product_transport_hard_risks(
                                forecasts=forecasts,
                                shipments=candidate_rows,
                                factories=factories,
                                days=days,
                                production_rows=production_rows,
                            )
                            if hard_risks:
                                continue
                            metrics = candidate_metrics(candidate_rows)
                            key = product_transport_priority(metrics, len(candidate_rows))
                            if key > product_transport_priority(current, len(rows)) and (best is None or key > best[0]):
                                best = (key, candidate_rows, metrics)
        if best is None:
            break
        rows = best[1]
        current = best[2]
    return sorted(rows, key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("route")), sv(row.get("destination"))))


def lane_routes(routes: list[Route], src: str, dst: str) -> list[Route]:
    exact = [route for route in routes if route.src == src and route.dst == dst]
    if exact:
        return exact
    return [route for route in routes if src in route.route and dst in route.route]


def pick_bulk_route(
    routes: list[Route],
    src: str,
    dst: str,
    amount: float,
    charge_ratio_value: float,
    *,
    urgent: bool = False,
    carriers: list[Any] | None = None,
) -> Route | None:
    candidates = lane_routes(routes, src, dst)
    if not candidates:
        return None
    reference_amount = max(ceil_int(amount), 1)
    return min(
        candidates,
        key=lambda route: (
            route_unit_cost(route, reference_amount, charge_ratio_value, carriers),
            route_effective_lead(route, reference_amount, charge_ratio_value, carriers) if urgent else 0,
            route_score(route, reference_amount, charge_ratio_value, urgent, carriers),
        ),
    )


def economic_lot(route: Route, charge_ratio_value: float, avg_daily: float, carriers: list[Any] | None = None) -> int:
    option = best_transport_option(route, max(avg_daily, 1.0), charge_ratio_value, carriers)
    lot = transport_option_economic_lot(option)
    if avg_daily > 0:
        lot = max(lot, ceil_int(avg_daily))
    return lot


def split_integer_lots(amount: int, trips: int) -> list[int]:
    trips = max(1, min(trips, amount)) if amount > 0 else 1
    base = amount // trips
    lots = [base for _ in range(trips)]
    lots[-1] += amount - base * trips
    return [lot for lot in lots if lot > 0]


def periodic_ship_days(
    *,
    trips: int,
    route: Route,
    days: int,
    init_cover_days: int = 0,
    lead: int | None = None,
) -> list[int]:
    lead_days = int(route.lead if lead is None else lead)
    latest_ship_day = max(1, days - lead_days)
    first_arrival_day = max(1, min(days, init_cover_days + 1))
    first_ship_day = max(1, min(latest_ship_day, first_arrival_day - lead_days))
    horizon = max(1, latest_ship_day - first_ship_day + 1)
    trips = max(1, min(trips, horizon))
    if trips == 1:
        return [first_ship_day]
    interval = max(1, math.floor((latest_ship_day - first_ship_day) / max(trips - 1, 1)))
    ship_days = [min(latest_ship_day, first_ship_day + idx * interval) for idx in range(trips)]
    # If the final cap collapsed days together, spread from the end backwards.
    if len(set(ship_days)) < len(ship_days):
        ship_days = list(range(max(1, latest_ship_day - trips + 1), latest_ship_day + 1))
    return ship_days


def build_periodic_transport_rows(
    *,
    source: str,
    destination: str,
    cargo: str,
    amount: float,
    route: Route,
    products: list[Product],
    carriers: list[Any] | None,
    days: int,
    initial_cover: float = 0.0,
    note: str = "快速整数周期运输",
) -> list[dict[str, Any]]:
    amount_int = ceil_int(amount)
    if amount_int <= 0:
        return []
    ratio = charge_ratio(products, cargo)
    avg_daily = amount_int / max(days, 1)
    lot = economic_lot(route, ratio, avg_daily, carriers)
    max_cycle = 5 if days <= 30 else 7
    cycle_days = max(1, min(max_cycle, ceil_int(lot / max(avg_daily, 1.0))))
    target_lot = max(lot, ceil_int(avg_daily * cycle_days))
    trips = max(1, ceil_int(amount_int / max(target_lot, 1)))
    while trips > 1 and amount_int // trips < max(1, min(lot, amount_int)):
        trips -= 1
    init_cover_days = int(math.floor(max(0.0, initial_cover) / max(avg_daily, 1.0)))
    preliminary_lots = split_integer_lots(amount_int, trips)
    rows: list[dict[str, Any]] = []
    lead_reference = route_effective_lead(route, max(preliminary_lots or [amount_int]), ratio, carriers)
    ship_days = periodic_ship_days(trips=trips, route=route, days=days, init_cover_days=init_cover_days, lead=lead_reference)
    lots = split_integer_lots(amount_int, len(ship_days))
    for ship_day, lot_amount in zip(ship_days, lots):
        lead = route_effective_lead(route, lot_amount, ratio, carriers)
        rows.append(
            {
                "cargo": cargo,
                "source": source,
                "factory": source,
                "destination": destination,
                "amount": int(lot_amount),
                "ship_day": int(ship_day),
                "arrival_day": int(ship_day) + lead,
                "route": route.route,
                "mode": route_mode(route),
                "lead": lead,
                "freight_cost": route_cost(route, lot_amount, ratio, carriers),
                "carrier": route_carrier(route, carriers, lot_amount, ratio),
                "note": note,
            }
        )
    return rows


def build_fast_product_transport(
    *,
    allocations: list[tuple[dict[str, Any], Factory, int, float, float]],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
) -> list[dict[str, Any]]:
    ratio = charge_ratio(products, cargo)
    rows: list[dict[str, Any]] = []
    for forecast, factory, allocated_amount, init_share, _demand_share in allocations:
        if allocated_amount <= 0:
            continue
        route = pick_bulk_route(
            routes,
            factory.name,
            forecast["node"],
            allocated_amount,
            ratio,
            urgent=False,
            carriers=carriers,
        )
        if not route:
            continue
        rows.extend(
            build_periodic_transport_rows(
                source=factory.name,
                destination=forecast["node"],
                cargo=cargo,
                amount=allocated_amount,
                route=route,
                products=products,
                carriers=carriers,
                days=days,
                initial_cover=init_share,
            )
        )
    return sorted(rows, key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("route")), sv(row.get("destination"))))


def build_safe_product_transport(
    *,
    forecasts: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    if not forecasts or not factories:
        return []
    ratio = charge_ratio(products, cargo)
    demand_by_node = {row["node"]: daily_demands_for_forecast(row, days) for row in forecasts}
    forecast_by_node = {row["node"]: row for row in forecasts}
    inventory = {row["node"]: float(row.get("init", 0.0) or 0.0) for row in forecasts}
    arrivals: dict[tuple[str, int], int] = defaultdict(int)
    shipped_by_factory_day: dict[tuple[str, int], int] = defaultdict(int)
    output_by_factory = factory_output_schedule(factories, days, production_rows)
    rows: list[dict[str, Any]] = []

    def factory_available(factory: Factory, day: int) -> int:
        produced = float(factory.init or 0.0) + sum(output_by_factory.get(factory.name, [0] * days)[:day])
        shipped = sum(qty for (factory_name, ship_day), qty in shipped_by_factory_day.items() if factory_name == factory.name and ship_day <= day)
        return max(0, int(math.floor(produced - shipped)))

    def projected_before_arrival(node: str, start_day: int, arrival_day: int) -> float:
        value = float(inventory.get(node, 0.0))
        demands = demand_by_node.get(node, [])
        for future_day in range(start_day + 1, arrival_day):
            value += arrivals.get((node, future_day), 0)
            if future_day - 1 < len(demands):
                value -= demands[future_day - 1]
        value += arrivals.get((node, arrival_day), 0)
        return value

    route_candidates: dict[str, list[tuple[float, int, Factory, Route]]] = {}
    for forecast in forecasts:
        node = forecast["node"]
        limit = float(forecast.get("limit") or 0.0)
        candidates: list[tuple[float, int, Factory, Route]] = []
        for factory in factories:
            for route in lane_routes(routes, factory.name, node):
                reference = max(1.0, min(max(float(forecast.get("daily_avg") or 1.0) * 4, 1.0), limit or 1e9))
                candidates.append((route_unit_cost(route, reference, ratio, carriers), route_effective_lead(route, reference, ratio, carriers), factory, route))
        route_candidates[node] = sorted(candidates, key=lambda item: (item[0], item[1], item[2].name))

    for day in range(1, days + 1):
        for forecast in forecasts:
            node = forecast["node"]
            inventory[node] = inventory.get(node, 0.0) + arrivals.get((node, day), 0)
            demand = demand_by_node.get(node, [0] * days)[day - 1]
            inventory[node] = max(0.0, inventory[node] - demand)

        for forecast in sorted(forecasts, key=lambda row: float(inventory.get(row["node"], 0.0)) / max(float(row.get("daily_avg") or 1.0), 1.0)):
            node = forecast["node"]
            limit = float(forecast.get("limit") or 0.0)
            avg_daily = max(float(forecast.get("daily_avg") or 0.0), 1.0)
            candidates = route_candidates.get(node, [])
            if not candidates:
                continue
            for _unit, _lead_rank, factory, route in candidates:
                reference_amount = max(1.0, avg_daily * 4)
                lead = route_effective_lead(route, reference_amount, ratio, carriers)
                arrival_day = day + lead
                if arrival_day > days:
                    continue
                projected = projected_before_arrival(node, day, arrival_day)
                target_inventory = min(limit if limit > 0 else avg_daily * 4, max(1.0, avg_daily * 4))
                if projected >= target_inventory - avg_daily * 0.5:
                    continue
                cap_at_arrival = (limit - projected) if limit > 0 else target_inventory - projected
                if cap_at_arrival < 1.0:
                    continue
                available = factory_available(factory, day)
                if available < 1:
                    continue
                amount = int(math.floor(min(cap_at_arrival, available, max(target_inventory - projected, 1.0))))
                if amount <= 0:
                    continue
                rows.append(
                    {
                        "destination": node,
                        "source": factory.name,
                        "factory": factory.name,
                        "cargo": cargo,
                        "amount": amount,
                        "route": route.route,
                        "mode": route_mode(route),
                        "carrier": route_carrier(route, carriers, amount, ratio),
                        "lead": lead,
                        "ship_day": day,
                        "arrival_day": arrival_day,
                        "freight_cost": route_cost(route, amount, ratio, carriers),
                        "note": "安全库存滚动补货",
                    }
                )
                arrivals[(node, arrival_day)] += amount
                shipped_by_factory_day[(factory.name, day)] += amount
                break

    return sorted(rows, key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("route")), sv(row.get("destination"))))


def build_service_first_product_transport(
    *,
    forecasts: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
    cover_days: int = 4,
) -> list[dict[str, Any]]:
    if not forecasts or not factories:
        return []
    ratio = charge_ratio(products, cargo)
    output_by_factory = factory_output_schedule(factories, days, production_rows)
    factory_shipments_by_day: dict[tuple[str, int], int] = defaultdict(int)
    destination_arrivals: dict[tuple[str, int], int] = defaultdict(int)
    rows: list[dict[str, Any]] = []

    def factory_available(factory: Factory, ship_day: int) -> int:
        produced = float(factory.init or 0.0) + sum(output_by_factory.get(factory.name, [0] * days)[:ship_day])
        shipped = sum(
            amount
            for (factory_name, day), amount in factory_shipments_by_day.items()
            if factory_name == factory.name and day <= ship_day
        )
        return max(0, int(math.floor(produced - shipped)))

    for forecast in forecasts:
        node = forecast["node"]
        demands = daily_demands_for_forecast(forecast, days)
        limit = int(forecast.get("limit") or 0)
        inventory = int(round(float(forecast.get("init") or 0.0)))
        for day in range(1, days + 1):
            inventory += destination_arrivals.get((node, day), 0)
            today_demand = int(demands[day - 1] if day - 1 < len(demands) else 0)
            if inventory >= today_demand:
                inventory -= today_demand
                continue

            target_cover = sum(int(value) for value in demands[day - 1 : min(days, day - 1 + max(1, cover_days))])
            if limit > 0:
                target_cover = min(target_cover, limit)
            required = max(today_demand - inventory, target_cover - inventory)
            remaining = int(required)
            candidates: list[tuple[int, float, Factory, Route, int]] = []
            for factory in factories:
                for route in lane_routes(routes, factory.name, node):
                    reference = max(remaining, 1)
                    lead = route_effective_lead(route, reference, ratio, carriers)
                    ship_day = day - lead
                    if ship_day < 1:
                        continue
                    available = factory_available(factory, ship_day)
                    if available <= 0:
                        continue
                    candidates.append((lead, route_unit_cost(route, min(reference, available), ratio, carriers), factory, route, ship_day))
            for lead, _unit, factory, route, ship_day in sorted(candidates, key=lambda item: (item[1], item[0], item[2].name)):
                if remaining <= 0:
                    break
                available = factory_available(factory, ship_day)
                amount = min(remaining, available)
                if amount <= 0:
                    continue
                row = {
                    "destination": node,
                    "source": factory.name,
                    "factory": factory.name,
                    "cargo": cargo,
                    "amount": int(amount),
                    "route": route.route,
                    "mode": route_mode(route),
                    "carrier": route_carrier(route, carriers, amount, ratio),
                    "lead": lead,
                    "ship_day": int(ship_day),
                    "arrival_day": int(day),
                    "freight_cost": route_cost(route, amount, ratio, carriers),
                    "note": "满足率优先滚动补货",
                }
                rows.append(row)
                factory_shipments_by_day[(factory.name, int(ship_day))] += int(amount)
                destination_arrivals[(node, int(day))] += int(amount)
                inventory += int(amount)
                remaining -= int(amount)
            inventory = max(0, inventory - today_demand)

    return sorted(effective_shipments(refresh_transport_rows(rows, routes, products, cargo, carriers), days), key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("route")), sv(row.get("destination"))))


def build_global_product_transport(
    *,
    forecasts: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    name: str,
    production_rows: list[dict[str, Any]] | None = None,
    warm_start_shipments: list[dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not forecasts or not factories:
        return [], {"method": "GlobalIntegerTransport", "status": "NoDemandOrSource"}
    output_by_factory = factory_output_schedule(factories, days, production_rows)
    sources = {
        factory.name: {
            "initial": factory.init,
            "supply": output_by_factory.get(factory.name, [0] * days),
            "supply_is_capacity": production_rows is None,
            "limit": 0.0,
        }
        for factory in factories
    }
    destinations = {
        row["node"]: {
            "initial": row["init"],
            "demand": daily_demands_for_forecast(row, days),
            "limit": row["limit"],
            "excess_fee": row.get("excess_fee", 0.0),
        }
        for row in forecasts
    }
    usable_routes = [route for route in routes if route.src in sources and route.dst in destinations]
    if not usable_routes:
        return [], {"method": "GlobalIntegerTransport", "status": "NoRoute"}
    transport = _solve_day_transport_milp(
        name=name,
        sources=sources,
        destinations=destinations,
        routes=usable_routes,
        products=products,
        cargo=cargo,
        carriers=carriers,
        days=days,
        ship_day_step=7 if days > 45 else 1,
        gap_rel=0.02 if days > 45 else 0.01,
        enforce_destination_limits=True,
        time_limit_sec=30 if days <= 45 and not is_extreme_mode() else None,
    )
    rows = [
        {
            "destination": row["destination"],
            "factory": row["source"],
            "source": row["source"],
            "cargo": row["cargo"],
            "amount": int(row["amount"]),
            "route": row["route"],
            "mode": row["mode"],
            "carrier": row.get("carrier", ""),
            "lead": int(row["lead"]),
            "ship_day": int(row["ship_day"]),
            "arrival_day": int(row["arrival_day"]),
            "freight_cost": float(row["freight_cost"]),
            "note": "全局逐日整数运输优化",
        }
        for row in transport.get("shipments", [])
    ]
    return rows, {
        "method": "GlobalIntegerTransport",
        "status": transport.get("status"),
        "shortage": transport.get("shortage", 0.0),
        "failures": transport.get("failures", []),
    }


def build_warm_window_global_product_transport(
    *,
    forecasts: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    name: str,
    warm_start_shipments: list[dict[str, Any]],
    production_rows: list[dict[str, Any]] | None = None,
    window: int = 2,
    time_limit_sec: int = 240,
    strict_freight_upper_bound: float | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not warm_start_shipments or not forecasts or not factories:
        return [], {"method": "WarmWindowGlobalIntegerTransport", "status": "NoWarmStart"}
    output_by_factory = factory_output_schedule(factories, days, production_rows)
    sources = {
        factory.name: {
            "initial": factory.init,
            "supply": output_by_factory.get(factory.name, [0] * days),
            "supply_is_capacity": production_rows is None,
            "limit": 0.0,
        }
        for factory in factories
    }
    destinations = {
        row["node"]: {
            "initial": row["init"],
            "demand": daily_demands_for_forecast(row, days),
            "limit": row["limit"],
            "excess_fee": row.get("excess_fee", 0.0),
        }
        for row in forecasts
    }
    usable_routes = [
        route
        for route in pruned_product_routes(
            forecasts=forecasts,
            factories=factories,
            routes=routes,
            products=products,
            cargo=cargo,
            carriers=carriers,
        )
        if route.src in sources and route.dst in destinations
    ]
    if not usable_routes:
        return [], {"method": "WarmWindowGlobalIntegerTransport", "status": "NoRoute"}

    ratio = charge_ratio(products, cargo)
    allowed_days: dict[tuple[str, str, str], set[int]] = defaultdict(set)
    allowed_option_days: dict[tuple[str, str, str, str], set[int]] = defaultdict(set)
    routes_by_destination: dict[str, list[Route]] = defaultdict(list)
    for route in usable_routes:
        routes_by_destination[route.dst].append(route)

    for row in warm_start_shipments:
        destination = sv(row.get("destination"))
        amount = max(1, int(row.get("amount") or 0))
        arrival_day = int(row.get("arrival_day") or (int(row.get("ship_day") or 1) + int(row.get("lead") or 0)))
        for route in routes_by_destination.get(destination, []):
            route_key = (route.route, route.src, route.dst)
            for option in route_transport_options(route, ratio, carriers, max_amount=max(amount, 1), prune_dominated=True):
                lead = int(option.lead or 0)
                center = max(1, min(days, arrival_day - lead))
                option_key = (route.route, route.src, route.dst, option.carrier)
                for day in range(center - window, center + window + 1):
                    if 1 <= day <= days - lead:
                        allowed_days[route_key].add(int(day))
                        allowed_option_days[option_key].add(int(day))
        route_key = (sv(row.get("route")), sv(row.get("source") or row.get("factory")), destination)
        option_key = (sv(row.get("route")), sv(row.get("source") or row.get("factory")), destination, sv(row.get("carrier")))
        ship_day = int(row.get("ship_day") or 1)
        for day in range(ship_day - window, ship_day + window + 1):
            if 1 <= day <= days:
                allowed_days[route_key].add(int(day))
                if option_key[3]:
                    allowed_option_days[option_key].add(int(day))

    transport = _solve_day_transport_milp(
        name=name,
        sources=sources,
        destinations=destinations,
        routes=usable_routes,
        products=products,
        cargo=cargo,
        carriers=carriers,
        days=days,
        ship_day_step=1,
        gap_rel=0.001,
        enforce_destination_limits=True,
        max_total_shortage=0.0,
        time_limit_sec=time_limit_sec,
        warm_start_shipments=warm_start_shipments,
        allowed_ship_days_by_route=allowed_days,
        allowed_ship_days_by_option=allowed_option_days,
        strict_freight_upper_bound=strict_freight_upper_bound,
        allow_fallback=strict_freight_upper_bound is None,
        strict_bound_feasibility_only=strict_freight_upper_bound is not None,
    )
    transport_status = transport.get("status")
    if strict_freight_upper_bound is not None and transport_status == "Infeasible":
        transport_status = "CounterexampleInfeasible"
    rows = [
        {
            "destination": row["destination"],
            "factory": row["source"],
            "source": row["source"],
            "cargo": row["cargo"],
            "amount": int(row["amount"]),
            "route": row["route"],
            "mode": row["mode"],
            "carrier": row.get("carrier", ""),
            "lead": int(row["lead"]),
            "ship_day": int(row["ship_day"]),
            "arrival_day": int(row["arrival_day"]),
            "freight_cost": float(row["freight_cost"]),
            "note": "暖启动窗口全局整数运输优化",
        }
        for row in transport.get("shipments", [])
    ]
    return rows, {
        "method": "WarmWindowGlobalIntegerTransport",
        "status": transport_status,
        "shortage": transport.get("shortage", 0.0),
        "failures": transport.get("failures", []),
        "solve_seconds": transport.get("solve_seconds"),
        "objective": transport.get("objective"),
        "allowed_route_windows": len(allowed_days),
        "allowed_option_windows": len(allowed_option_days),
        "model_stats": transport.get("model_stats", {}),
        "raw_status": transport.get("raw_status"),
        "time_limit_hit": transport.get("time_limit_hit"),
        "strict_freight_upper_bound": strict_freight_upper_bound,
    }


def optimize_destination_subset_transport(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    target_nodes: set[str],
    production_rows: list[dict[str, Any]] | None = None,
    time_limit_sec: int = 120,
    ship_day_window: int | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    clean_targets = {sv(node) for node in target_nodes if sv(node)}
    if not clean_targets:
        return [], {"method": "DestinationSubsetMILP", "status": "NoTarget"}
    fixed_rows = [dict(row) for row in shipments if sv(row.get("destination")) not in clean_targets]
    target_rows = [dict(row) for row in shipments if sv(row.get("destination")) in clean_targets]
    if not target_rows:
        return [], {"method": "DestinationSubsetMILP", "status": "NoTargetShipment"}

    output_by_factory = factory_output_schedule(factories, days, production_rows)
    fixed_daily_usage: dict[str, list[int]] = {factory.name: [0] * days for factory in factories}
    fixed_total_usage: dict[str, int] = defaultdict(int)
    for row in fixed_rows:
        factory_name = sv(row.get("source") or row.get("factory"))
        if factory_name not in fixed_daily_usage:
            continue
        ship_day = max(1, min(days, int(row.get("ship_day") or 1)))
        amount = int(row.get("amount") or 0)
        fixed_daily_usage[factory_name][ship_day - 1] += amount
        fixed_total_usage[factory_name] += amount

    sources: dict[str, dict[str, Any]] = {}
    for factory in factories:
        remaining_supply = [
            float(output_by_factory.get(factory.name, [0] * days)[day]) - float(fixed_daily_usage.get(factory.name, [0] * days)[day])
            for day in range(days)
        ]
        sources[factory.name] = {
            "initial": factory.init,
            "supply": remaining_supply,
            "supply_is_capacity": False,
            "limit": 0.0,
        }

    forecast_by_node = {sv(row.get("node")): row for row in forecasts}
    destinations = {
        node: {
            "initial": forecast_by_node[node]["init"],
            "demand": daily_demands_for_forecast(forecast_by_node[node], days),
            "limit": forecast_by_node[node]["limit"],
            "excess_fee": forecast_by_node[node].get("excess_fee", 0.0),
        }
        for node in clean_targets
        if node in forecast_by_node
    }
    if not destinations:
        return [], {"method": "DestinationSubsetMILP", "status": "NoTargetForecast"}

    usable_routes = [
        route
        for route in routes
        if route.src in sources and route.dst in destinations
    ]
    if not usable_routes:
        return [], {"method": "DestinationSubsetMILP", "status": "NoRoute"}

    ratio = charge_ratio(products, cargo)
    allowed_days: dict[tuple[str, str, str], set[int]] | None = None
    allowed_option_days: dict[tuple[str, str, str, str], set[int]] | None = None
    if ship_day_window is not None and int(ship_day_window) >= 0:
        window = max(0, int(ship_day_window))
        allowed_days = defaultdict(set)
        allowed_option_days = defaultdict(set)
        routes_by_destination: dict[str, list[Route]] = defaultdict(list)
        for route in usable_routes:
            routes_by_destination[route.dst].append(route)
        for row in target_rows:
            destination = sv(row.get("destination"))
            amount = max(1, int(row.get("amount") or 0))
            arrival_day = int(row.get("arrival_day") or (int(row.get("ship_day") or 1) + int(row.get("lead") or 0)))
            for route in routes_by_destination.get(destination, []):
                route_key = (route.route, route.src, route.dst)
                for option in route_transport_options(route, ratio, carriers, max_amount=max(amount, 1), prune_dominated=True):
                    lead = int(option.lead or 0)
                    center = max(1, min(days, arrival_day - lead))
                    option_key = (route.route, route.src, route.dst, option.carrier)
                    for day in range(center - window, center + window + 1):
                        if 1 <= day <= days - lead:
                            allowed_days[route_key].add(int(day))
                            allowed_option_days[option_key].add(int(day))
            route_key = (sv(row.get("route")), sv(row.get("source") or row.get("factory")), destination)
            option_key = (sv(row.get("route")), sv(row.get("source") or row.get("factory")), destination, sv(row.get("carrier")))
            ship_day = int(row.get("ship_day") or 1)
            for day in range(ship_day - window, ship_day + window + 1):
                if 1 <= day <= days:
                    allowed_days[route_key].add(int(day))
                    if option_key[3]:
                        allowed_option_days[option_key].add(int(day))

    transport = _solve_day_transport_milp(
        name=f"{cargo}_destination_subset_{safe_stem('_'.join(sorted(clean_targets)))}",
        sources=sources,
        destinations=destinations,
        routes=usable_routes,
        products=products,
        cargo=cargo,
        carriers=carriers,
        days=days,
        ship_day_step=1,
        gap_rel=0.0,
        enforce_destination_limits=True,
        max_total_shortage=0.0,
        allow_fallback=False,
        time_limit_sec=time_limit_sec,
        warm_start_shipments=target_rows,
        allowed_ship_days_by_route=allowed_days,
        allowed_ship_days_by_option=allowed_option_days,
        relax_integrality=False,
        prune_transport_options=True,
    )
    if transport.get("status") not in {"Optimal", "TimeLimitFeasible"}:
        return [], {
            "method": "DestinationSubsetMILP",
            "status": transport.get("status"),
            "raw_status": transport.get("raw_status"),
            "solve_seconds": transport.get("solve_seconds"),
            "time_limit_hit": transport.get("time_limit_hit"),
            "model_stats": transport.get("model_stats", {}),
            "ship_day_window": ship_day_window,
            "failures": transport.get("failures", []),
        }

    replacement = [
        {
            "destination": row["destination"],
            "factory": row["source"],
            "source": row["source"],
            "cargo": row["cargo"],
            "amount": int(row["amount"]),
            "route": row["route"],
            "mode": row["mode"],
            "carrier": row.get("carrier", ""),
            "lead": int(row["lead"]),
            "ship_day": int(row["ship_day"]),
            "arrival_day": int(row["arrival_day"]),
            "freight_cost": float(row["freight_cost"]),
            "note": "目的地子集全局整数优化",
        }
        for row in transport.get("shipments", [])
    ]
    candidate = sorted(
        effective_shipments(fixed_rows + replacement, days),
        key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("destination")), sv(row.get("route"))),
    )
    old_cost = sum(float(row.get("freight_cost") or 0.0) for row in shipments)
    new_cost = sum(float(row.get("freight_cost") or 0.0) for row in candidate)
    replay = daily_market_replay(forecasts, candidate, days)
    hard_risks = product_transport_hard_risks(
        forecasts=forecasts,
        shipments=candidate,
        factories=factories,
        days=days,
        production_rows=production_rows,
    )
    status = {
        "method": "DestinationSubsetMILP",
        "status": transport.get("status"),
        "targets": sorted(clean_targets),
        "delta": round(old_cost - new_cost, 4),
        "freight_cost": round(new_cost, 4),
        "old_freight_cost": round(old_cost, 4),
        "shortage": replay.get("shortage", 0.0),
        "market_satisfaction": replay.get("market_satisfaction", 0.0),
        "hard_risks": hard_risks,
        "solve_seconds": transport.get("solve_seconds"),
        "time_limit_hit": transport.get("time_limit_hit"),
        "model_stats": transport.get("model_stats", {}),
        "ship_day_window": ship_day_window,
    }
    if not is_full_satisfaction(float(replay.get("market_satisfaction", 0.0)), float(replay.get("shortage", 0.0))) or hard_risks:
        return [], {**status, "status": f"{transport.get('status')}RejectedByReplay"}
    return candidate, status


def build_full_global_audit_product_transport(
    *,
    forecasts: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    name: str,
    warm_start_shipments: list[dict[str, Any]],
    production_rows: list[dict[str, Any]] | None = None,
    time_limit_sec: int = 300,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not warm_start_shipments or not forecasts or not factories:
        return [], {"method": "FullGlobalIntegerAudit", "status": "NoWarmStart"}
    output_by_factory = factory_output_schedule(factories, days, production_rows)
    sources = {
        factory.name: {
            "initial": factory.init,
            "supply": output_by_factory.get(factory.name, [0] * days),
            "supply_is_capacity": production_rows is None,
            "limit": 0.0,
        }
        for factory in factories
    }
    destinations = {
        row["node"]: {
            "initial": row["init"],
            "demand": daily_demands_for_forecast(row, days),
            "limit": row["limit"],
            "excess_fee": row.get("excess_fee", 0.0),
        }
        for row in forecasts
    }
    usable_routes = [
        route
        for route in pruned_product_routes(
            forecasts=forecasts,
            factories=factories,
            routes=routes,
            products=products,
            cargo=cargo,
            carriers=carriers,
            top_k=99,
        )
        if route.src in sources and route.dst in destinations
    ]
    if not usable_routes:
        return [], {"method": "FullGlobalIntegerAudit", "status": "NoRoute"}
    transport = _solve_day_transport_milp(
        name=name,
        sources=sources,
        destinations=destinations,
        routes=usable_routes,
        products=products,
        cargo=cargo,
        carriers=carriers,
        days=days,
        ship_day_step=1,
        gap_rel=0.0,
        enforce_destination_limits=True,
        max_total_shortage=0.0,
        allow_fallback=False,
        time_limit_sec=time_limit_sec,
        warm_start_shipments=warm_start_shipments,
        relax_integrality=False,
        prune_transport_options=True,
    )
    rows = [
        {
            "destination": row["destination"],
            "factory": row["source"],
            "source": row["source"],
            "cargo": row["cargo"],
            "amount": int(row["amount"]),
            "route": row["route"],
            "mode": row["mode"],
            "carrier": row.get("carrier", ""),
            "lead": int(row["lead"]),
            "ship_day": int(row["ship_day"]),
            "arrival_day": int(row["arrival_day"]),
            "freight_cost": float(row["freight_cost"]),
            "note": "全路线全日期全局整数审查",
        }
        for row in transport.get("shipments", [])
    ]
    return rows, {
        "method": "FullGlobalIntegerAudit",
        "status": transport.get("status"),
        "shortage": transport.get("shortage", 0.0),
        "failures": transport.get("failures", []),
        "solve_seconds": transport.get("solve_seconds"),
        "objective": transport.get("objective"),
        "raw_status": transport.get("raw_status"),
        "model_stats": transport.get("model_stats", {}),
        "time_limit_hit": transport.get("time_limit_hit"),
    }


def pruned_product_routes(
    *,
    forecasts: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None = None,
    top_k: int = 3,
) -> list[Route]:
    if is_extreme_mode():
        seen_all: set[tuple[str, str, str]] = set()
        all_routes: list[Route] = []
        factory_names = {factory.name for factory in factories}
        node_names = {forecast["node"] for forecast in forecasts}
        for route in routes:
            if route.src not in factory_names or route.dst not in node_names:
                continue
            key = (route.route, route.src, route.dst)
            if key in seen_all:
                continue
            seen_all.add(key)
            all_routes.append(route)
        return all_routes
    ratio = charge_ratio(products, cargo)
    selected: list[Route] = []
    seen: set[tuple[str, str, str]] = set()
    for forecast in forecasts:
        node = forecast["node"]
        reference = max(
            float(forecast.get("forecast") or 0.0) - float(forecast.get("init") or 0.0),
            float(forecast.get("daily_avg") or 1.0) * 3,
            1.0,
        )
        for factory in factories:
            candidates = lane_routes(routes, factory.name, node)
            candidates = sorted(
                candidates,
                key=lambda route: (
                    route_unit_cost(route, max(reference, 1.0), ratio, carriers),
                    route_effective_lead(route, max(reference, 1.0), ratio, carriers),
                    route.route,
                ),
            )
            for route in candidates[:max(1, top_k)]:
                key = (route.route, route.src, route.dst)
                if key not in seen:
                    selected.append(route)
                    seen.add(key)
    return selected


def _future_inventory_states(demand: list[int], day: int, limit: int) -> list[int]:
    """Candidate end-of-day stocks that correspond to covering an integer future window."""
    if limit <= 0:
        return [0]
    states = {0, int(limit)}
    total = 0
    for idx in range(day, len(demand)):
        total += int(demand[idx] or 0)
        if total > limit:
            break
        states.add(total)
    return sorted(states)


def _factory_cumulative_used(rows: list[dict[str, Any]], factory_name: str, ship_day: int) -> int:
    return sum(
        int(row.get("amount") or 0)
        for row in rows
        if sv(row.get("source") or row.get("factory")) == factory_name
        and int(row.get("ship_day") or 1) <= ship_day
    )


def _factory_usage_prefix(rows: list[dict[str, Any]], factories: list[Factory], days: int) -> dict[str, list[int]]:
    usage = {factory.name: [0] * (days + 1) for factory in factories}
    for row in rows:
        factory_name = sv(row.get("source") or row.get("factory"))
        if factory_name not in usage:
            continue
        day = max(1, min(days, int(row.get("ship_day") or 1)))
        usage[factory_name][day] += int(row.get("amount") or 0)
    for values in usage.values():
        running = 0
        for day in range(1, days + 1):
            running += values[day]
            values[day] = running
    return usage


def _factory_cumulative_capacity(factory: Factory, output: list[int], ship_day: int) -> int:
    day = max(1, int(ship_day or 1))
    return int(math.floor(float(factory.init or 0.0) + sum(float(value or 0.0) for value in output[:day])))


def _factory_capacity_prefix(factories: list[Factory], output_by_factory: dict[str, list[int]], days: int) -> dict[str, list[int]]:
    capacity: dict[str, list[int]] = {}
    for factory in factories:
        values = [0] * (days + 1)
        running = float(factory.init or 0.0)
        output = output_by_factory.get(factory.name, [0] * days)
        for day in range(1, days + 1):
            running += float(output[day - 1] if day - 1 < len(output) else 0.0)
            values[day] = int(math.floor(running))
        capacity[factory.name] = values
    return capacity


def _shipment_shadow_cost(
    *,
    factory_name: str,
    amount: int,
    ship_day: int,
    days: int,
    factory_day_penalty: dict[tuple[str, int], float] | None,
) -> float:
    if not factory_day_penalty or amount <= 0:
        return 0.0
    start = max(1, int(ship_day or 1))
    return float(amount) * sum(
        float(factory_day_penalty.get((factory_name, day), 0.0) or 0.0)
        for day in range(start, days + 1)
    )


def _factory_shadow_penalties(
    *,
    rows: list[dict[str, Any]],
    factories: list[Factory],
    output_by_factory: dict[str, list[int]],
    days: int,
    strength: float,
) -> dict[tuple[str, int], float]:
    """Generate search-only prices for tight factory/date capacity.

    The values are not real costs and are never included in the reported score.
    They only make the per-destination DP produce alternative patterns that use
    less congested factories or ship days, which the final selector then judges
    by true freight and hard inventory constraints.
    """
    if strength <= 0:
        return {}
    usage = _factory_usage_prefix(rows, factories, days)
    capacity = _factory_capacity_prefix(factories, output_by_factory, days)
    penalties: dict[tuple[str, int], float] = {}
    for factory in factories:
        daily_scale = max(1.0, float(factory.daily or 0.0) * 3.0)
        for day in range(1, days + 1):
            slack = float(capacity.get(factory.name, [0] * (days + 1))[day] - usage.get(factory.name, [0] * (days + 1))[day])
            tightness = max(0.0, 1.0 - max(0.0, slack) / daily_scale)
            if tightness <= 0:
                continue
            remaining = max(1, days - day + 1)
            penalties[(factory.name, day)] = float(strength) * tightness / remaining
    return penalties


def _best_arrival_option_row(
    *,
    node: str,
    amount: int,
    arrival_day: int,
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    base_rows: list[dict[str, Any]],
    local_rows: list[dict[str, Any]],
    output_by_factory: dict[str, list[int]],
    option_cache: dict[tuple[str, int, int], list[dict[str, Any]]] | None = None,
    base_usage_prefix: dict[str, list[int]] | None = None,
    capacity_prefix: dict[str, list[int]] | None = None,
    factory_day_penalty: dict[tuple[str, int], float] | None = None,
) -> dict[str, Any] | None:
    rows = _arrival_option_rows(
        node=node,
        amount=amount,
        arrival_day=arrival_day,
        factories=factories,
        routes=routes,
        products=products,
        cargo=cargo,
        carriers=carriers,
        days=days,
        base_rows=base_rows,
        local_rows=local_rows,
        output_by_factory=output_by_factory,
        option_cache=option_cache,
        base_usage_prefix=base_usage_prefix,
        capacity_prefix=capacity_prefix,
        factory_day_penalty=factory_day_penalty,
        max_options=1,
    )
    return rows[0] if rows else None


def _arrival_option_rows(
    *,
    node: str,
    amount: int,
    arrival_day: int,
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    base_rows: list[dict[str, Any]],
    local_rows: list[dict[str, Any]],
    output_by_factory: dict[str, list[int]],
    option_cache: dict[tuple[str, int, int], list[dict[str, Any]]] | None = None,
    base_usage_prefix: dict[str, list[int]] | None = None,
    capacity_prefix: dict[str, list[int]] | None = None,
    factory_day_penalty: dict[tuple[str, int], float] | None = None,
    max_options: int = 1,
) -> list[dict[str, Any]]:
    if amount <= 0:
        return []
    ratio = charge_ratio(products, cargo)
    cache_key = (node, int(amount), int(arrival_day))
    use_cache = option_cache is not None and not factory_day_penalty
    candidates = option_cache.get(cache_key) if use_cache else None
    local_usage_prefix = _factory_usage_prefix(local_rows, factories, days) if local_rows else None
    if candidates is None:
        candidates = []
        for factory in factories:
            for route in lane_routes(routes, factory.name, node):
                options = route_transport_options(
                    route,
                    ratio,
                    carriers,
                    max_amount=max(float(amount), 1.0),
                    prune_dominated=True,
                )
                for option in options:
                    lead = int(option.lead or 0)
                    ship_day = int(arrival_day) - lead
                    if ship_day < 1 or ship_day > days:
                        continue
                    freight = transport_option_cost(option, amount)
                    shadow = _shipment_shadow_cost(
                        factory_name=factory.name,
                        amount=int(amount),
                        ship_day=ship_day,
                        days=days,
                        factory_day_penalty=factory_day_penalty,
                    )
                    candidates.append(
                        {
                            "destination": node,
                            "factory": factory.name,
                            "source": factory.name,
                            "cargo": cargo,
                            "amount": int(amount),
                            "route": route.route,
                            "mode": route_mode(route),
                            "carrier": option.carrier,
                            "lead": lead,
                            "ship_day": ship_day,
                            "arrival_day": int(arrival_day),
                            "freight_cost": float(freight),
                            "_search_cost": float(freight) + shadow,
                            "note": "极限版库存窗口补货",
                        }
                    )
        candidates.sort(key=lambda row: (float(row.get("_search_cost", row.get("freight_cost") or 0.0) or 0.0), int(row.get("lead") or 0), sv(row.get("route"))))
        if use_cache:
            option_cache[cache_key] = candidates
    feasible_rows: list[dict[str, Any]] = []
    seen_signatures: set[tuple[Any, ...]] = set()
    for row in candidates:
        factory_name = sv(row.get("factory") or row.get("source"))
        factory = next((item for item in factories if item.name == factory_name), None)
        if factory is None:
            continue
        ship_day = int(row.get("ship_day") or 1)
        factory_output = output_by_factory.get(factory.name, [0] * days)
        feasible_capacity = True
        for check_day in range(ship_day, days + 1):
            available = (
                capacity_prefix.get(factory.name, [0] * (days + 1))[check_day]
                if capacity_prefix is not None
                else _factory_cumulative_capacity(factory, factory_output, check_day)
            )
            base_used = (
                base_usage_prefix.get(factory.name, [0] * (days + 1))[check_day]
                if base_usage_prefix is not None
                else _factory_cumulative_used(base_rows, factory.name, check_day)
            )
            local_used = (
                local_usage_prefix.get(factory.name, [0] * (days + 1))[check_day]
                if local_usage_prefix is not None
                else _factory_cumulative_used(local_rows, factory.name, check_day)
            )
            used = base_used + local_used
            if used + amount > available + 1e-6:
                feasible_capacity = False
                break
        if feasible_capacity:
            signature = (
                factory_name,
                sv(row.get("route")),
                sv(row.get("carrier")),
                int(row.get("ship_day") or 1),
                int(row.get("arrival_day") or 1),
            )
            if signature in seen_signatures:
                continue
            seen_signatures.add(signature)
            feasible_rows.append(dict(row))
            if len(feasible_rows) >= max(1, int(max_options or 1)):
                break
    return feasible_rows


def _path_usage_signature(rows: list[dict[str, Any]], factories: list[Factory], days: int) -> tuple[Any, ...]:
    checkpoints = sorted({max(1, min(days, day)) for day in (7, 14, 21, 30, 45, days)})
    signature: list[Any] = []
    for factory in factories:
        factory_name = factory.name
        for day in checkpoints:
            used = sum(
                int(row.get("amount") or 0)
                for row in rows
                if sv(row.get("factory") or row.get("source")) == factory_name
                and int(row.get("ship_day") or 1) <= day
            )
            signature.append((factory_name, day, used))
    return tuple(signature)


def inventory_window_beam_width(cargo: str) -> int:
    if sv(cargo) == "热水器" and is_extreme_mode():
        return max(1, env_int("SUPPLY_CHAIN_HEATWATER_DP_BEAM", 1))
    return max(1, env_int("SUPPLY_CHAIN_DP_BEAM", 1))


def inventory_window_arrival_options(cargo: str) -> int:
    if sv(cargo) == "热水器" and is_extreme_mode():
        return max(1, env_int("SUPPLY_CHAIN_HEATWATER_DP_OPTIONS", 1))
    return max(1, env_int("SUPPLY_CHAIN_DP_OPTIONS", 1))


def _append_dp_state_candidate(
    bucket: list[tuple[float, list[dict[str, Any]]]],
    candidate: tuple[float, list[dict[str, Any]]],
    *,
    factories: list[Factory],
    days: int,
    beam_width: int,
) -> None:
    cost, rows = candidate
    new_sig = _path_usage_signature(rows, factories, days)
    for idx, (old_cost, old_rows) in enumerate(list(bucket)):
        old_sig = _path_usage_signature(old_rows, factories, days)
        if old_sig == new_sig:
            if cost < old_cost - 1e-6 or (abs(cost - old_cost) <= 1e-6 and len(rows) < len(old_rows)):
                bucket[idx] = candidate
            break
    else:
        bucket.append(candidate)
    bucket.sort(key=lambda item: (item[0], len(item[1])))
    del bucket[max(1, int(beam_width or 1)) :]


def _solve_one_destination_inventory_window(
    *,
    forecast: dict[str, Any],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    base_rows: list[dict[str, Any]],
    output_by_factory: dict[str, list[int]],
    option_cache: dict[tuple[str, int, int], list[dict[str, Any]]] | None = None,
    capacity_prefix: dict[str, list[int]] | None = None,
    factory_day_penalty: dict[tuple[str, int], float] | None = None,
    beam_width: int = 1,
    arrival_options_per_state: int = 1,
) -> list[dict[str, Any]] | None:
    node = sv(forecast.get("node"))
    demand = daily_demands_for_forecast(forecast, days)
    limit = int(math.floor(float(forecast.get("limit") or 0.0)))
    initial = int(math.floor(float(forecast.get("init") or 0.0)))
    if not node or not demand:
        return []
    if limit <= 0:
        limit = max(initial, sum(demand), 1)

    base_usage_prefix = _factory_usage_prefix(base_rows, factories, days)
    capacity_prefix = capacity_prefix or _factory_capacity_prefix(factories, output_by_factory, days)
    beam_width = max(1, int(beam_width or 1))
    arrival_options_per_state = max(1, int(arrival_options_per_state or 1))
    states: dict[int, list[tuple[float, list[dict[str, Any]]]]] = {initial: [(0.0, [])]}
    for day in range(1, days + 1):
        today_demand = int(demand[day - 1] if day - 1 < len(demand) else 0)
        next_states: dict[int, list[tuple[float, list[dict[str, Any]]]]] = {}
        candidates = _future_inventory_states(demand, day, limit)
        for inv_before, state_items in states.items():
            for cost_so_far, rows_so_far in state_items:
                natural_inv = int(inv_before - today_demand)
                if natural_inv >= 0:
                    bucket = next_states.setdefault(natural_inv, [])
                    if beam_width <= 1:
                        if not bucket or cost_so_far < bucket[0][0] - 1e-6 or (abs(cost_so_far - bucket[0][0]) <= 1e-6 and len(rows_so_far) < len(bucket[0][1])):
                            next_states[natural_inv] = [(cost_so_far, rows_so_far)]
                    else:
                        _append_dp_state_candidate(
                            bucket,
                            (cost_so_far, rows_so_far),
                            factories=factories,
                            days=days,
                            beam_width=beam_width,
                        )
                for inv_after in candidates:
                    amount = int(today_demand + inv_after - inv_before)
                    if amount < 0:
                        continue
                    if amount == 0:
                        option_rows = [None]
                    else:
                        option_rows = _arrival_option_rows(
                            node=node,
                            amount=amount,
                            arrival_day=day,
                            factories=factories,
                            routes=routes,
                            products=products,
                            cargo=cargo,
                            carriers=carriers,
                            days=days,
                            base_rows=base_rows,
                            local_rows=rows_so_far,
                            output_by_factory=output_by_factory,
                            option_cache=option_cache,
                            base_usage_prefix=base_usage_prefix,
                            capacity_prefix=capacity_prefix,
                            factory_day_penalty=factory_day_penalty,
                            max_options=arrival_options_per_state,
                        )
                        if not option_rows:
                            continue
                    for row in option_rows:
                        add_cost = 0.0 if row is None else float(row.get("_search_cost", row.get("freight_cost") or 0.0) or 0.0)
                        new_rows = rows_so_far + ([row] if row else [])
                        new_cost = cost_so_far + add_cost
                        bucket = next_states.setdefault(inv_after, [])
                        if beam_width <= 1:
                            if not bucket or new_cost < bucket[0][0] - 1e-6 or (abs(new_cost - bucket[0][0]) <= 1e-6 and len(new_rows) < len(bucket[0][1])):
                                next_states[inv_after] = [(new_cost, new_rows)]
                        else:
                            _append_dp_state_candidate(
                                bucket,
                                (new_cost, new_rows),
                                factories=factories,
                                days=days,
                                beam_width=beam_width,
                            )
        states = next_states
        if not states:
            return None
    if not states:
        return None
    _ending_inv, state_items = min(
        states.items(),
        key=lambda item: (min(candidate[0] for candidate in item[1]), item[0], min(len(candidate[1]) for candidate in item[1])),
    )
    _cost, rows = min(state_items, key=lambda item: (item[0], len(item[1])))
    return rows


def independent_destination_lower_bound(
    *,
    forecasts: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
    include_shipments: bool = False,
) -> dict[str, Any]:
    """Optimistic lower bound from solving each destination independently.

    This deliberately ignores cross-destination factory capacity competition, so
    it is not a feasible global plan. It is useful as a cheap audit bound: if the
    selected feasible plan is close to this value, remaining improvement is
    necessarily small inside the current route/cost model.
    """
    if not forecasts or not factories:
        return {"status": "Skipped", "reason": "NoDemandOrFactory"}
    started_at = time.time()
    output_by_factory = factory_output_schedule(factories, days, production_rows)
    capacity_prefix = _factory_capacity_prefix(factories, output_by_factory, days)
    option_cache: dict[tuple[str, int, int], list[dict[str, Any]]] = {}
    node_costs: dict[str, float] = {}
    node_shipments: dict[str, list[dict[str, Any]]] = {}
    failures: list[str] = []
    total = 0.0
    for forecast in forecasts:
        node = sv(forecast.get("node"))
        rows = _solve_one_destination_inventory_window(
            forecast=forecast,
            factories=factories,
            routes=routes,
            products=products,
            cargo=cargo,
            carriers=carriers,
            days=days,
            base_rows=[],
            output_by_factory=output_by_factory,
            option_cache=option_cache,
            capacity_prefix=capacity_prefix,
            beam_width=inventory_window_beam_width(cargo),
            arrival_options_per_state=inventory_window_arrival_options(cargo),
        )
        if rows is None:
            failures.append(node or "未知销售网点")
            continue
        cost = sum(float(row.get("freight_cost") or 0.0) for row in rows)
        node_costs[node] = cost
        if include_shipments:
            node_shipments[node] = [
                {key: value for key, value in dict(row).items() if not str(key).startswith("_")}
                for row in rows
            ]
        total += cost
    result = {
        "status": "Computed" if not failures else "Partial",
        "freight_lower_bound": round(total, 4),
        "node_costs": {node: round(cost, 4) for node, cost in node_costs.items()},
        "failures": failures,
        "seconds": round(time.time() - started_at, 3),
    }
    if include_shipments:
        result["node_shipments"] = node_shipments
    return result


def heatwater_gap_diagnostics(
    *,
    forecasts: list[dict[str, Any]],
    factories: list[Factory],
    final_shipments: list[dict[str, Any]],
    independent_status: dict[str, Any],
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Explain the remaining heatwater gap without changing the objective.

    The independent destination lower bound ignores cross-destination factory
    capacity. This diagnostic checks whether that bound is physically feasible
    and shows which destinations pay the conflict cost in the selected plan.
    """
    node_costs = independent_status.get("node_costs") or {}
    node_shipments = independent_status.get("node_shipments") or {}
    if not isinstance(node_costs, dict) or not isinstance(node_shipments, dict):
        return {"status": "Skipped", "reason": "NoIndependentShipmentEvidence"}

    def total_cost(items: list[dict[str, Any]]) -> float:
        return sum(float(row.get("freight_cost") or 0.0) for row in items)

    final_by_node: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in final_shipments:
        node = sv(row.get("destination"))
        if node:
            final_by_node[node].append(dict(row))

    node_deltas: dict[str, dict[str, Any]] = {}
    for forecast in forecasts:
        node = sv(forecast.get("node"))
        final_rows = final_by_node.get(node, [])
        final_cost = total_cost(final_rows)
        independent_cost = float(node_costs.get(node, 0.0) or 0.0)
        node_deltas[node] = {
            "final_cost": round(final_cost, 4),
            "independent_cost": round(independent_cost, 4),
            "delta": round(final_cost - independent_cost, 4),
            "final_shipments": len(final_rows),
            "final_amount": int(sum(int(row.get("amount") or 0) for row in final_rows)),
            "independent_shipments": len(node_shipments.get(node, [])),
        }

    output_by_factory = factory_output_schedule(factories, days, production_rows)
    capacity_prefix = _factory_capacity_prefix(factories, output_by_factory, days)
    independent_rows = [
        dict(row)
        for rows in node_shipments.values()
        for row in (rows if isinstance(rows, list) else [])
    ]
    overflows: list[dict[str, Any]] = []
    max_overflow = 0.0
    for factory in factories:
        capacity_values = capacity_prefix.get(factory.name, [0] * (days + 1))
        for day in range(1, days + 1):
            used = sum(
                int(row.get("amount") or 0)
                for row in independent_rows
                if sv(row.get("source") or row.get("factory")) == factory.name
                and int(row.get("ship_day") or 1) <= day
            )
            capacity = float(capacity_values[day] if day < len(capacity_values) else 0.0)
            overflow = max(0.0, float(used) - capacity)
            if overflow > 1e-6:
                max_overflow = max(max_overflow, overflow)
                overflows.append(
                    {
                        "factory": factory.name,
                        "day": day,
                        "used": int(used),
                        "capacity": int(round(capacity)),
                        "overflow": int(math.ceil(overflow)),
                    }
                )

    top_overflows = sorted(overflows, key=lambda row: (-int(row["overflow"]), row["factory"], row["day"]))[:8]
    total_final = total_cost(final_shipments)
    total_independent = float(independent_status.get("freight_lower_bound") or 0.0)
    return {
        "status": "Computed",
        "final_freight": round(total_final, 4),
        "independent_lower_bound": round(total_independent, 4),
        "gap": round(total_final - total_independent, 4),
        "gap_ratio": round((total_final - total_independent) / max(total_final, 0.001), 6),
        "node_deltas": node_deltas,
        "independent_capacity_feasible": not overflows,
        "max_independent_capacity_overflow": int(math.ceil(max_overflow)),
        "top_independent_capacity_overflows": top_overflows,
    }


def capacity_aware_transport_lower_bound(
    *,
    forecasts: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """LP-relaxed global lower bound with factory and store inventory constraints."""
    if not forecasts or not factories:
        return {"status": "Skipped", "reason": "NoDemandOrFactory"}
    output_by_factory = factory_output_schedule(factories, days, production_rows)
    sources = {
        factory.name: {
            "initial": factory.init,
            "supply": output_by_factory.get(factory.name, [0] * days),
            "supply_is_capacity": production_rows is None,
            "limit": 0.0,
        }
        for factory in factories
    }
    destinations = {
        row["node"]: {
            "initial": row["init"],
            "demand": daily_demands_for_forecast(row, days),
            "limit": row["limit"],
            "excess_fee": row.get("excess_fee", 0.0),
        }
        for row in forecasts
    }
    usable_routes = [
        route
        for route in pruned_product_routes(
            forecasts=forecasts,
            factories=factories,
            routes=routes,
            products=products,
            cargo=cargo,
            carriers=carriers,
            top_k=99,
        )
        if route.src in sources and route.dst in destinations
    ]
    if not usable_routes:
        return {"status": "NoRoute", "freight_lower_bound": 0.0}
    transport = _solve_day_transport_milp(
        name=f"{cargo}_capacity_aware_lower_bound",
        sources=sources,
        destinations=destinations,
        routes=usable_routes,
        products=products,
        cargo=cargo,
        carriers=carriers,
        days=days,
        ship_day_step=1,
        gap_rel=0.0,
        enforce_destination_limits=True,
        max_total_shortage=0.0,
        allow_fallback=False,
        time_limit_sec=env_int("SUPPLY_CHAIN_LB_TIMELIMIT", 120),
        relax_integrality=True,
        prune_transport_options=False,
    )
    objective = float(transport.get("objective") or transport.get("freight_cost") or 0.0)
    return {
        "status": transport.get("status", "Unknown"),
        "raw_status": transport.get("raw_status"),
        "freight_lower_bound": round(objective, 4),
        "solve_seconds": round(float(transport.get("solve_seconds") or 0.0), 3),
        "time_limit_hit": bool(transport.get("time_limit_hit")),
        "model_stats": transport.get("model_stats", {}),
        "routes": len(usable_routes),
    }


def build_inventory_window_extreme_sales_transport(
    *,
    forecasts: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not is_extreme_mode() or not forecasts or not factories:
        return [], {"method": "InventoryWindowExtremeSalesDP", "status": "Skipped"}
    output_by_factory = factory_output_schedule(factories, days, production_rows)
    capacity_prefix = _factory_capacity_prefix(factories, output_by_factory, days)
    orders = list(iter_permutations(forecasts)) if len(forecasts) <= 6 else [tuple(sorted(forecasts, key=lambda row: float(row.get("forecast") or 0.0), reverse=True))]
    best_rows: list[dict[str, Any]] = []
    best_metrics: tuple[float, int] | None = None
    failures = 0
    option_cache: dict[tuple[str, int, int], list[dict[str, Any]]] = {}
    for order in orders:
        rows: list[dict[str, Any]] = []
        ok = True
        for forecast in order:
            node_rows = _solve_one_destination_inventory_window(
                forecast=forecast,
                factories=factories,
                routes=routes,
                products=products,
                cargo=cargo,
                carriers=carriers,
                days=days,
                base_rows=rows,
                output_by_factory=output_by_factory,
                option_cache=option_cache,
                capacity_prefix=capacity_prefix,
                beam_width=inventory_window_beam_width(cargo),
                arrival_options_per_state=inventory_window_arrival_options(cargo),
            )
            if node_rows is None:
                ok = False
                failures += 1
                break
            rows.extend(node_rows)
        if not ok:
            continue
        rows = sorted(effective_shipments(refresh_transport_rows(rows, routes, products, cargo, carriers), days), key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("route")), sv(row.get("destination"))))
        hard_risks = product_transport_hard_risks(
            forecasts=forecasts,
            shipments=rows,
            factories=factories,
            days=days,
            production_rows=production_rows,
        )
        replay = daily_market_replay(forecasts, rows, days)
        if hard_risks or not is_full_satisfaction(float(replay.get("market_satisfaction", 0.0)), float(replay.get("shortage", 0.0))):
            failures += 1
            continue
        freight = sum(float(row.get("freight_cost") or 0.0) for row in rows)
        metric = (freight, len(rows))
        if best_metrics is None or metric < best_metrics:
            best_metrics = metric
            best_rows = rows
    best_rows, polish_status = polish_inventory_window_transport_by_destination(
        forecasts=forecasts,
        shipments=best_rows,
        factories=factories,
        routes=routes,
        products=products,
        cargo=cargo,
        carriers=carriers,
        days=days,
        production_rows=production_rows,
        option_cache=option_cache,
        capacity_prefix=capacity_prefix,
    )
    best_metrics = (sum(float(row.get("freight_cost") or 0.0) for row in best_rows), len(best_rows)) if best_rows else best_metrics
    status = {
        "method": "InventoryWindowExtremeSalesDP",
        "status": "OptimalWindowCandidate" if best_rows else "NoFeasibleWindowCandidate",
        "orders_tried": len(orders),
        "failed_orders": failures,
        "option_cache_entries": len(option_cache),
        "polish": polish_status,
        "freight_cost": best_metrics[0] if best_metrics else 0.0,
        "shipments": len(best_rows),
    }
    return best_rows, status


def polish_inventory_window_transport_by_destination(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
    option_cache: dict[tuple[str, int, int], list[dict[str, Any]]] | None = None,
    capacity_prefix: dict[str, list[int]] | None = None,
    max_rounds: int = 3,
    include_pair_search: bool = True,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not shipments or not forecasts:
        return shipments, {"rounds": 0, "improvements": 0, "delta": 0.0}
    output_by_factory = factory_output_schedule(factories, days, production_rows)
    capacity_prefix = capacity_prefix or _factory_capacity_prefix(factories, output_by_factory, days)
    rows = sorted([dict(row) for row in shipments], key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("destination")), sv(row.get("route"))))
    cache = option_cache if option_cache is not None else {}

    def total_cost(items: list[dict[str, Any]]) -> float:
        return sum(float(row.get("freight_cost") or 0.0) for row in items)

    start_cost = total_cost(rows)
    improvements = 0
    pair_improvements = 0
    rounds_run = 0
    for round_idx in range(max_rounds):
        rounds_run = round_idx + 1
        changed = False
        for forecast in sorted(forecasts, key=lambda row: sv(row.get("node"))):
            node = sv(forecast.get("node"))
            base_rows = [dict(row) for row in rows if sv(row.get("destination")) != node]
            replacement = _solve_one_destination_inventory_window(
                forecast=forecast,
                factories=factories,
                routes=routes,
                products=products,
                cargo=cargo,
                carriers=carriers,
                days=days,
                base_rows=base_rows,
                output_by_factory=output_by_factory,
                option_cache=cache,
                capacity_prefix=capacity_prefix,
                beam_width=inventory_window_beam_width(cargo),
                arrival_options_per_state=inventory_window_arrival_options(cargo),
            )
            if replacement is None:
                continue
            candidate = sorted(
                effective_shipments(refresh_transport_rows(base_rows + replacement, routes, products, cargo, carriers), days),
                key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("destination")), sv(row.get("route"))),
            )
            if total_cost(candidate) >= total_cost(rows) - 1e-6:
                continue
            replay = daily_market_replay(forecasts, candidate, days)
            hard_risks = product_transport_hard_risks(
                forecasts=forecasts,
                shipments=candidate,
                factories=factories,
                days=days,
                production_rows=production_rows,
            )
            if hard_risks or not is_full_satisfaction(float(replay.get("market_satisfaction", 0.0)), float(replay.get("shortage", 0.0))):
                continue
            rows = candidate
            improvements += 1
            changed = True
        if not changed:
            break
    if include_pair_search:
        forecast_by_node = {sv(row.get("node")): row for row in forecasts}
        for left, right in iter_combinations(sorted(forecast_by_node), 2):
            base_rows = [dict(row) for row in rows if sv(row.get("destination")) not in {left, right}]
            best_pair_rows: list[dict[str, Any]] | None = None
            best_pair_cost = total_cost(rows)
            for node_order in iter_permutations([left, right]):
                candidate_rows = [dict(row) for row in base_rows]
                feasible = True
                for node in node_order:
                    replacement = _solve_one_destination_inventory_window(
                        forecast=forecast_by_node[node],
                        factories=factories,
                        routes=routes,
                        products=products,
                        cargo=cargo,
                        carriers=carriers,
                        days=days,
                        base_rows=candidate_rows,
                        output_by_factory=output_by_factory,
                        option_cache=cache,
                        capacity_prefix=capacity_prefix,
                        beam_width=inventory_window_beam_width(cargo),
                        arrival_options_per_state=inventory_window_arrival_options(cargo),
                    )
                    if replacement is None:
                        feasible = False
                        break
                    candidate_rows.extend(replacement)
                if not feasible:
                    continue
                candidate_rows = sorted(
                    effective_shipments(refresh_transport_rows(candidate_rows, routes, products, cargo, carriers), days),
                    key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("destination")), sv(row.get("route"))),
                )
                candidate_cost = total_cost(candidate_rows)
                if candidate_cost >= best_pair_cost - 1e-6:
                    continue
                replay = daily_market_replay(forecasts, candidate_rows, days)
                hard_risks = product_transport_hard_risks(
                    forecasts=forecasts,
                    shipments=candidate_rows,
                    factories=factories,
                    days=days,
                    production_rows=production_rows,
                )
                if hard_risks or not is_full_satisfaction(float(replay.get("market_satisfaction", 0.0)), float(replay.get("shortage", 0.0))):
                    continue
                best_pair_rows = candidate_rows
                best_pair_cost = candidate_cost
            if best_pair_rows is not None and best_pair_cost < total_cost(rows) - 1e-6:
                rows = best_pair_rows
                pair_improvements += 1
    end_cost = total_cost(rows)
    return rows, {
        "rounds": rounds_run,
        "improvements": improvements,
        "pair_improvements": pair_improvements,
        "pair_search": bool(include_pair_search),
        "delta": round(start_cost - end_cost, 4),
    }


def polish_shift_expensive_early_to_later(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
    max_rounds: int = 4,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Move surplus early arrivals from expensive lanes into cheaper later arrivals."""
    if not shipments or not forecasts:
        return shipments, {"rounds": 0, "improvements": 0, "delta": 0.0}
    route_by_name = {route.route: route for route in routes}
    rows = [dict(row) for row in shipments]

    def total_cost(items: list[dict[str, Any]]) -> float:
        return sum(float(row.get("freight_cost") or 0.0) for row in items)

    def refresh_row(row: dict[str, Any]) -> None:
        route = route_by_name.get(sv(row.get("route")))
        if route is not None:
            refresh_shipment_route_fields(row, route, products, cargo, carriers)

    def shifted_row_cost(row: dict[str, Any], amount: int) -> float:
        if amount <= 0:
            return 0.0
        route = route_by_name.get(sv(row.get("route")))
        if route is None:
            original_amount = max(1, int(row.get("amount") or 0))
            return float(row.get("freight_cost") or 0.0) * amount / original_amount
        ratio = charge_ratio(products, cargo)
        return route_cost(route, amount, ratio, carriers)

    def shift_saving_estimate(from_row: dict[str, Any], to_row: dict[str, Any], delta: int) -> float:
        from_amount = int(from_row.get("amount") or 0)
        to_amount = int(to_row.get("amount") or 0)
        if delta <= 0 or delta > from_amount:
            return -1.0
        old_cost = float(from_row.get("freight_cost") or 0.0) + float(to_row.get("freight_cost") or 0.0)
        new_cost = shifted_row_cost(from_row, from_amount - delta) + shifted_row_cost(to_row, to_amount + delta)
        return old_cost - new_cost

    def candidate_after_shift(idx_from: int, idx_to: int, delta: int) -> list[dict[str, Any]] | None:
        if delta <= 0:
            return None
        candidate = [dict(row) for row in rows]
        candidate[idx_from]["amount"] = int(candidate[idx_from].get("amount") or 0) - delta
        candidate[idx_to]["amount"] = int(candidate[idx_to].get("amount") or 0) + delta
        if int(candidate[idx_from]["amount"]) < 0:
            return None
        refresh_row(candidate[idx_from])
        refresh_row(candidate[idx_to])
        candidate = [row for row in candidate if int(row.get("amount") or 0) > 0]
        candidate = sorted(
            effective_shipments(candidate, days),
            key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("destination")), sv(row.get("route")), int(row.get("amount") or 0)),
        )
        return candidate

    start_cost = total_cost(rows)
    improvements = 0
    rounds_run = 0
    estimated_candidates = 0
    checked_candidates = 0
    rejected_estimate = 0
    rejected_cost = 0
    rejected_market = 0
    rejected_hard = 0
    limit_by_dest = {sv(row.get("node")): int(float(row.get("limit") or 0.0)) for row in forecasts}
    for round_idx in range(max_rounds):
        rounds_run = round_idx + 1
        replay = daily_market_replay(forecasts, rows, days)
        inventory_by_dest_day = {
            (sv(item.get("node")), int(item.get("day") or 0)): int(item.get("ending_inventory") or 0)
            for item in replay.get("daily_rows", [])
        }
        best_rows: list[dict[str, Any]] | None = None
        best_saving = 0.0
        current_cost = total_cost(rows)
        for idx_from, early in enumerate(rows):
            amount_from = int(early.get("amount") or 0)
            if amount_from <= 1:
                continue
            dest = sv(early.get("destination"))
            early_arrival = int(early.get("arrival_day") or early.get("ship_day") or 1)
            early_unit = float(early.get("freight_cost") or 0.0) / max(amount_from, 1)
            for idx_to, later in enumerate(rows):
                if idx_to == idx_from or sv(later.get("destination")) != dest:
                    continue
                later_arrival = int(later.get("arrival_day") or later.get("ship_day") or 1)
                if later_arrival <= early_arrival:
                    continue
                amount_to = int(later.get("amount") or 0)
                if amount_to <= 0:
                    continue
                later_unit = float(later.get("freight_cost") or 0.0) / max(amount_to, 1)
                slack_days = range(early_arrival, later_arrival)
                max_delta = min(
                    amount_from,
                    min((inventory_by_dest_day.get((dest, day), 0) for day in slack_days), default=0),
                )
                if max_delta > 0:
                    for delta in shipment_shift_delta_candidates(
                        from_row=early,
                        to_row=later,
                        max_delta=int(max_delta),
                        route_by_name=route_by_name,
                        products=products,
                        cargo=cargo,
                        carriers=carriers,
                    ):
                        estimated_candidates += 1
                        estimated_saving = shift_saving_estimate(early, later, int(delta))
                        if estimated_saving <= best_saving + 1e-6:
                            rejected_estimate += 1
                            continue
                        candidate = candidate_after_shift(idx_from, idx_to, int(delta))
                        if candidate is None:
                            continue
                        candidate_cost = total_cost(candidate)
                        saving = current_cost - candidate_cost
                        if saving <= best_saving + 1e-6:
                            rejected_cost += 1
                            continue
                        checked_candidates += 1
                        replay_after = daily_market_replay(forecasts, candidate, days)
                        hard_risks = product_transport_hard_risks(
                            forecasts=forecasts,
                            shipments=candidate,
                            factories=factories,
                            days=days,
                            production_rows=production_rows,
                        )
                        if hard_risks:
                            rejected_hard += 1
                            continue
                        if not is_full_satisfaction(float(replay_after.get("market_satisfaction", 0.0)), float(replay_after.get("shortage", 0.0))):
                            rejected_market += 1
                            continue
                        best_rows = candidate
                        best_saving = saving
                if later_unit <= early_unit + 1e-6:
                    continue
                limit = int(limit_by_dest.get(dest, 0) or 0)
                if limit <= 0:
                    continue
                headroom_days = range(early_arrival, later_arrival)
                max_delta = min(
                    amount_to,
                    min((limit - inventory_by_dest_day.get((dest, day), limit) for day in headroom_days), default=0),
                )
                if max_delta <= 0:
                    continue
                for delta in shipment_shift_delta_candidates(
                    from_row=later,
                    to_row=early,
                    max_delta=int(max_delta),
                    route_by_name=route_by_name,
                    products=products,
                    cargo=cargo,
                    carriers=carriers,
                ):
                    estimated_candidates += 1
                    estimated_saving = shift_saving_estimate(later, early, int(delta))
                    if estimated_saving <= best_saving + 1e-6:
                        rejected_estimate += 1
                        continue
                    candidate = candidate_after_shift(idx_to, idx_from, int(delta))
                    if candidate is None:
                        continue
                    candidate_cost = total_cost(candidate)
                    saving = current_cost - candidate_cost
                    if saving <= best_saving + 1e-6:
                        rejected_cost += 1
                        continue
                    checked_candidates += 1
                    replay_after = daily_market_replay(forecasts, candidate, days)
                    hard_risks = product_transport_hard_risks(
                        forecasts=forecasts,
                        shipments=candidate,
                        factories=factories,
                        days=days,
                        production_rows=production_rows,
                    )
                    if hard_risks:
                        rejected_hard += 1
                        continue
                    if not is_full_satisfaction(float(replay_after.get("market_satisfaction", 0.0)), float(replay_after.get("shortage", 0.0))):
                        rejected_market += 1
                        continue
                    best_rows = candidate
                    best_saving = saving
        if best_rows is None or best_saving <= 1e-6:
            break
        rows = best_rows
        improvements += 1
    end_cost = total_cost(rows)
    return rows, {
        "rounds": rounds_run,
        "improvements": improvements,
        "delta": round(start_cost - end_cost, 4),
        "estimated_candidates": estimated_candidates,
        "checked_candidates": checked_candidates,
        "rejected_estimate": rejected_estimate,
        "rejected_cost": rejected_cost,
        "rejected_market": rejected_market,
        "rejected_hard": rejected_hard,
    }


def polish_single_arrival_replacements(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
    max_rounds: int = 3,
    option_cache: dict[tuple[str, int, int], list[dict[str, Any]]] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Replace individual shipments with the cheapest feasible lane for the same arrival.

    Keeping destination, amount, and arrival day fixed preserves the market
    inventory trajectory; only factory/source/route/carrier changes. Each
    accepted replacement is still checked against factory inventory and store
    hard limits.
    """
    if not shipments or not forecasts or not factories:
        return shipments, {"rounds": 0, "improvements": 0, "delta": 0.0}
    rows = sorted(
        [dict(row) for row in shipments],
        key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("destination")), sv(row.get("route"))),
    )
    output_by_factory = factory_output_schedule(factories, days, production_rows)
    capacity_prefix = _factory_capacity_prefix(factories, output_by_factory, days)
    cache = option_cache if option_cache is not None else {}

    def total_cost(items: list[dict[str, Any]]) -> float:
        return sum(float(row.get("freight_cost") or 0.0) for row in items)

    start_cost = total_cost(rows)
    improvements = 0
    rounds_run = 0
    for round_idx in range(max_rounds):
        rounds_run = round_idx + 1
        changed = False
        for idx, row in enumerate(list(rows)):
            destination = sv(row.get("destination"))
            amount = int(row.get("amount") or 0)
            arrival_day = int(row.get("arrival_day") or (int(row.get("ship_day") or 1) + int(row.get("lead") or 0)))
            if not destination or amount <= 0 or arrival_day <= 0:
                continue
            base_rows = [dict(item) for pos, item in enumerate(rows) if pos != idx]
            replacement = _best_arrival_option_row(
                node=destination,
                amount=amount,
                arrival_day=arrival_day,
                factories=factories,
                routes=routes,
                products=products,
                cargo=cargo,
                carriers=carriers,
                days=days,
                base_rows=base_rows,
                local_rows=[],
                output_by_factory=output_by_factory,
                option_cache=cache,
                base_usage_prefix=_factory_usage_prefix(base_rows, factories, days),
                capacity_prefix=capacity_prefix,
            )
            if replacement is None:
                continue
            old_cost = float(row.get("freight_cost") or 0.0)
            new_cost = float(replacement.get("freight_cost") or 0.0)
            if new_cost >= old_cost - 1e-6:
                continue
            candidate = sorted(
                effective_shipments(base_rows + [replacement], days),
                key=lambda item: (int(item.get("ship_day") or 1), sv(item.get("destination")), sv(item.get("route"))),
            )
            hard_risks = factory_finished_goods_risks(
                shipments=candidate,
                factories=factories,
                days=days,
                production_rows=production_rows,
            )
            if hard_risks:
                continue
            rows = candidate
            improvements += 1
            changed = True
            break
        if not changed:
            break
    return rows, {
        "rounds": rounds_run,
        "improvements": improvements,
        "delta": round(start_cost - total_cost(rows), 4),
        "option_cache_entries": len(cache),
    }


def polish_delay_to_cheaper_lanes(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
    max_rounds: int = 3,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Replace urgent expensive shipments with slower cheaper full-lane choices.

    This is a true plan change, not a display rewrite: every accepted candidate
    is replayed against store demand and factory finished-goods inventory.
    """
    if not shipments or not forecasts or not factories:
        return shipments, {"rounds": 0, "improvements": 0, "delta": 0.0}
    ratio = charge_ratio(products, cargo)
    rows = sorted(
        effective_shipments(refresh_transport_rows([dict(row) for row in shipments], routes, products, cargo, carriers), days),
        key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("destination")), sv(row.get("route"))),
    )

    def total_cost(items: list[dict[str, Any]]) -> float:
        return sum(float(row.get("freight_cost") or 0.0) for row in items)

    start_cost = total_cost(rows)
    improvements = 0
    rounds_run = 0
    candidates_checked = 0
    rejected_market = 0
    rejected_hard = 0
    rejected_cost = 0
    for round_idx in range(max_rounds):
        rounds_run = round_idx + 1
        current_cost = total_cost(rows)
        best_rows: list[dict[str, Any]] | None = None
        best_saving = 0.0
        for idx, row in enumerate(rows):
            destination = sv(row.get("destination"))
            source = sv(row.get("source") or row.get("factory"))
            amount = int(row.get("amount") or 0)
            old_cost = float(row.get("freight_cost") or 0.0)
            original_arrival = int(row.get("arrival_day") or (int(row.get("ship_day") or 1) + int(row.get("lead") or 0)))
            if not destination or not source or amount <= 0 or old_cost <= 0:
                continue
            trial_rows_without_current = [dict(item) for pos, item in enumerate(rows) if pos != idx]
            lane_candidates: list[dict[str, Any]] = []
            for route in lane_routes(routes, source, destination):
                for option in route_transport_options(route, ratio, carriers, max_amount=max(amount, 1), prune_dominated=True):
                    new_cost = transport_option_cost(option, amount)
                    if new_cost >= old_cost - 1e-6:
                        rejected_cost += 1
                        continue
                    lead = int(option.lead or 0)
                    earliest_arrival = max(1, 1 + lead, original_arrival)
                    if earliest_arrival > days:
                        continue
                    lane_candidates.append(
                        {
                            "route": route,
                            "option": option,
                            "lead": lead,
                            "freight_cost": float(new_cost),
                            "earliest_arrival": int(earliest_arrival),
                        }
                    )
            lane_candidates.sort(key=lambda item: (float(item["freight_cost"]), int(item["lead"]), sv(item["route"].route)))
            for lane in lane_candidates[:12]:
                route = lane["route"]
                option = lane["option"]
                lead = int(lane["lead"])
                # Earlier arrivals are safer; later arrivals may avoid store over-limit.
                candidate_arrivals = list(range(int(lane["earliest_arrival"]), days + 1))
                for arrival_day in candidate_arrivals:
                    ship_day = int(arrival_day) - lead
                    if ship_day < 1 or ship_day > days:
                        continue
                    replacement = {
                        "destination": destination,
                        "factory": source,
                        "source": source,
                        "cargo": cargo,
                        "amount": amount,
                        "route": route.route,
                        "mode": route_mode(route),
                        "carrier": option.carrier,
                        "lead": lead,
                        "ship_day": ship_day,
                        "arrival_day": int(arrival_day),
                        "freight_cost": float(lane["freight_cost"]),
                        "note": "延后改走低成本完整路线",
                    }
                    candidate = sorted(
                        effective_shipments(trial_rows_without_current + [replacement], days),
                        key=lambda item: (int(item.get("ship_day") or 1), sv(item.get("destination")), sv(item.get("route"))),
                    )
                    saving = current_cost - total_cost(candidate)
                    if saving <= best_saving + 1e-6:
                        rejected_cost += 1
                        continue
                    candidates_checked += 1
                    replay = daily_market_replay(forecasts, candidate, days)
                    if not is_full_satisfaction(float(replay.get("market_satisfaction", 0.0)), float(replay.get("shortage", 0.0))):
                        rejected_market += 1
                        continue
                    hard_risks = product_transport_hard_risks(
                        forecasts=forecasts,
                        shipments=candidate,
                        factories=factories,
                        days=days,
                        production_rows=production_rows,
                    )
                    if hard_risks:
                        rejected_hard += 1
                        continue
                    best_rows = candidate
                    best_saving = saving
                    break
        if best_rows is None or best_saving <= 1e-6:
            break
        rows = best_rows
        improvements += 1
    return rows, {
        "status": "Improved" if improvements > 0 else "NoCostImprovement",
        "rounds": rounds_run,
        "improvements": improvements,
        "delta": round(start_cost - total_cost(rows), 4),
        "checked": candidates_checked,
        "rejected_cost": rejected_cost,
        "rejected_market": rejected_market,
        "rejected_hard": rejected_hard,
    }


def optimize_fixed_arrival_reroute(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
    time_limit_sec: int = 30,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Globally re-route fixed shipment quantities while preserving arrivals.

    Destination, quantity, and arrival day stay fixed, so store inventory and
    market satisfaction are unchanged. The MILP only chooses source/route/carrier
    options and checks factory cumulative inventory.
    """
    if not shipments or not factories:
        return shipments, {"method": "FixedArrivalRerouteMILP", "status": "Skipped"}
    try:
        import pulp
    except ImportError:
        return shipments, {"method": "FixedArrivalRerouteMILP", "status": "Unavailable"}

    started_at = time.time()
    ratio = charge_ratio(products, cargo)
    output_by_factory = factory_output_schedule(factories, days, production_rows)
    capacity_prefix = _factory_capacity_prefix(factories, output_by_factory, days)
    clean_shipments = [dict(row) for row in effective_shipments(shipments, days)]
    options_by_idx: dict[int, list[dict[str, Any]]] = {}
    for idx, row in enumerate(clean_shipments):
        node = sv(row.get("destination"))
        amount = int(row.get("amount") or 0)
        arrival_day = int(row.get("arrival_day") or (int(row.get("ship_day") or 1) + int(row.get("lead") or 0)))
        options: list[dict[str, Any]] = []
        if not node or amount <= 0:
            continue
        for factory in factories:
            for route in lane_routes(routes, factory.name, node):
                for option in route_transport_options(route, ratio, carriers, max_amount=max(amount, 1), prune_dominated=True):
                    lead = int(option.lead or 0)
                    ship_day = arrival_day - lead
                    if ship_day < 1 or ship_day > days:
                        continue
                    options.append(
                        {
                            "destination": node,
                            "factory": factory.name,
                            "source": factory.name,
                            "cargo": cargo,
                            "amount": amount,
                            "route": route.route,
                            "mode": route_mode(route),
                            "carrier": option.carrier,
                            "lead": lead,
                            "ship_day": ship_day,
                            "arrival_day": arrival_day,
                            "freight_cost": transport_option_cost(option, amount),
                            "note": "固定到货日全局重路由",
                        }
                    )
        if not options:
            return shipments, {"method": "FixedArrivalRerouteMILP", "status": "NoOption", "shipment_index": idx}
        options_by_idx[idx] = sorted(options, key=lambda item: (float(item.get("freight_cost") or 0.0), int(item.get("lead") or 0), sv(item.get("route"))))

    model = pulp.LpProblem("fixed_arrival_reroute", pulp.LpMinimize)
    choose: dict[tuple[int, int], Any] = {}
    for idx, options in options_by_idx.items():
        for option_idx, _option in enumerate(options):
            choose[(idx, option_idx)] = pulp.LpVariable(f"reroute_{idx}_{option_idx}", lowBound=0, upBound=1, cat="Binary")
        model += pulp.lpSum(choose[(idx, option_idx)] for option_idx in range(len(options))) == 1

    for factory in factories:
        capacity_values = capacity_prefix.get(factory.name, [0] * (days + 1))
        for day in range(1, days + 1):
            terms = []
            for idx, options in options_by_idx.items():
                for option_idx, option in enumerate(options):
                    if sv(option.get("source") or option.get("factory")) == factory.name and int(option.get("ship_day") or 1) <= day:
                        terms.append(int(option.get("amount") or 0) * choose[(idx, option_idx)])
            if terms:
                model += pulp.lpSum(terms) <= int(capacity_values[day])

    model += pulp.lpSum(
        float(option.get("freight_cost") or 0.0) * choose[(idx, option_idx)]
        for idx, options in options_by_idx.items()
        for option_idx, option in enumerate(options)
    )
    status_code = model.solve(pulp.PULP_CBC_CMD(msg=False, gapRel=0.0, timeLimit=max(1, int(time_limit_sec))))
    raw_status = pulp.LpStatus[status_code]
    if raw_status != "Optimal":
        return shipments, {
            "method": "FixedArrivalRerouteMILP",
            "status": raw_status,
            "solve_seconds": round(time.time() - started_at, 3),
            "shipments": len(clean_shipments),
            "options": sum(len(options) for options in options_by_idx.values()),
        }

    selected: list[dict[str, Any]] = []
    for idx, options in options_by_idx.items():
        selected_option: dict[str, Any] | None = None
        for option_idx, option in enumerate(options):
            if float(pulp.value(choose[(idx, option_idx)]) or 0.0) >= 0.5:
                selected_option = option
                break
        if selected_option is None:
            selected_option = options[0]
        selected.append({key: value for key, value in selected_option.items() if not str(key).startswith("_")})
    selected = sorted(
        effective_shipments(selected, days),
        key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("destination")), sv(row.get("route"))),
    )
    replay = daily_market_replay(forecasts, selected, days)
    hard_risks = product_transport_hard_risks(
        forecasts=forecasts,
        shipments=selected,
        factories=factories,
        days=days,
        production_rows=production_rows,
    )
    start_cost = sum(float(row.get("freight_cost") or 0.0) for row in clean_shipments)
    selected_cost = sum(float(row.get("freight_cost") or 0.0) for row in selected)
    if hard_risks or not is_full_satisfaction(float(replay.get("market_satisfaction", 0.0)), float(replay.get("shortage", 0.0))):
        return shipments, {
            "method": "FixedArrivalRerouteMILP",
            "status": "OptimalRejectedByReplay",
            "shortage": replay.get("shortage", 0.0),
            "hard_risks": hard_risks,
            "solve_seconds": round(time.time() - started_at, 3),
        }
    return selected, {
        "method": "FixedArrivalRerouteMILP",
        "status": "Optimal",
        "solve_seconds": round(time.time() - started_at, 3),
        "shipments": len(clean_shipments),
        "options": sum(len(options) for options in options_by_idx.values()),
        "freight_cost": round(selected_cost, 4),
        "base_freight": round(start_cost, 4),
        "delta": round(start_cost - selected_cost, 4),
    }


def polish_extreme_transport_cycles(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
    max_cycles: int = 4,
    option_cache: dict[tuple[str, int, int], list[dict[str, Any]]] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Alternate low-cost local optimizers until the full candidate stops improving."""
    if not shipments:
        return shipments, {"cycles": 0, "delta": 0.0, "steps": []}
    rows = [dict(row) for row in shipments]

    def total_cost(items: list[dict[str, Any]]) -> float:
        return sum(float(row.get("freight_cost") or 0.0) for row in items)

    def rows_signature(items: list[dict[str, Any]]) -> tuple[Any, ...]:
        return tuple(
            sorted(
                (
                    sv(row.get("destination")),
                    sv(row.get("factory") or row.get("source")),
                    sv(row.get("route")),
                    sv(row.get("carrier")),
                    int(row.get("ship_day") or 1),
                    int(row.get("arrival_day") or 1),
                    int(row.get("amount") or 0),
                )
                for row in items
            )
        )

    start_cost = total_cost(rows)
    steps: list[dict[str, Any]] = []
    cycles_run = 0
    no_gain_cache: set[tuple[str, tuple[Any, ...]]] = set()
    cache = option_cache if option_cache is not None else {}
    for cycle_idx in range(max_cycles):
        before = total_cost(rows)
        input_signature = rows_signature(rows)
        if ("shift", input_signature) in no_gain_cache:
            shift_status = {"status": "SkippedNoGainCached", "seconds": 0.0, "delta": 0.0}
        else:
            step_started = time.time()
            shifted_rows, shift_status = polish_shift_expensive_early_to_later(
                forecasts=forecasts,
                shipments=rows,
                factories=factories,
                routes=routes,
                products=products,
                cargo=cargo,
                carriers=carriers,
                days=days,
                production_rows=production_rows,
                max_rounds=4,
            )
            shift_status = {**shift_status, "seconds": round(time.time() - step_started, 3)}
            if float(shift_status.get("delta") or 0.0) <= 1e-6:
                no_gain_cache.add(("shift", input_signature))
            if shifted_rows:
                rows = shifted_rows

        input_signature = rows_signature(rows)
        if ("repolish", input_signature) in no_gain_cache:
            repolished_status = {"status": "SkippedNoGainCached", "seconds": 0.0, "delta": 0.0}
        else:
            step_started = time.time()
            repolished_rows, repolished_status = polish_inventory_window_transport_by_destination(
                forecasts=forecasts,
                shipments=rows,
                factories=factories,
                routes=routes,
                products=products,
                cargo=cargo,
                carriers=carriers,
                days=days,
                production_rows=production_rows,
                max_rounds=4,
                include_pair_search=False,
                option_cache=cache,
            )
            repolished_status = {**repolished_status, "seconds": round(time.time() - step_started, 3)}
            if float(repolished_status.get("delta") or 0.0) <= 1e-6:
                no_gain_cache.add(("repolish", input_signature))
            if repolished_rows:
                rows = repolished_rows

        input_signature = rows_signature(rows)
        if ("single", input_signature) in no_gain_cache:
            replacement_status = {"status": "SkippedNoGainCached", "seconds": 0.0, "delta": 0.0}
        else:
            step_started = time.time()
            replaced_rows, replacement_status = polish_single_arrival_replacements(
                forecasts=forecasts,
                shipments=rows,
                factories=factories,
                routes=routes,
                products=products,
                cargo=cargo,
                carriers=carriers,
                days=days,
                production_rows=production_rows,
                max_rounds=4,
                option_cache=cache,
            )
            replacement_status = {**replacement_status, "seconds": round(time.time() - step_started, 3)}
            if float(replacement_status.get("delta") or 0.0) <= 1e-6:
                no_gain_cache.add(("single", input_signature))
            if replaced_rows:
                rows = replaced_rows
        after = total_cost(rows)
        cycles_run = cycle_idx + 1
        steps.append(
            {
                "cycle": cycles_run,
                "delta": round(before - after, 4),
                "shift": shift_status,
                "repolish": repolished_status,
                "single_replacement": replacement_status,
            }
        )
        if after >= before - 1e-6:
            break
    return rows, {
        "cycles": cycles_run,
        "delta": round(start_cost - total_cost(rows), 4),
        "steps": steps,
    }


def polish_destination_blocks(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
    max_block_size: int = 2,
    max_rounds: int = 3,
    option_cache: dict[tuple[str, int, int], list[dict[str, Any]]] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Re-solve destination blocks against the rest of the current plan."""
    if not forecasts or not shipments:
        return shipments, {"rounds": 0, "improvements": 0, "delta": 0.0, "blocks": []}
    forecast_by_node = {sv(row.get("node")): row for row in forecasts}
    nodes = sorted(node for node in forecast_by_node if node)
    if len(nodes) < 2:
        return shipments, {"rounds": 0, "improvements": 0, "delta": 0.0, "blocks": []}
    rows = [dict(row) for row in shipments]
    output_by_factory = factory_output_schedule(factories, days, production_rows)
    capacity_prefix = _factory_capacity_prefix(factories, output_by_factory, days)

    def total_cost(items: list[dict[str, Any]]) -> float:
        return sum(float(row.get("freight_cost") or 0.0) for row in items)

    start_cost = total_cost(rows)
    improvements = 0
    rounds_run = 0
    accepted_blocks: list[dict[str, Any]] = []
    cache = option_cache if option_cache is not None else {}
    max_size = max(2, min(int(max_block_size or 2), len(nodes)))
    blocks_checked = 0
    orders_checked = 0
    for round_idx in range(max_rounds):
        rounds_run = round_idx + 1
        current_cost = total_cost(rows)
        best_rows: list[dict[str, Any]] | None = None
        best_cost = current_cost
        best_block: tuple[str, ...] | None = None
        best_order: tuple[str, ...] | None = None
        for block_size in range(2, max_size + 1):
            for block in iter_combinations(nodes, block_size):
                blocks_checked += 1
                block_set = set(block)
                base_rows = [dict(row) for row in rows if sv(row.get("destination")) not in block_set]
                for order in iter_permutations(block):
                    orders_checked += 1
                    candidate_rows = [dict(row) for row in base_rows]
                    feasible = True
                    for node in order:
                        replacement = _solve_one_destination_inventory_window(
                            forecast=forecast_by_node[node],
                            factories=factories,
                            routes=routes,
                            products=products,
                            cargo=cargo,
                            carriers=carriers,
                            days=days,
                            base_rows=candidate_rows,
                            output_by_factory=output_by_factory,
                            option_cache=cache,
                            capacity_prefix=capacity_prefix,
                        )
                        if replacement is None:
                            feasible = False
                            break
                        candidate_rows.extend(replacement)
                    if not feasible:
                        continue
                    candidate_rows = sorted(
                        effective_shipments(candidate_rows, days),
                        key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("destination")), sv(row.get("route"))),
                    )
                    candidate_cost = total_cost(candidate_rows)
                    if candidate_cost >= best_cost - 1e-6:
                        continue
                    replay = daily_market_replay(forecasts, candidate_rows, days)
                    if not is_full_satisfaction(float(replay.get("market_satisfaction", 0.0)), float(replay.get("shortage", 0.0))):
                        continue
                    hard_risks = product_transport_hard_risks(
                        forecasts=forecasts,
                        shipments=candidate_rows,
                        factories=factories,
                        days=days,
                        production_rows=production_rows,
                    )
                    if hard_risks:
                        continue
                    best_rows = candidate_rows
                    best_cost = candidate_cost
                    best_block = tuple(block)
                    best_order = tuple(order)
        if best_rows is None or best_cost >= current_cost - 1e-6:
            break
        rows = best_rows
        improvements += 1
        accepted_blocks.append(
            {
                "round": rounds_run,
                "block": list(best_block or ()),
                "order": list(best_order or ()),
                "delta": round(current_cost - best_cost, 4),
            }
        )
    return rows, {
        "rounds": rounds_run,
        "improvements": improvements,
        "delta": round(start_cost - total_cost(rows), 4),
        "blocks": accepted_blocks,
        "blocks_checked": blocks_checked,
        "orders_checked": orders_checked,
        "option_cache_entries": len(cache),
    }


def polish_destination_pattern_combinations(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
    max_block_size: int = 3,
    max_patterns_per_destination: int = 8,
    max_combinations: int = 5000,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Collect per-destination schedule patterns and enumerate feasible combinations."""
    if not forecasts or not shipments:
        return shipments, {"patterns": {}, "checked": 0, "improvements": 0, "delta": 0.0}
    forecast_by_node = {sv(row.get("node")): row for row in forecasts}
    nodes = sorted(node for node in forecast_by_node if node)
    if len(nodes) < 2:
        return shipments, {"patterns": {}, "checked": 0, "improvements": 0, "delta": 0.0}
    rows = [dict(row) for row in shipments]
    output_by_factory = factory_output_schedule(factories, days, production_rows)
    capacity_prefix = _factory_capacity_prefix(factories, output_by_factory, days)

    def total_cost(items: list[dict[str, Any]]) -> float:
        return sum(float(row.get("freight_cost") or 0.0) for row in items)

    def pattern_signature(items: list[dict[str, Any]]) -> tuple[Any, ...]:
        return tuple(
            sorted(
                (
                    sv(row.get("destination")),
                    sv(row.get("factory") or row.get("source")),
                    sv(row.get("route")),
                    sv(row.get("carrier")),
                    int(row.get("ship_day") or 1),
                    int(row.get("arrival_day") or 1),
                    int(row.get("amount") or 0),
                )
                for row in items
            )
        )

    def clean_pattern(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return [
            {key: value for key, value in dict(row).items() if not str(key).startswith("_")}
            for row in items
        ]

    option_cache: dict[tuple[str, int, int], list[dict[str, Any]]] = {}
    patterns: dict[str, dict[tuple[Any, ...], list[dict[str, Any]]]] = defaultdict(dict)
    independent_patterns: dict[str, list[dict[str, Any]]] = {}
    for node in nodes:
        current = clean_pattern([dict(row) for row in rows if sv(row.get("destination")) == node])
        if current:
            patterns[node][pattern_signature(current)] = current
        alone = _solve_one_destination_inventory_window(
            forecast=forecast_by_node[node],
            factories=factories,
            routes=routes,
            products=products,
            cargo=cargo,
            carriers=carriers,
            days=days,
            base_rows=[],
            output_by_factory=output_by_factory,
            option_cache=option_cache,
            capacity_prefix=capacity_prefix,
            beam_width=inventory_window_beam_width(cargo),
            arrival_options_per_state=inventory_window_arrival_options(cargo),
        )
        if alone:
            alone = clean_pattern(alone)
            independent_patterns[node] = alone
            patterns[node][pattern_signature(alone)] = alone

    shadow_pattern_count = 0
    shadow_strengths = [10.0, 35.0, 90.0]
    if heatwater_pattern_audit_enabled():
        shadow_strengths.extend([180.0, 360.0, 720.0, 1200.0, 2000.0])
    for strength in shadow_strengths:
        penalties = _factory_shadow_penalties(
            rows=rows,
            factories=factories,
            output_by_factory=output_by_factory,
            days=days,
            strength=strength,
        )
        if not penalties:
            continue
        for node in nodes:
            shadow_rows = _solve_one_destination_inventory_window(
                forecast=forecast_by_node[node],
                factories=factories,
                routes=routes,
                products=products,
                cargo=cargo,
                carriers=carriers,
                days=days,
                base_rows=[],
                output_by_factory=output_by_factory,
                option_cache=None,
                capacity_prefix=capacity_prefix,
                factory_day_penalty=penalties,
            )
            if shadow_rows:
                shadow_rows = clean_pattern(shadow_rows)
                signature = pattern_signature(shadow_rows)
                if signature not in patterns[node]:
                    patterns[node][signature] = shadow_rows
                    shadow_pattern_count += 1

    avoidance_pattern_count = 0
    if heatwater_pattern_audit_enabled():
        for factory in factories:
            for strength in (250.0, 750.0, 1500.0, 3000.0, 6000.0):
                avoid_penalties = {
                    (factory.name, day): float(strength) / max(1, days - day + 1)
                    for day in range(1, days + 1)
                }
                for node in nodes:
                    avoid_rows = _solve_one_destination_inventory_window(
                        forecast=forecast_by_node[node],
                        factories=factories,
                        routes=routes,
                        products=products,
                        cargo=cargo,
                        carriers=carriers,
                        days=days,
                        base_rows=[],
                        output_by_factory=output_by_factory,
                        option_cache=None,
                        capacity_prefix=capacity_prefix,
                        factory_day_penalty=avoid_penalties,
                    )
                    if avoid_rows:
                        avoid_rows = clean_pattern(avoid_rows)
                        signature = pattern_signature(avoid_rows)
                        if signature not in patterns[node]:
                            patterns[node][signature] = avoid_rows
                            avoidance_pattern_count += 1

    conflict_shadow_pattern_count = 0
    independent_rows = [
        dict(row)
        for pattern in independent_patterns.values()
        for row in pattern
    ]
    if independent_rows:
        for strength in (80.0, 160.0, 320.0, 640.0, 1000.0, 1600.0, 2500.0, 4000.0):
            conflict_penalties: dict[tuple[str, int], float] = {}
            for factory in factories:
                capacity_values = capacity_prefix.get(factory.name, [0] * (days + 1))
                daily_scale = max(1.0, float(factory.daily or 0.0) * 3.0)
                for day in range(1, days + 1):
                    used = sum(
                        int(row.get("amount") or 0)
                        for row in independent_rows
                        if sv(row.get("source") or row.get("factory")) == factory.name
                        and int(row.get("ship_day") or 1) <= day
                    )
                    capacity = float(capacity_values[day] if day < len(capacity_values) else 0.0)
                    overflow = max(0.0, float(used) - capacity)
                    if overflow <= 1e-6:
                        continue
                    remaining = max(1, days - day + 1)
                    conflict_penalties[(factory.name, day)] = float(strength) * overflow / daily_scale / remaining
            if not conflict_penalties:
                continue
            for node in nodes:
                conflict_rows = _solve_one_destination_inventory_window(
                    forecast=forecast_by_node[node],
                    factories=factories,
                    routes=routes,
                    products=products,
                    cargo=cargo,
                    carriers=carriers,
                    days=days,
                    base_rows=[],
                    output_by_factory=output_by_factory,
                    option_cache=None,
                    capacity_prefix=capacity_prefix,
                    factory_day_penalty=conflict_penalties,
                )
                if conflict_rows:
                    conflict_rows = clean_pattern(conflict_rows)
                    signature = pattern_signature(conflict_rows)
                    if signature not in patterns[node]:
                        patterns[node][signature] = conflict_rows
                        conflict_shadow_pattern_count += 1

    max_size = max(2, min(int(max_block_size or 2), len(nodes)))
    for block_size in range(2, max_size + 1):
        for block in iter_combinations(nodes, block_size):
            base_rows = [dict(row) for row in rows if sv(row.get("destination")) not in set(block)]
            for order in iter_permutations(block):
                candidate_rows = [dict(row) for row in base_rows]
                by_node: dict[str, list[dict[str, Any]]] = {}
                feasible = True
                for node in order:
                    replacement = _solve_one_destination_inventory_window(
                        forecast=forecast_by_node[node],
                        factories=factories,
                        routes=routes,
                        products=products,
                        cargo=cargo,
                        carriers=carriers,
                        days=days,
                        base_rows=candidate_rows,
                        output_by_factory=output_by_factory,
                        option_cache=option_cache,
                        capacity_prefix=capacity_prefix,
                        beam_width=inventory_window_beam_width(cargo),
                        arrival_options_per_state=inventory_window_arrival_options(cargo),
                    )
                    if replacement is None:
                        feasible = False
                        break
                    replacement = clean_pattern(replacement)
                    by_node[node] = replacement
                    candidate_rows.extend(replacement)
                if not feasible:
                    continue
                for node, pattern in by_node.items():
                    patterns[node][pattern_signature(pattern)] = pattern

    compact_patterns: dict[str, list[list[dict[str, Any]]]] = {}
    independent_lower_bound = 0.0
    for node in nodes:
        sorted_patterns = sorted(patterns[node].values(), key=lambda pattern: (total_cost(pattern), len(pattern)))
        if sorted_patterns:
            independent_lower_bound += total_cost(sorted_patterns[0])
        compact_patterns[node] = sorted_patterns[: max(1, int(max_patterns_per_destination or 1))]

    combination_count = 1
    for node in nodes:
        combination_count *= max(1, len(compact_patterns.get(node, [])))

    milp_status: dict[str, Any] = {"status": "Skipped"}
    milp_rows: list[dict[str, Any]] | None = None
    milp_cost = total_cost(rows)
    try:
        import pulp

        model = pulp.LpProblem("destination_pattern_selection", pulp.LpMinimize)
        choose: dict[tuple[str, int], Any] = {}
        for node in nodes:
            node_patterns = compact_patterns.get(node, [])
            if not node_patterns:
                raise ValueError(f"{node} has no candidate patterns")
            for idx, _pattern in enumerate(node_patterns):
                choose[(node, idx)] = pulp.LpVariable(f"pat_{len(choose)}", lowBound=0, upBound=1, cat="Binary")
            model += pulp.lpSum(choose[(node, idx)] for idx in range(len(node_patterns))) == 1

        for factory in factories:
            capacity_values = capacity_prefix.get(factory.name, [0] * (days + 1))
            for day in range(1, days + 1):
                terms = []
                for node in nodes:
                    for idx, pattern in enumerate(compact_patterns.get(node, [])):
                        used = sum(
                            int(row.get("amount") or 0)
                            for row in pattern
                            if sv(row.get("source") or row.get("factory")) == factory.name
                            and int(row.get("ship_day") or 1) <= day
                        )
                        if used:
                            terms.append(used * choose[(node, idx)])
                if terms:
                    model += pulp.lpSum(terms) <= int(capacity_values[day])

        model += pulp.lpSum(
            total_cost(pattern) * choose[(node, idx)]
            for node in nodes
            for idx, pattern in enumerate(compact_patterns.get(node, []))
        )
        status_code = model.solve(pulp.PULP_CBC_CMD(msg=False, gapRel=0.0, timeLimit=60))
        status = pulp.LpStatus[status_code]
        milp_status = {
            "status": status,
            "patterns": {node: len(compact_patterns.get(node, [])) for node in nodes},
            "shadow_patterns": shadow_pattern_count,
            "variables": len(choose),
            "combination_equivalent": combination_count,
        }
        if status == "Optimal":
            selected: list[dict[str, Any]] = []
            for node in nodes:
                for idx, pattern in enumerate(compact_patterns.get(node, [])):
                    if float(pulp.value(choose[(node, idx)]) or 0.0) >= 0.5:
                        selected.extend(dict(row) for row in pattern)
                        break
            selected = sorted(
                effective_shipments(clean_pattern(selected), days),
                key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("destination")), sv(row.get("route"))),
            )
            replay = daily_market_replay(forecasts, selected, days)
            hard_risks = product_transport_hard_risks(
                forecasts=forecasts,
                shipments=selected,
                factories=factories,
                days=days,
                production_rows=production_rows,
            )
            if is_full_satisfaction(float(replay.get("market_satisfaction", 0.0)), float(replay.get("shortage", 0.0))) and not hard_risks:
                milp_rows = selected
                milp_cost = total_cost(selected)
                milp_status = {
                    **milp_status,
                    "selected_cost": round(milp_cost, 4),
                    "hard_risks": 0,
                }
            else:
                milp_status = {
                    **milp_status,
                    "status": "OptimalRejectedByReplay",
                    "shortage": replay.get("shortage", 0.0),
                    "hard_risks": hard_risks,
                }
    except Exception as exc:
        milp_status = {"status": "Unavailable", "error": str(exc)}

    if milp_rows is not None and milp_cost < total_cost(rows) - 1e-6:
        return milp_rows, {
            "patterns": {node: len(compact_patterns.get(node, [])) for node in nodes},
            "checked": 0,
            "status": "PatternMILPImproved",
            "pattern_milp": milp_status,
            "shadow_patterns": shadow_pattern_count,
            "avoidance_patterns": avoidance_pattern_count,
            "conflict_shadow_patterns": conflict_shadow_pattern_count,
            "option_cache_entries": len(option_cache),
            "combinations": combination_count,
            "improvements": 1,
            "delta": round(total_cost(rows) - milp_cost, 4),
            "independent_lower_bound": round(independent_lower_bound, 4),
            "gap_to_independent_lower_bound": round(milp_cost - independent_lower_bound, 4),
        }

    if combination_count > max_combinations:
        return rows, {
            "patterns": {node: len(compact_patterns.get(node, [])) for node in nodes},
            "checked": 0,
            "status": "SkippedTooManyCombinations",
            "pattern_milp": milp_status,
            "shadow_patterns": shadow_pattern_count,
            "avoidance_patterns": avoidance_pattern_count,
            "conflict_shadow_patterns": conflict_shadow_pattern_count,
            "option_cache_entries": len(option_cache),
            "combinations": combination_count,
            "improvements": 0,
            "delta": 0.0,
            "independent_lower_bound": round(independent_lower_bound, 4),
            "gap_to_independent_lower_bound": round(total_cost(rows) - independent_lower_bound, 4),
        }

    start_cost = total_cost(rows)
    best_rows = rows
    best_cost = start_cost
    checked = 0
    feasible_improving = 0
    for combo in iter_product(*(compact_patterns[node] for node in nodes)):
        checked += 1
        candidate: list[dict[str, Any]] = []
        for pattern in combo:
            candidate.extend(dict(row) for row in pattern)
        candidate_cost = total_cost(candidate)
        if candidate_cost >= best_cost - 1e-6:
            continue
        replay = daily_market_replay(forecasts, candidate, days)
        if not is_full_satisfaction(float(replay.get("market_satisfaction", 0.0)), float(replay.get("shortage", 0.0))):
            continue
        hard_risks = product_transport_hard_risks(
            forecasts=forecasts,
            shipments=candidate,
            factories=factories,
            days=days,
            production_rows=production_rows,
        )
        if hard_risks:
            continue
        feasible_improving += 1
        best_rows = sorted(
            effective_shipments(candidate, days),
            key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("destination")), sv(row.get("route"))),
        )
        best_cost = candidate_cost
    return best_rows, {
        "patterns": {node: len(compact_patterns.get(node, [])) for node in nodes},
        "checked": checked,
        "feasible_improving": feasible_improving,
        "improvements": 1 if best_cost < start_cost - 1e-6 else 0,
        "delta": round(start_cost - best_cost, 4),
        "pattern_milp": milp_status,
        "shadow_patterns": shadow_pattern_count,
        "avoidance_patterns": avoidance_pattern_count,
        "conflict_shadow_patterns": conflict_shadow_pattern_count,
        "option_cache_entries": len(option_cache),
        "independent_lower_bound": round(independent_lower_bound, 4),
        "gap_to_independent_lower_bound": round(best_cost - independent_lower_bound, 4),
    }


def shipment_shift_delta_candidates(
    *,
    from_row: dict[str, Any],
    to_row: dict[str, Any],
    max_delta: int,
    route_by_name: dict[str, Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
) -> list[int]:
    """Integer deltas worth testing when moving quantity between two shipments.

    Transport cost is piecewise linear because each carrier segment has a hard
    lowest charge. Sampling only round numbers can miss the cheapest point just
    before or after a lowest-charge breakpoint, so include breakpoints for both
    the reduced row and the increased row.
    """
    limit = int(max_delta or 0)
    if limit <= 0:
        return []

    amount_from = int(from_row.get("amount") or 0)
    amount_to = int(to_row.get("amount") or 0)
    if amount_from <= 0:
        return []

    points: set[int] = {1, limit}
    if limit <= 500:
        points.update(range(1, limit + 1))
    else:
        points.update({2, 3, 5, 8, 10, 15, 20, 30, 50, 80, 100, 150, 200, 240, 300, 400, 500})

    ratio = charge_ratio(products, cargo)

    def add_near_delta(raw_delta: float) -> None:
        for candidate in (
            math.floor(raw_delta) - 2,
            math.floor(raw_delta) - 1,
            math.floor(raw_delta),
            math.ceil(raw_delta),
            math.ceil(raw_delta) + 1,
            math.ceil(raw_delta) + 2,
        ):
            if 1 <= candidate <= limit:
                points.add(int(candidate))

    def add_breakpoints_for_row(row: dict[str, Any], max_amount: int, mapper: Any) -> None:
        route = route_by_name.get(sv(row.get("route")))
        if route is None or max_amount <= 0:
            return
        options = route_transport_options(
            route,
            ratio,
            carriers,
            max_amount=max(float(max_amount), 1.0),
            prune_dominated=False,
        )
        for option in options:
            for breakpoint in option_breakpoints(option, max(float(max_amount), 1.0)):
                add_near_delta(float(mapper(breakpoint)))

    add_breakpoints_for_row(from_row, amount_from, lambda qty: amount_from - float(qty))
    add_breakpoints_for_row(to_row, amount_to + limit, lambda qty: float(qty) - amount_to)
    return sorted(delta for delta in points if 1 <= delta <= limit)


def score_search_service_levels(days: int) -> list[float]:
    if days > 45:
        return [1.0, 0.95, 0.90]
    return [1.0, 0.97, 0.95, 0.90]


def build_score_aware_product_transport(
    *,
    forecasts: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    xls_path: Path,
    qtype: str,
    days: int,
    name: str,
    production_rows: list[dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not forecasts or not factories:
        return [], {"method": "ScoreAwareServiceLevelMILP", "status": "NoDemandOrSource"}
    output_by_factory = factory_output_schedule(factories, days, production_rows)
    sources = {
        factory.name: {
            "initial": factory.init,
            "supply": output_by_factory.get(factory.name, [0] * days),
            "supply_is_capacity": production_rows is None,
            "limit": 0.0,
        }
        for factory in factories
    }
    destinations = {
        row["node"]: {
            "initial": row["init"],
            "demand": daily_demands_for_forecast(row, days),
            "limit": row["limit"],
            "excess_fee": row.get("excess_fee", 0.0),
        }
        for row in forecasts
    }
    total_demand = sum(sum(info.get("demand", [])) for info in destinations.values())
    if total_demand <= 0:
        return [], {"method": "ScoreAwareServiceLevelMILP", "status": "NoDemand"}
    usable_routes = pruned_product_routes(
        forecasts=forecasts,
        factories=factories,
        routes=routes,
        products=products,
        cargo=cargo,
        carriers=carriers,
        top_k=2 if days > 45 else 3,
    )
    usable_routes = [route for route in usable_routes if route.src in sources and route.dst in destinations]
    if not usable_routes:
        return [], {"method": "ScoreAwareServiceLevelMILP", "status": "NoRoute"}

    best_rows: list[dict[str, Any]] = []
    best_status: dict[str, Any] = {"method": "ScoreAwareServiceLevelMILP", "status": "NoCandidate"}
    best_metrics: dict[str, float] | None = None
    summaries: list[dict[str, Any]] = []
    for service_level in score_search_service_levels(days):
        max_shortage = math.floor(total_demand * max(0.0, 1.0 - service_level))
        transport = _solve_day_transport_milp(
            name=f"{name}_svc_{int(service_level * 1000)}",
            sources=sources,
            destinations=destinations,
            routes=usable_routes,
            products=products,
            cargo=cargo,
            carriers=carriers,
            days=days,
            ship_day_step=6 if days > 45 else 2,
            gap_rel=0.02 if days > 45 else 0.005,
            enforce_destination_limits=True,
            max_total_shortage=max_shortage,
            shortage_penalty=0.0,
            time_limit_sec=120 if qtype == "生产" and is_extreme_mode() else 20,
        )
        rows = [
            {
                "destination": row["destination"],
                "factory": row["source"],
                "source": row["source"],
                "cargo": row["cargo"],
                "amount": int(row["amount"]),
                "route": row["route"],
                "mode": row["mode"],
                "carrier": row.get("carrier", ""),
                "lead": int(row["lead"]),
                "ship_day": int(row["ship_day"]),
                "arrival_day": int(row["arrival_day"]),
                "freight_cost": float(row["freight_cost"]),
                "note": "平台分数服务率枚举",
            }
            for row in transport.get("shipments", [])
        ]
        rows = effective_shipments(rows, days)
        metrics = score_product_transport_candidate(
            forecasts=forecasts,
            shipments=rows,
            xls_path=xls_path,
            qtype=qtype,
            days=days,
        )
        hard_risks = product_transport_hard_risks(
            forecasts=forecasts,
            shipments=rows,
            factories=factories,
            days=days,
            production_rows=production_rows,
        )
        summaries.append(
            {
                "service_level": round(service_level, 4),
                "status": transport.get("status"),
                "score": round(metrics["score"], 4),
                "unit_logistics": round(metrics["unit_logistics"], 4),
                "market_satisfaction": round(metrics["market_satisfaction"], 4),
                "shortage": round(metrics["shortage"], 4),
                "hard_risks": len(hard_risks),
                "shipments": len(rows),
            }
        )
        if hard_risks:
            continue
        if best_metrics is None or product_transport_priority(metrics, len(rows)) > product_transport_priority(best_metrics, len(best_rows)):
            best_rows = rows
            best_metrics = metrics
            best_status = {
                "method": "ScoreAwareServiceLevelMILP",
                "status": transport.get("status"),
                "service_level": service_level,
                "max_total_shortage": max_shortage,
                "solver_shortage": transport.get("shortage", 0.0),
            }
    best_status["service_candidates"] = summaries
    return best_rows, best_status


def build_budget_score_product_transport(
    *,
    forecasts: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    xls_path: Path,
    qtype: str,
    days: int,
    name: str,
    production_rows: list[dict[str, Any]] | None = None,
    warm_start_shipments: list[dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not forecasts or not factories:
        return [], {"method": "ScoreBudgetMILP", "status": "NoDemandOrSource"}
    targets = score_targets(xls_path, qtype)
    target_logistics = float(targets.get("unit_logistics_cost", 0.0) or 0.0)
    if target_logistics <= 0:
        return [], {"method": "ScoreBudgetMILP", "status": "NoLogisticsTarget"}
    output_by_factory = factory_output_schedule(factories, days, production_rows)
    sources = {
        factory.name: {
            "initial": factory.init,
            "supply": output_by_factory.get(factory.name, [0] * days),
            "supply_is_capacity": production_rows is None,
            "limit": 0.0,
        }
        for factory in factories
    }
    destinations = {
        row["node"]: {
            "initial": row["init"],
            "demand": daily_demands_for_forecast(row, days),
            "limit": row["limit"],
            "excess_fee": row.get("excess_fee", 0.0),
        }
        for row in forecasts
    }
    usable_routes = pruned_product_routes(
        forecasts=forecasts,
        factories=factories,
        routes=routes,
        products=products,
        cargo=cargo,
        carriers=carriers,
        top_k=3 if days <= 45 else 2,
    )
    usable_routes = [route for route in usable_routes if route.src in sources and route.dst in destinations]
    if not usable_routes:
        return [], {"method": "ScoreBudgetMILP", "status": "NoRoute"}

    if is_extreme_mode():
        cost_factors = [1.0]
    elif qtype == "销售" and days > 45:
        cost_factors = [0.0]
    else:
        cost_factors = [1.0, 1.02, 1.05]
    best_rows: list[dict[str, Any]] = []
    best_status: dict[str, Any] = {"method": "ScoreBudgetMILP", "status": "NoCandidate"}
    best_metrics: dict[str, float] | None = None
    summaries: list[dict[str, Any]] = []
    for factor in cost_factors:
        budget = target_logistics * factor if factor > 0 else 0.0
        transport = _solve_day_transport_milp(
            name=f"{name}_budget_{int(factor * 100)}",
            sources=sources,
            destinations=destinations,
            routes=usable_routes,
            products=products,
            cargo=cargo,
            carriers=carriers,
            days=days,
            ship_day_step=(1 if is_extreme_mode() else 3) if days > 45 else 1,
            gap_rel=0.0 if is_extreme_mode() else (0.02 if days > 45 else 0.01),
            enforce_destination_limits=True,
            shortage_penalty=100_000_000.0,
            max_total_shortage=0.0,
            freight_budget_per_served=budget if qtype == "销售" and factor > 0 else None,
            freight_budget_per_shipped=budget if qtype != "销售" and factor > 0 else None,
            allow_fallback=False,
            time_limit_sec=None if is_extreme_mode() else (180 if qtype == "生产" else 60),
            warm_start_shipments=warm_start_shipments,
        )
        rows = [
            {
                "destination": row["destination"],
                "factory": row["source"],
                "source": row["source"],
                "cargo": row["cargo"],
                "amount": int(row["amount"]),
                "route": row["route"],
                "mode": row["mode"],
                "carrier": row.get("carrier", ""),
                "lead": int(row["lead"]),
                "ship_day": int(row["ship_day"]),
                "arrival_day": int(row["arrival_day"]),
                "freight_cost": float(row["freight_cost"]),
                "note": "平台分数成本预算优化",
            }
            for row in transport.get("shipments", [])
        ]
        rows = effective_shipments(rows, days)
        metrics = score_product_transport_candidate(
            forecasts=forecasts,
            shipments=rows,
            xls_path=xls_path,
            qtype=qtype,
            days=days,
        )
        hard_risks = product_transport_hard_risks(
            forecasts=forecasts,
            shipments=rows,
            factories=factories,
            days=days,
            production_rows=production_rows,
        )
        summaries.append(
            {
                "cost_factor": round(factor, 4),
                "status": transport.get("status"),
                "score": round(metrics["score"], 4),
                "unit_logistics": round(metrics["unit_logistics"], 4),
                "market_satisfaction": round(metrics["market_satisfaction"], 4),
                "shortage": round(metrics["shortage"], 4),
                "hard_risks": len(hard_risks),
                "shipments": len(rows),
            }
        )
        if hard_risks or not rows:
            continue
        if not is_extreme_mode() and not is_full_satisfaction(metrics["market_satisfaction"], metrics["shortage"]):
            continue
        if best_metrics is None or product_transport_priority(metrics, len(rows)) > product_transport_priority(best_metrics, len(best_rows)):
            best_rows = rows
            best_metrics = metrics
            best_status = {
                "method": "ScoreBudgetMILP",
                "status": transport.get("status"),
                "cost_factor": factor if factor > 0 else None,
                "unit_budget": budget if factor > 0 else None,
                "solver_shortage": transport.get("shortage", 0.0),
            }
    best_status["budget_candidates"] = summaries
    return best_rows, best_status


def shipment_assignment(
    *,
    forecasts: list[dict[str, Any]],
    factories: list[Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None = None,
) -> dict[str, Factory]:
    ratio = charge_ratio(products, cargo)
    assignment: dict[str, Factory] = {}
    for forecast in forecasts:
        ranked: list[tuple[float, Factory]] = []
        for factory in factories:
            route = pick_best_route(routes, factory.name, forecast["node"], max(forecast["forecast"], 1.0), ratio, urgent=True, carriers=carriers)
            if not route:
                continue
            ranked.append((route_score(route, max(forecast["forecast"], 1.0), ratio, urgent=True, carriers=carriers), factory))
        if ranked:
            assignment[forecast["node"]] = min(ranked, key=lambda item: item[0])[1]
    return assignment


def score_product_transport_candidate(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    xls_path: Path,
    qtype: str,
    days: int,
) -> dict[str, float]:
    replay = daily_market_replay(forecasts, shipments, days)
    freight = sum(float(row.get("freight_cost") or 0.0) for row in shipments)
    if qtype == "销售":
        targets = score_targets(xls_path, "销售")
        points = score_points(xls_path, "销售")
        denominator = replay["served"]
        unit = freight / max(denominator, 0.001)
        score = deviation_score(0.0, targets["prediction_deviation"], points["预测偏差率"])
        score += cost_score(unit, targets["unit_logistics_cost"], points["单位物流成本"])
        score += satisfaction_score(replay["market_satisfaction"], points["市场满足率"])
    else:
        targets = score_targets(xls_path, "生产")
        points = score_points(xls_path, "生产")
        denominator = sum(float(row.get("amount") or 0.0) for row in shipments)
        unit = freight / max(denominator, 0.001)
        score = cost_score(unit, targets.get("unit_logistics_cost", 0.0), points["单位物流成本"])
        score += satisfaction_score(replay["market_satisfaction"], points["市场满足率"])
    return {
        "score": float(score),
        "unit_logistics": float(unit),
        "market_satisfaction": float(replay["market_satisfaction"]),
        "shortage": float(replay["shortage"]),
        "served": float(replay["served"]),
        "freight": float(freight),
    }


def product_transport_hard_risks(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    factories: list[Factory],
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
) -> list[str]:
    risks: list[str] = []
    risks.extend(
        factory_finished_goods_risks(
            shipments=shipments,
            factories=factories,
            days=days,
            production_rows=production_rows,
        )
    )

    replay = daily_market_replay(forecasts, shipments, days) if forecasts else None
    if replay:
        risks.extend(risk for risk in replay.get("risks", []) if "库存超过上限" in risk)
    return risks


def factory_finished_goods_risks(
    *,
    shipments: list[dict[str, Any]],
    factories: list[Factory],
    days: int,
    production_rows: list[dict[str, Any]] | None = None,
) -> list[str]:
    risks: list[str] = []
    shipped_by_factory_day: dict[tuple[str, int], float] = defaultdict(float)
    for row in shipments:
        factory_name = sv(row.get("factory") or row.get("source"))
        if not factory_name:
            continue
        day = max(1, min(days, int(row.get("ship_day") or 1)))
        shipped_by_factory_day[(factory_name, day)] += float(row.get("amount") or 0.0)

    output_by_factory = factory_output_schedule(factories, days, production_rows)
    for factory in factories:
        inventory = float(factory.init or 0.0)
        min_inventory = inventory
        first_negative_day = None
        for day in range(1, days + 1):
            inventory += float(output_by_factory.get(factory.name, [0] * days)[day - 1])
            inventory -= shipped_by_factory_day.get((factory.name, day), 0.0)
            min_inventory = min(min_inventory, inventory)
            if inventory < -1e-6 and first_negative_day is None:
                first_negative_day = day
        if first_negative_day is not None:
            risks.append(f"{factory.name} 第{first_negative_day}天成品库存为负，最低 {min_inventory:,.0f}")
    return risks


def polish_product_transport(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    assignment: dict[str, Factory],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    factories: list[Factory],
    xls_path: Path,
    qtype: str,
    days: int,
    fill_rounds: int = 2,
    production_rows: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    rows = repair_factory_ship_days([dict(row) for row in shipments], factories, days, production_rows)
    rows = fill_market_shortages_with_fast_routes(
        forecasts=forecasts,
        shipments=rows,
        assignment=assignment,
        routes=routes,
        products=products,
        cargo=cargo,
        carriers=carriers,
        factories=factories,
        days=days,
        max_rounds=fill_rounds,
        production_rows=production_rows,
    )
    rows = repair_factory_ship_days(rows, factories, days, production_rows)
    rows = trim_expensive_shipments_for_score(
        forecasts=forecasts,
        shipments=rows,
        xls_path=xls_path,
        qtype=qtype,
        days=days,
    )
    return sorted(rows, key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("route")), sv(row.get("destination"))))


def fill_market_shortages_score_aware(
    *,
    forecasts: list[dict[str, Any]],
    shipments: list[dict[str, Any]],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    factories: list[Factory],
    xls_path: Path,
    qtype: str,
    days: int,
    max_additions: int = 16,
) -> list[dict[str, Any]]:
    if qtype != "销售" or not forecasts or not shipments:
        return shipments
    rows = effective_shipments([dict(row) for row in shipments], days)
    ratio = charge_ratio(products, cargo)

    def metrics_for(rows_to_score: list[dict[str, Any]]) -> dict[str, float]:
        return score_product_transport_candidate(
            forecasts=forecasts,
            shipments=rows_to_score,
            xls_path=xls_path,
            qtype=qtype,
            days=days,
        )

    current_metrics = metrics_for(rows)
    finished_goods_total = {
        factory.name: int(math.floor(float(factory.init or 0.0) + float(factory.daily or 0.0) * days))
        for factory in factories
    }
    for _ in range(max_additions):
        replay = daily_market_replay(forecasts, rows, days)
        shortage_by_node: dict[str, int] = defaultdict(int)
        first_short_day: dict[str, int] = {}
        for day_row in replay.get("daily_rows", []):
            shortage = ceil_int(day_row.get("shortage") or 0)
            if shortage <= 0:
                continue
            node = sv(day_row.get("node"))
            shortage_by_node[node] += shortage
            first_short_day.setdefault(node, int(day_row.get("day") or 1))
        if not shortage_by_node:
            break

        shipped_by_source: dict[str, int] = defaultdict(int)
        for row in rows:
            shipped_by_source[sv(row.get("source") or row.get("factory"))] += int(row.get("amount") or 0)

        best: tuple[tuple[float, float, float, float], dict[str, Any], dict[str, float]] | None = None
        for node, shortage in sorted(shortage_by_node.items(), key=lambda item: item[1], reverse=True):
            if shortage <= 0:
                continue
            for factory in factories:
                remaining = finished_goods_total.get(factory.name, 0) - shipped_by_source.get(factory.name, 0)
                if remaining <= 0:
                    continue
                for route in lane_routes(routes, factory.name, node):
                    amount = min(shortage, remaining)
                    amount = int(amount)
                    if amount <= 0:
                        continue
                    lead = route_effective_lead(route, amount, ratio, carriers)
                    ship_day = max(1, min(days, first_short_day[node] - lead))
                    arrival_day = ship_day + lead
                    if arrival_day > days:
                        continue
                    candidate = {
                        "destination": node,
                        "source": factory.name,
                        "factory": factory.name,
                        "cargo": cargo,
                        "amount": amount,
                        "route": route.route,
                        "mode": route_mode(route),
                        "carrier": route_carrier(route, carriers, amount, ratio),
                        "lead": lead,
                        "ship_day": ship_day,
                        "arrival_day": arrival_day,
                        "freight_cost": route_cost(route, amount, ratio, carriers),
                        "note": "评分预算补货",
                    }
                    candidate_metrics = metrics_for(rows + [candidate])
                    priority = product_transport_priority(candidate_metrics, len(rows) + 1)
                    unit_cost = float(candidate["freight_cost"]) / max(amount, 1)
                    item = (priority, candidate, candidate_metrics)
                    if priority > product_transport_priority(current_metrics, len(rows)) and (
                        best is None or (priority, -unit_cost) > (best[0], -float(best[1].get("freight_cost") or 0.0) / max(float(best[1].get("amount") or 0.0), 1.0))
                    ):
                        best = item
        if best is None:
            break
        rows.append(best[1])
        current_metrics = best[2]

    return sorted(rows, key=lambda row: (int(row.get("ship_day") or 1), sv(row.get("route")), sv(row.get("destination"))))


def _transport_greedy_fallback(
    *,
    status: str,
    sources: dict[str, dict[str, Any]],
    destinations: dict[str, dict[str, Any]],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None,
    days: int,
) -> dict[str, Any]:
    ratio = charge_ratio(products, cargo)
    source_remaining = {
        src: ceil_int(float(info.get("initial", 0.0) or 0.0) + sum(float(v or 0.0) for v in info.get("supply", [])))
        for src, info in sources.items()
    }
    shipments: list[dict[str, Any]] = []
    failures: list[str] = []
    for dst, info in destinations.items():
        need = ceil_int(max(0.0, sum(info.get("demand", [])) - float(info.get("initial", 0.0) or 0.0)))
        if need <= 0:
            continue
        candidates = [
            route for route in routes
            if route.dst == dst and route.src in sources and source_remaining.get(route.src, 0) > 0
        ]
        candidates.sort(key=lambda route: (route_unit_cost(route, max(need, 1.0), ratio, carriers), route_effective_lead(route, max(need, 1.0), ratio, carriers)))
        for route in candidates:
            if need <= 0:
                break
            available = source_remaining.get(route.src, 0)
            if available <= 0:
                continue
            amount = min(need, available)
            amount = ceil_int(amount)
            if amount <= 0:
                continue
            ship_day = max(1, min(days, 1))
            lead = route_effective_lead(route, amount, ratio, carriers)
            shipments.append({
                "cargo": cargo,
                "source": route.src,
                "destination": route.dst,
                "amount": amount,
                "ship_day": ship_day,
                "arrival_day": ship_day + lead,
                "route": route.route,
                "mode": route_mode(route),
                "lead": lead,
                "freight_cost": route_cost(route, amount, ratio, carriers),
                "carrier": route_carrier(route, carriers, amount, ratio),
            })
            source_remaining[route.src] = available - amount
            need -= amount
        if need > 0:
            failures.append(f"{cargo}-{dst} fallback 总量缺口 {need:.0f}")
    return {
        "status": f"{status}-GreedyFallback",
        "shipments": sorted(shipments, key=lambda row: (row["ship_day"], row["route"], row["amount"])),
        "freight_cost": sum(row["freight_cost"] for row in shipments),
        "shortage": 0.0,
        "failures": failures,
        "source_supply": {src: 0 for src in sources},
    }


def material_plan_product_units(factory_materials: list[FactoryMaterial], products: list[Product], days: int) -> dict[str, float]:
    by_factory: dict[str, list[float]] = defaultdict(list)
    for item in factory_materials:
        bom = material_bom(products, item.material)
        if bom > 0:
            by_factory[item.factory].append(float(item.daily or 0.0) * days / bom)
    return {factory: min(values) for factory, values in by_factory.items() if values}


def initial_material_inventory_cost(
    factory_materials: list[FactoryMaterial],
    suppliers: list[Supplier],
    rates: dict[tuple[str, str], float],
) -> float:
    price_by_material: dict[str, float] = {}
    for supplier in suppliers:
        if supplier.material and supplier.material not in price_by_material:
            price_by_material[supplier.material] = currency_to_cny(supplier.price, supplier.currency, rates)
    return sum(float(item.init or 0.0) * price_by_material.get(item.material, 0.0) for item in factory_materials)


def supplier_supply_profile(supplier: Supplier, days: int) -> list[float]:
    if supplier.available > 0:
        remaining = max(0.0, float(supplier.available) - float(supplier.init or 0.0))
    else:
        remaining = max(0.0, float(supplier.daily or 0.0) * days)
    profile: list[float] = []
    for _day in range(1, days + 1):
        produced = min(float(supplier.daily or 0.0), remaining)
        profile.append(max(0.0, produced))
        remaining -= produced
    return profile


def solve_material_procurement_daily(
    *,
    material: FactoryMaterial,
    daily_demand: list[float],
    products: list[Product],
    suppliers: list[Supplier],
    routes: list[Route],
    rates: dict[tuple[str, str], float],
    carriers: list[Any] | None,
    days: int,
    name: str,
) -> dict[str, Any]:
    supplier_sources: dict[str, dict[str, Any]] = {}
    supplier_by_name: dict[str, Supplier] = {}
    for supplier in suppliers:
        if supplier.material != material.material:
            continue
        supplier_by_name[supplier.name] = supplier
        supplier_sources[supplier.name] = {
            "initial": supplier.init,
            "supply": supplier_supply_profile(supplier, days),
            "unit_cost": currency_to_cny(supplier.price, supplier.currency, rates),
            "max_total": supplier.available if supplier.available > 0 else 0.0,
        }
    factory_raw_material_destinations = {
        material.factory: {
            "initial": material.init,
            "demand": daily_demand,
            "limit": material.limit,
            "excess_fee": material.excess_fee,
        }
    }
    usable_routes = [route for route in routes if route.dst == material.factory and route.src in supplier_sources]
    transport = _solve_day_transport_milp(
        name=name,
        sources=supplier_sources,
        destinations=factory_raw_material_destinations,
        routes=usable_routes,
        products=products,
        cargo=material.material,
        carriers=carriers,
        days=days,
        ship_day_step=3 if days > 60 else 1,
    )
    purchase_cost = 0.0
    for row in transport.get("shipments", []):
        supplier = supplier_by_name.get(row["source"])
        unit_price = currency_to_cny(supplier.price, supplier.currency, rates) if supplier else 0.0
        row["material"] = material.material
        row["supplier"] = row["source"]
        row["factory"] = material.factory
        row["unit_price"] = unit_price
        row["purchase_cost"] = int(row["amount"]) * unit_price
        row["note"] = "逐日整数原料补货"
        purchase_cost += row["purchase_cost"]
    transport["purchase_cost"] = purchase_cost
    return transport


def simulate_plan(
    *,
    sales_fc: list[dict[str, Any]],
    factories: list[Factory],
    factory_materials: list[FactoryMaterial],
    production_rows: list[dict[str, Any]],
    procurement_rows: list[dict[str, Any]],
    product_transport_rows: list[dict[str, Any]],
    products: list[Product],
    material_daily_requirements: bool = False,
    days: int = PLAN_DAYS,
) -> dict[str, Any]:
    risks: list[str] = []

    factory_production_daily = production_daily_schedule(production_rows, factories, days)
    material_init = {(item.factory, item.material): float(item.init or 0.0) for item in factory_materials}
    material_need_by_factory: dict[tuple[str, str], float] = {}
    material_daily_by_key: dict[tuple[str, str], list[float]] = {}
    if material_daily_requirements:
        for item in factory_materials:
            key = (item.factory, item.material)
            daily = [float(item.daily or 0.0)] * days
            material_daily_by_key[key] = daily
            material_need_by_factory[key] = sum(daily)
    else:
        for prod in production_rows:
            factory = prod["factory"]
            amount = prod["amount"]
            for item in factory_materials:
                if item.factory != factory:
                    continue
                key = (factory, item.material)
                bom = material_bom(products, item.material)
                daily = [qty * bom for qty in factory_production_daily.get(factory, [0] * days)]
                material_daily_by_key[key] = daily
                material_need_by_factory[key] = amount * bom

    proc_by_key: dict[tuple[str, str], float] = {}
    earliest_arrival: dict[tuple[str, str], int] = {}
    material_arrivals: dict[tuple[str, str, int], float] = defaultdict(float)
    for row in procurement_rows:
        if row.get("supplier") == "缺口":
            continue
        key = (row["factory"], row["material"])
        proc_by_key[key] = proc_by_key.get(key, 0.0) + row["amount"]
        earliest_arrival[key] = min(earliest_arrival.get(key, 999), int(row.get("lead") or 0))
        arrival_day = int(row.get("arrival_day") or (int(row.get("ship_day") or 1) + int(row.get("lead") or 0)))
        if 1 <= arrival_day <= days:
            material_arrivals[(key[0], key[1], arrival_day)] += float(row.get("amount") or 0.0)

    for key, need in material_need_by_factory.items():
        init = material_init.get(key, 0.0)
        got = proc_by_key.get(key, 0.0)
        lead = earliest_arrival.get(key, 0)
        daily_need = need / days if need else 0.0
        if init + got + 1e-6 < need:
            risks.append(f"{key[0]}-{key[1]} 原料总量不足，缺 {need - init - got:,.0f}")
        if lead > 0 and init < daily_need * lead:
            risks.append(f"{key[0]}-{key[1]} 首批到货 {lead} 天，期初库存可能撑不到到货")
        inventory = float(init)
        min_inventory = inventory
        first_negative_day = None
        for day, demand in enumerate(material_daily_by_key.get(key, [0.0] * days), start=1):
            inventory += material_arrivals.get((key[0], key[1], day), 0.0)
            inventory -= float(demand or 0.0)
            min_inventory = min(min_inventory, inventory)
            if inventory < -1e-6 and first_negative_day is None:
                first_negative_day = day
        if first_negative_day is not None:
            risks.append(f"{key[0]}-{key[1]} 第{first_negative_day}天工厂原料库存为负，最低 {min_inventory:,.0f}")

    factory_init = {factory.name: factory.init for factory in factories}
    for prod in production_rows:
        cap = prod.get("capacity", 0.0)
        if prod["amount"] > cap + 1e-6:
            risks.append(f"{prod['factory']} 生产量超过 {days} 天产能")

    factory_by_name = {factory.name: factory for factory in factories}
    shipped_by_factory_day: dict[tuple[str, int], float] = {}
    for row in product_transport_rows:
        factory_name = sv(row.get("factory") or row.get("source"))
        if not factory_name:
            continue
        day = max(1, min(days, int(row.get("ship_day") or 1)))
        shipped_by_factory_day[(factory_name, day)] = shipped_by_factory_day.get((factory_name, day), 0.0) + float(row.get("amount") or 0.0)

    for factory_name, factory in factory_by_name.items():
        inventory = float(factory.init or 0.0)
        min_inventory = inventory
        first_negative_day = None
        for day in range(1, days + 1):
            inventory += float(factory_production_daily.get(factory_name, [0] * days)[day - 1])
            inventory -= shipped_by_factory_day.get((factory_name, day), 0.0)
            min_inventory = min(min_inventory, inventory)
            if inventory < -1e-6 and first_negative_day is None:
                first_negative_day = day
        if first_negative_day is not None:
            risks.append(f"{factory_name} 第{first_negative_day}天成品库存为负，最低 {min_inventory:,.0f}")

    forecast_by_node = {row["node"]: row for row in sales_fc}
    ship_by_node: dict[str, float] = {}
    arrival_by_node: dict[str, int] = {}
    for row in product_transport_rows:
        node = row["destination"]
        ship_by_node[node] = ship_by_node.get(node, 0.0) + row["amount"]
        arrival_by_node[node] = min(arrival_by_node.get(node, 999), int(row.get("lead") or 0))

    for node, forecast in forecast_by_node.items():
        required = forecast["forecast"]
        init = forecast["init"]
        shipped = ship_by_node.get(node, 0.0)
        lead = arrival_by_node.get(node, 0)
        if init + shipped + 1e-6 < required:
            risks.append(f"{node} 成品总量不足，市场缺口 {required - init - shipped:,.0f}")
        if lead > 0 and init < required / days * lead:
            risks.append(f"{node} 成品到货 {lead} 天，期初库存可能撑不到到货")
        if init + shipped > forecast["limit"] + required:
            risks.append(f"{node} 补货偏多，可能推高库存")

    total_production = sum(sum(daily) for daily in factory_production_daily.values())
    available_product = sum(factory_init.values()) + total_production
    total_market_need = sum(row["forecast"] for row in sales_fc)
    total_ship = sum(row["amount"] for row in product_transport_rows)
    if total_ship > available_product + 1e-6:
        risks.append("成品发运量超过工厂期初库存+生产量")
    if sales_fc and total_ship + sum(row["init"] for row in sales_fc) < total_market_need:
        risks.append("销售网点总供给不足")
    if sales_fc:
        market_replay = daily_market_replay(sales_fc, product_transport_rows, days)
        risks.extend(market_replay["risks"])
    else:
        market_replay = None

    return {
        "ok": not risks,
        "risks": risks,
        "market_replay": market_replay,
        "summary": "校验通过：库存、产能、运输总量未发现硬缺口" if not risks else "存在风险：需关注断料/断货或超量",
    }


def solve_procurement(sections: dict[str, list[tuple[Any, ...]]], xls_path: Path) -> dict[str, Any]:
    plan_days = plan_days_for_case(xls_path, "采购")
    products = parse_products(sections)
    factories = parse_factories(sections)
    factory_materials = parse_factory_materials(sections)
    suppliers = parse_suppliers(sections)
    routes = parse_routes(sections)
    rates = parse_rates(sections)
    carriers = parse_carriers_safe(sections, rates)

    factory_by_name = {factory.name: factory for factory in factories}
    plan_units_by_factory = {
        factory.name: max(0.0, float(factory.daily or 0.0) * plan_days)
        for factory in factories
        if factory.daily > 0
    }
    for factory, units in material_plan_product_units(factory_materials, products, plan_days).items():
        plan_units_by_factory.setdefault(factory, units)

    production_rows: list[dict[str, Any]] = []
    for factory in factories:
        production_rows.append(
            {
                "factory": factory.name,
                "product": factory.product,
                "amount": ceil_int(plan_units_by_factory.get(factory.name, factory.daily * plan_days)),
                "capacity": factory.daily * plan_days,
                "init": factory.init,
            }
        )
    for factory, units in plan_units_by_factory.items():
        if factory in factory_by_name:
            continue
        production_rows.append(
            {
                "factory": factory,
                "product": next((product.name for product in products if product.kind == "产品"), ""),
                "amount": ceil_int(units),
                "capacity": ceil_int(units),
                "init": 0.0,
            }
        )

    material_transport: list[dict[str, Any]] = []
    material_failures: list[str] = []
    material_status: dict[str, str] = {}
    product_loss_by_factory: dict[str, float] = defaultdict(float)
    total_raw_demand = 0.0
    total_raw_shortage = 0.0

    for idx, item in enumerate(factory_materials):
        daily_demand = [float(item.daily or 0.0)] * plan_days
        total_raw_demand += sum(daily_demand)
        transport = solve_material_procurement_daily(
            material=item,
            daily_demand=daily_demand,
            products=products,
            suppliers=suppliers,
            routes=routes,
            rates=rates,
            carriers=carriers,
            days=plan_days,
            name=f"procurement_material_{idx}",
        )
        material_status[f"{item.factory}-{item.material}"] = str(transport.get("status", "Unknown"))
        material_failures.extend(transport.get("failures", []))
        material_transport.extend(transport.get("shipments", []))
        shortage = float(transport.get("shortage", 0.0) or 0.0)
        total_raw_shortage += shortage
        bom = material_bom(products, item.material)
        if bom > 0 and shortage > 0:
            product_loss_by_factory[item.factory] = max(product_loss_by_factory[item.factory], shortage / bom)

    procurement = procurement_summary_rows(material_transport, "逐日整数采购汇总")
    total_purchase = sum(float(row["purchase_cost"]) for row in procurement)
    total_freight = sum(float(row["freight_cost"]) for row in procurement)
    planned_product_units = sum(plan_units_by_factory.values())
    actual_product_units = sum(max(0.0, units - product_loss_by_factory.get(factory, 0.0)) for factory, units in plan_units_by_factory.items())
    if planned_product_units <= 0 and total_raw_demand > 0:
        planned_product_units = total_raw_demand
        actual_product_units = max(0.0, total_raw_demand - total_raw_shortage)
    production_satisfaction = min(1.0, actual_product_units / max(planned_product_units, 0.001)) if planned_product_units else 1.0
    initial_cost = initial_material_inventory_cost(factory_materials, suppliers, rates)
    unit_procurement = (total_purchase + total_freight + initial_cost) / max(actual_product_units, 0.001)
    targets = score_targets(xls_path, "采购")
    target_cost = targets.get("unit_procurement_cost")
    score_rows: list[dict[str, Any]] = []
    if target_cost:
        procurement_score = cost_score(unit_procurement, target_cost, 60)
        satisfaction_points = satisfaction_score(production_satisfaction, 40)
        score_rows = [
            {"item": "单位采购成本", "target": target_cost, "actual": unit_procurement, "points": procurement_score, "max": 60},
            {"item": "生产满足率", "target": 1.0, "actual": production_satisfaction, "points": satisfaction_points, "max": 40},
        ]
        score: float | None = sum(row["points"] for row in score_rows)
        score_note = "按已配置采购评分指标计算"
    else:
        score = None
        score_note = "未配置该场次正式单位采购成本指标：只输出优化计划、单位采购成本和生产满足率，不冒充平台真实分。"
    simulation = simulate_plan(
        sales_fc=[],
        factories=factories,
        factory_materials=factory_materials,
        production_rows=production_rows,
        procurement_rows=material_transport,
        product_transport_rows=[],
        products=products,
        material_daily_requirements=True,
        days=plan_days,
    )
    if material_failures:
        simulation["risks"] = list(dict.fromkeys(list(simulation.get("risks", [])) + material_failures))
        simulation["ok"] = False
        simulation["summary"] = "逐日采购校验存在断料风险"
    return {
        "qtype": "采购",
        "products": products,
        "sales_fc": [],
        "production": production_rows,
        "procurement": procurement,
        "material_transport": material_transport,
        "product_transport": [],
        "transport_sections": _compact_transport_rows(material_transport),
        "score_rows": score_rows,
        "score": score,
        "score_note": score_note,
        "unit_procurement": unit_procurement,
        "unit_logistics": total_freight / max(actual_product_units, 0.001),
        "production_satisfaction": production_satisfaction,
        "market_satisfaction": 1.0,
        "simulation": simulation,
        "plan_days": plan_days,
        "solver_status": material_status,
        "assumptions": [
            f"{plan_days} 天逐日原料消耗建模",
            "供应商总可供量、日产能、路线提前期和最低运费均进入整数模型",
            "路线/趟次惩罚为 0，最终分数只按平台评分公式或正式配置计算",
        ],
    }


def solve_sales_or_production(
    sections: dict[str, list[tuple[Any, ...]]],
    xls_path: Path,
    qtype: str,
) -> dict[str, Any]:
    plan_days = plan_days_for_case(xls_path, qtype)
    products = parse_products(sections)
    factories = parse_factories(sections)
    factory_materials = parse_factory_materials(sections)
    suppliers = parse_suppliers(sections)
    routes = parse_routes(sections)
    rates = parse_rates(sections)
    carriers = parse_carriers_safe(sections, rates)
    sales = parse_sales(sections, plan_days)
    use_bias = forecast_bias_for_case(xls_path, qtype)
    forecasts = [forecast_node(node, plan_days, use_bias=use_bias) for node in sales]
    stage_timings: list[dict[str, Any]] = []

    def mark_stage(name: str, started_at: float) -> None:
        stage_timings.append({"stage": name, "seconds": round(time.time() - started_at, 3)})

    product_name = products[0].name if products else (factories[0].product if factories else "")
    assignment: dict[str, Factory] = {}
    demand_by_factory: dict[str, float] = {}
    explicit_production_rows: list[dict[str, Any]] | None = None
    if qtype == "生产":
        production_amounts = {
            factory.name: int(math.floor(float(factory.daily or 0.0) * plan_days))
            for factory in factories
        }
        explicit_production_rows = production_rows_from_amounts(factories, product_name, production_amounts, plan_days)

    stage_started = time.time()
    for forecast in forecasts:
        if factories:
            ranked = []
            for candidate in factories:
                amount = max(forecast["forecast"], 1.0)
                ratio = charge_ratio(products, product_name)
                candidate_route = pick_best_route(routes, candidate.name, forecast["node"], amount, ratio, urgent=False, carriers=carriers)
                route_cost_score = route_score(candidate_route, amount, ratio, urgent=False, carriers=carriers) if candidate_route else 999999
                ranked.append((route_cost_score, candidate, candidate_route))
            _, factory, _route = min(ranked, key=lambda item: item[0])
            assignment[forecast["node"]] = factory

    product_transport: list[dict[str, Any]] = []
    product_solver_status: dict[str, Any] = {"method": "FastIntegerPeriodicTransport"}
    initial_capacity_schedule = factory_output_schedule(factories, plan_days, explicit_production_rows)
    finished_goods_capacity_remaining = {
        factory.name: max(0.0, float(factory.init or 0.0) + sum(initial_capacity_schedule.get(factory.name, [0] * plan_days)))
        for factory in factories
    }
    transport_allocations: list[tuple[dict[str, Any], Factory, int, float, float]] = []
    for forecast in sorted(forecasts, key=lambda row: float(row.get("forecast") or 0.0), reverse=True):
        if not factories:
            continue
        net_need = ceil_int(max(0.0, float(forecast.get("forecast") or 0.0) - float(forecast.get("init") or 0.0)))
        if net_need <= 0:
            continue
        ranked: list[tuple[float, Factory, Route | None]] = []
        ratio = charge_ratio(products, product_name)
        for candidate in factories:
            route = pick_best_route(routes, candidate.name, forecast["node"], max(net_need, 1.0), ratio, urgent=False, carriers=carriers)
            if route:
                ranked.append((route_score(route, max(net_need, 1.0), ratio, urgent=False, carriers=carriers), candidate, route))
        if not ranked:
            continue
        ranked.sort(key=lambda item: item[0])
        assignment.setdefault(forecast["node"], ranked[0][1])
        remaining = net_need
        first = True
        for _score, factory, _route in ranked:
            if remaining <= 0:
                break
            available = max(0, int(math.floor(finished_goods_capacity_remaining.get(factory.name, 0.0))))
            if available <= 0:
                continue
            amount = min(remaining, available)
            init_share = float(forecast.get("init") or 0.0) if first else 0.0
            demand_share = (amount + init_share) / max(float(forecast.get("forecast") or 0.0), 1.0)
            transport_allocations.append((forecast, factory, int(amount), init_share, demand_share))
            demand_by_factory[factory.name] = demand_by_factory.get(factory.name, 0.0) + amount
            finished_goods_capacity_remaining[factory.name] = finished_goods_capacity_remaining.get(factory.name, 0.0) - amount
            remaining -= amount
            first = False
        if remaining > 0:
            factory = ranked[0][1]
            demand_share = remaining / max(float(forecast.get("forecast") or 0.0), 1.0)
            transport_allocations.append((forecast, factory, int(remaining), 0.0, demand_share))
            demand_by_factory[factory.name] = demand_by_factory.get(factory.name, 0.0) + remaining
    mark_stage("initial_assignment", stage_started)

    stage_started = time.time()
    fast_transport_raw = build_fast_product_transport(
        allocations=transport_allocations,
        routes=routes,
        products=products,
        cargo=product_name,
        carriers=carriers,
        days=plan_days,
    )

    fast_transport = effective_shipments(repair_factory_ship_days(fast_transport_raw, factories, plan_days, explicit_production_rows), plan_days)
    mark_stage("fast_transport", stage_started)

    stage_started = time.time()
    safe_transport = build_safe_product_transport(
        forecasts=forecasts,
        factories=factories,
        routes=routes,
        products=products,
        cargo=product_name,
        carriers=carriers,
        days=plan_days,
        production_rows=explicit_production_rows,
    )
    mark_stage("safe_transport", stage_started)

    stage_started = time.time()
    service_first_transport = build_service_first_product_transport(
        forecasts=forecasts,
        factories=factories,
        routes=routes,
        products=products,
        cargo=product_name,
        carriers=carriers,
        days=plan_days,
        production_rows=explicit_production_rows,
        cover_days=4 if plan_days <= 30 else 6,
    )
    mark_stage("service_first_transport", stage_started)

    stage_started = time.time()
    polished_fast_transport = polish_product_transport(
        forecasts=forecasts,
        shipments=fast_transport,
        assignment=assignment,
        routes=routes,
        products=products,
        cargo=product_name,
        carriers=carriers,
        factories=factories,
        xls_path=xls_path,
        qtype=qtype,
        days=plan_days,
        fill_rounds=2,
        production_rows=explicit_production_rows,
    )
    mark_stage("polish_fast_transport", stage_started)

    stage_started = time.time()
    full_fill_transport = fill_total_market_gap_fast(
        forecasts=forecasts,
        shipments=fill_early_shortages_from_any_factory(
            forecasts=forecasts,
            shipments=polished_fast_transport,
            routes=routes,
            products=products,
            cargo=product_name,
            carriers=carriers,
            factories=factories,
            days=plan_days,
            production_rows=explicit_production_rows,
        ),
        assignment=assignment,
        routes=routes,
        products=products,
        cargo=product_name,
        carriers=carriers,
        factories=factories,
        days=plan_days,
        production_rows=explicit_production_rows,
    )
    full_fill_transport = polish_product_transport(
        forecasts=forecasts,
        shipments=full_fill_transport,
        assignment=assignment,
        routes=routes,
        products=products,
        cargo=product_name,
        carriers=carriers,
        factories=factories,
        xls_path=xls_path,
        qtype=qtype,
        days=plan_days,
        fill_rounds=4,
        production_rows=explicit_production_rows,
    )
    mark_stage("full_fill_transport", stage_started)

    stage_started = time.time()
    budget_fill_transport = fill_market_shortages_score_aware(
        forecasts=forecasts,
        shipments=fast_transport,
        routes=routes,
        products=products,
        cargo=product_name,
        carriers=carriers,
        factories=factories,
        xls_path=xls_path,
        qtype=qtype,
        days=plan_days,
    )
    mark_stage("budget_fill_transport", stage_started)
    inventory_window_transport: list[dict[str, Any]] = []
    inventory_window_status: dict[str, Any] = {
        "method": "InventoryWindowExtremeSalesDP",
        "status": "Skipped",
    }
    if is_extreme_mode() and qtype == "销售" and plan_days > 45:
        stage_started = time.time()
        inventory_window_transport, inventory_window_status = build_inventory_window_extreme_sales_transport(
            forecasts=forecasts,
            factories=factories,
            routes=routes,
            products=products,
            cargo=product_name,
            carriers=carriers,
            days=plan_days,
            production_rows=explicit_production_rows,
        )
        mark_stage("inventory_window_extreme", stage_started)

    def make_transport_candidate(
        name: str,
        rows: list[dict[str, Any]],
        status: dict[str, Any],
    ) -> tuple[str, list[dict[str, Any]], dict[str, Any], dict[str, float]]:
        if status.get("skip_sanitize"):
            rows = effective_shipments([dict(row) for row in rows], plan_days)
        else:
            rows = sanitize_product_transport(
                forecasts=forecasts,
                shipments=rows,
                factories=factories,
                routes=routes,
                products=products,
                cargo=product_name,
                carriers=carriers,
                days=plan_days,
                production_rows=explicit_production_rows,
            )
        hard_risks = product_transport_hard_risks(
            forecasts=forecasts,
            shipments=rows,
            factories=factories,
            days=plan_days,
            production_rows=explicit_production_rows,
        )
        metrics = score_product_transport_candidate(
            forecasts=forecasts,
            shipments=rows,
            xls_path=xls_path,
            qtype=qtype,
            days=plan_days,
        )
        metrics["hard_risk_count"] = float(len(hard_risks))
        return name, rows, {**status, "hard_risks": hard_risks}, metrics

    candidates = [
        make_transport_candidate(
            "快速周期整数运输",
            fast_transport,
            {"method": "FastIntegerPeriodicTransport"},
        )
    ]
    candidates.append(
        make_transport_candidate(
            "保守不补货硬约束基线",
            [],
            {"method": "ZeroShipmentHardConstraintBaseline"},
        )
    )
    if safe_transport and safe_transport != fast_transport:
        candidates.append(
            make_transport_candidate(
                "安全库存滚动补货",
                safe_transport,
                {"method": "SafeRollingReplenishment"},
            )
        )
    if service_first_transport and service_first_transport not in (fast_transport, safe_transport):
        candidates.append(
            make_transport_candidate(
                "满足率优先滚动补货",
                service_first_transport,
                {"method": "ServiceFirstRollingReplenishment"},
            )
        )
    if polished_fast_transport and polished_fast_transport != fast_transport:
        candidates.append(
            make_transport_candidate(
                "快速补缺口整数运输",
                polished_fast_transport,
                {"method": "FastIntegerPolishedTransport"},
            )
        )
    if full_fill_transport and full_fill_transport not in (fast_transport, polished_fast_transport):
        candidates.append(
            make_transport_candidate(
                "满足率快速补齐运输",
                full_fill_transport,
                {"method": "FastFullSatisfactionFillTransport"},
            )
        )
    if budget_fill_transport and budget_fill_transport not in (fast_transport, polished_fast_transport):
        candidates.append(
            make_transport_candidate(
                "评分预算补货整数运输",
                budget_fill_transport,
                {"method": "ScoreAwareBudgetFillTransport"},
            )
        )
    if inventory_window_transport:
        candidates.append(
            make_transport_candidate(
                "极限版库存窗口最低成本补货",
                inventory_window_transport,
                {**inventory_window_status, "skip_sanitize": True},
            )
        )

    if is_extreme_mode() and qtype == "销售" and plan_days > 45:
        extreme_post_polish_status: dict[str, Any] = {}
        full_valid_candidates = [
            item for item in candidates
            if not item[2].get("hard_risks")
            and is_full_satisfaction(item[3].get("market_satisfaction", 0.0), item[3].get("shortage", 0.0))
        ]
        if full_valid_candidates:
            stage_started = time.time()
            post_polish_timings: list[dict[str, Any]] = []

            def mark_post_stage(name: str, started_at: float, detail: dict[str, Any] | None = None) -> None:
                row: dict[str, Any] = {"stage": name, "seconds": round(time.time() - started_at, 3)}
                if isinstance(detail, dict):
                    for key in ("status", "delta", "improvements", "checked", "feasible_improving", "rounds", "cycles"):
                        if key in detail:
                            row[key] = detail[key]
                post_polish_timings.append(row)

            base_name, base_rows, _base_status, base_metrics = max(
                full_valid_candidates,
                key=lambda item: product_transport_priority(item[3], len(item[1])),
            )
            post_option_cache: dict[tuple[str, int, int], list[dict[str, Any]]] = {}
            post_stage_started = time.time()
            polished_rows, polished_status = polish_inventory_window_transport_by_destination(
                forecasts=forecasts,
                shipments=base_rows,
                factories=factories,
                routes=routes,
                products=products,
                cargo=product_name,
                carriers=carriers,
                days=plan_days,
                production_rows=explicit_production_rows,
                max_rounds=5,
                option_cache=post_option_cache,
            )
            mark_post_stage("destination_repolish", post_stage_started, polished_status)
            post_stage_started = time.time()
            cycled_rows, cycle_status = polish_extreme_transport_cycles(
                forecasts=forecasts,
                shipments=polished_rows,
                factories=factories,
                routes=routes,
                products=products,
                cargo=product_name,
                carriers=carriers,
                days=plan_days,
                production_rows=explicit_production_rows,
                max_cycles=3 if case_keyword(xls_path) == "热水器" else 4,
                option_cache=post_option_cache,
            )
            if cycled_rows:
                polished_rows = cycled_rows
            mark_post_stage("cycle_polish", post_stage_started, cycle_status)
            post_stage_started = time.time()
            fixed_reroute_rows, fixed_reroute_status = optimize_fixed_arrival_reroute(
                forecasts=forecasts,
                shipments=polished_rows,
                factories=factories,
                routes=routes,
                products=products,
                cargo=product_name,
                carriers=carriers,
                days=plan_days,
                production_rows=explicit_production_rows,
                time_limit_sec=env_int("SUPPLY_CHAIN_FIXED_REROUTE_TIMELIMIT", 30),
            )
            if fixed_reroute_rows and float(fixed_reroute_status.get("delta") or 0.0) > 1e-6:
                polished_rows = fixed_reroute_rows
            mark_post_stage("fixed_arrival_reroute", post_stage_started, fixed_reroute_status)
            post_stage_started = time.time()
            block_rows, block_status = polish_destination_blocks(
                forecasts=forecasts,
                shipments=polished_rows,
                factories=factories,
                routes=routes,
                products=products,
                cargo=product_name,
                carriers=carriers,
                days=plan_days,
                production_rows=explicit_production_rows,
                max_block_size=3,
                max_rounds=1 if case_keyword(xls_path) == "热水器" else 3,
                option_cache=post_option_cache,
            )
            if block_rows:
                polished_rows = block_rows
            mark_post_stage("destination_blocks", post_stage_started, block_status)
            post_stage_started = time.time()
            if case_keyword(xls_path) == "热水器" and not heatwater_pattern_audit_enabled():
                pattern_rows = []
                pattern_status = {
                    "status": "SkippedHeatwaterNoRecentGain",
                    "delta": 0.0,
                    "improvements": 0,
                    "checked": 0,
                    "reason": "热水器近轮模式组合审查耗时高且未降本，保留有收益的块重排后跳过",
                }
            else:
                pattern_rows, pattern_status = polish_destination_pattern_combinations(
                    forecasts=forecasts,
                    shipments=polished_rows,
                    factories=factories,
                    routes=routes,
                    products=products,
                    cargo=product_name,
                    carriers=carriers,
                    days=plan_days,
                    production_rows=explicit_production_rows,
                    max_block_size=3,
                    max_patterns_per_destination=env_int("SUPPLY_CHAIN_HEATWATER_PATTERN_MAX_PATTERNS", 8) if case_keyword(xls_path) == "热水器" else 8,
                    max_combinations=env_int("SUPPLY_CHAIN_HEATWATER_PATTERN_MAX_COMBINATIONS", 5000) if case_keyword(xls_path) == "热水器" else 5000,
                )
                if pattern_rows:
                    polished_rows = pattern_rows
            mark_post_stage("pattern_selection", post_stage_started, pattern_status)
            post_stage_started = time.time()
            final_cycle_rows, final_cycle_status = polish_extreme_transport_cycles(
                forecasts=forecasts,
                shipments=polished_rows,
                factories=factories,
                routes=routes,
                products=products,
                cargo=product_name,
                carriers=carriers,
                days=plan_days,
                production_rows=explicit_production_rows,
                max_cycles=3,
                option_cache=post_option_cache,
            )
            if final_cycle_rows:
                polished_rows = final_cycle_rows
            mark_post_stage("final_cycle_polish", post_stage_started, final_cycle_status)
            extreme_post_polish_status = {
                "method": "ExtremePostPolishAudit",
                "status": "Computed",
                "base_candidate": base_name,
                "destination_repolish": polished_status,
                "post_polish_cycles": cycle_status,
                "fixed_arrival_reroute": fixed_reroute_status,
                "destination_block_polish": block_status,
                "destination_pattern_polish": pattern_status,
                "final_polish_cycles": final_cycle_status,
                "post_option_cache_entries": len(post_option_cache),
                "post_polish_timings": post_polish_timings,
            }
            polished_freight = sum(float(row.get("freight_cost") or 0.0) for row in polished_rows)
            if polished_rows and polished_freight < float(base_metrics.get("freight", 0.0) or 0.0) - 1e-6:
                candidates.append(
                    make_transport_candidate(
                        f"{base_name}+极限库存窗口复优化",
                        polished_rows,
                        {
                            **polished_status,
                            **extreme_post_polish_status,
                            "method": "InventoryWindowPolishFromFeasibleCandidate",
                            "base_candidate": base_name,
                            "skip_sanitize": True,
                        },
                    )
                )
            mark_stage("extreme_post_polish", stage_started)
    else:
        extreme_post_polish_status = {}

    warm_window_status: dict[str, Any] = {
        "method": "WarmWindowGlobalIntegerTransport",
        "status": "Skipped",
    }
    full_global_audit_status: dict[str, Any] = {
        "method": "FullGlobalIntegerAudit",
        "status": "Skipped",
    }
    if is_extreme_mode() and qtype == "销售" and plan_days > 45:
        full_valid_candidates = [
            item for item in candidates
            if not item[2].get("hard_risks")
            and is_full_satisfaction(item[3].get("market_satisfaction", 0.0), item[3].get("shortage", 0.0))
        ]
        best_full_metric = max(
            (item[3] for item in full_valid_candidates),
            key=lambda metrics: product_transport_priority(metrics, 0),
            default=None,
        )
        skip_warm_window = (
            case_keyword(xls_path) == "热水器"
            and not heatwater_global_audit_enabled()
            and isinstance(best_full_metric, dict)
            and is_full_satisfaction(best_full_metric.get("market_satisfaction", 0.0), best_full_metric.get("shortage", 0.0))
        )
        if skip_warm_window:
            warm_window_status = {
                **warm_window_status,
                "status": "SkippedHeatwaterFullSatisfactionCandidate",
                "reason": "热水器极限版已有满满足候选；跳过未被采纳且易超时的暖窗口全局审查",
            }
        elif full_valid_candidates:
            stage_started = time.time()
            warm_name, warm_rows, _warm_status, _warm_metrics = max(
                full_valid_candidates,
                key=lambda item: product_transport_priority(item[3], len(item[1])),
            )
            warm_base_freight = float(_warm_metrics.get("freight", 0.0) or 0.0)
            strict_bound = warm_base_freight - 0.01 if heatwater_counterexample_audit_enabled() and warm_base_freight > 0 else None
            warm_window_transport, warm_window_status = build_warm_window_global_product_transport(
                forecasts=forecasts,
                factories=factories,
                routes=routes,
                products=products,
                cargo=product_name,
                carriers=carriers,
                days=plan_days,
                name=f"{case_keyword(xls_path)}_{qtype}_warm_window_global_transport",
                warm_start_shipments=warm_rows,
                production_rows=explicit_production_rows,
                window=env_int("SUPPLY_CHAIN_HEATWATER_GLOBAL_WINDOW", 2) if case_keyword(xls_path) == "热水器" else 2,
                time_limit_sec=env_int("SUPPLY_CHAIN_HEATWATER_GLOBAL_TIMELIMIT", 30) if case_keyword(xls_path) == "热水器" else 30,
                strict_freight_upper_bound=strict_bound,
            )
            warm_window_freight = sum(float(row.get("freight_cost") or 0.0) for row in warm_window_transport)
            warm_window_optimal = sv(warm_window_status.get("status")) == "Optimal"
            warm_window_status = {
                **warm_window_status,
                "warm_window_freight": warm_window_freight if warm_window_transport else None,
                "warm_base_freight": warm_base_freight,
            }
            if warm_window_optimal and warm_window_transport and warm_window_freight < warm_base_freight - 1e-6:
                candidates.append(
                    make_transport_candidate(
                        f"{warm_name}+暖启动窗口全局重配",
                        warm_window_transport,
                        {**warm_window_status, "skip_sanitize": True},
                    )
                )
            elif warm_window_transport:
                warm_window_status = {
                    **warm_window_status,
                    "status": (
                        f"{warm_window_status.get('status', 'Solved')}NotSelected"
                        if not warm_window_optimal
                        else f"{warm_window_status.get('status', 'Solved')}NoCostImprovement"
                    ),
                }
            mark_stage("warm_window_global", stage_started)
        else:
            warm_window_status["status"] = "SkippedNoFullSatisfactionWarmStart"

        if case_keyword(xls_path) == "热水器" and heatwater_full_global_audit_enabled() and full_valid_candidates:
            stage_started = time.time()
            audit_name, audit_rows, _audit_status, audit_metrics = max(
                full_valid_candidates,
                key=lambda item: product_transport_priority(item[3], len(item[1])),
            )
            full_global_audit_transport, full_global_audit_status = build_full_global_audit_product_transport(
                forecasts=forecasts,
                factories=factories,
                routes=routes,
                products=products,
                cargo=product_name,
                carriers=carriers,
                days=plan_days,
                name=f"{case_keyword(xls_path)}_{qtype}_full_global_audit",
                warm_start_shipments=audit_rows,
                production_rows=explicit_production_rows,
                time_limit_sec=env_int("SUPPLY_CHAIN_HEATWATER_FULL_GLOBAL_TIMELIMIT", 300),
            )
            audit_freight = sum(float(row.get("freight_cost") or 0.0) for row in full_global_audit_transport)
            audit_base_freight = float(audit_metrics.get("freight", 0.0) or 0.0)
            if full_global_audit_transport and audit_freight < audit_base_freight - 1e-6 and sv(full_global_audit_status.get("status")) == "Optimal":
                candidates.append(
                    make_transport_candidate(
                        f"{audit_name}+全路线全日期整数审查",
                        full_global_audit_transport,
                        {**full_global_audit_status, "skip_sanitize": True},
                    )
                )
            elif full_global_audit_transport:
                full_global_audit_status = {
                    **full_global_audit_status,
                    "status": (
                        f"{full_global_audit_status.get('status', 'Solved')}NotSelected"
                        if sv(full_global_audit_status.get("status")) != "Optimal"
                        else f"{full_global_audit_status.get('status', 'Solved')}NoCostImprovement"
                    ),
                    "full_global_freight": audit_freight,
                    "full_global_base_freight": audit_base_freight,
                }
            mark_stage("full_global_audit", stage_started)

    global_transport: list[dict[str, Any]] = []
    global_status: dict[str, Any] = {
        "method": "GlobalIntegerTransport",
        "status": "WillRun" if plan_days <= 45 else "SkippedLongHorizon",
    }
    score_aware_transport: list[dict[str, Any]] = []
    score_aware_status: dict[str, Any] = {
        "method": "ScoreAwareServiceLevelMILP",
        "status": "WillRun",
    }
    budget_score_transport: list[dict[str, Any]] = []
    budget_score_status: dict[str, Any] = {
        "method": "ScoreBudgetMILP",
        "status": "WillRun",
    }

    def has_full_valid_satisfaction() -> bool:
        valid = [item for item in candidates if not item[2].get("hard_risks")]
        return any(is_full_satisfaction(item[3].get("market_satisfaction", 0.0), item[3].get("shortage", 0.0)) for item in valid)

    run_global_transport = plan_days <= 45
    run_budget_transport = plan_days <= 45 or (qtype == "销售" and plan_days <= 90)
    if is_extreme_mode() and qtype == "销售" and plan_days > 45 and has_full_valid_satisfaction():
        run_budget_transport = False

    if run_global_transport:
        global_transport, global_status = build_global_product_transport(
            forecasts=forecasts,
            factories=factories,
            routes=routes,
            products=products,
            cargo=product_name,
            carriers=carriers,
            days=plan_days,
            name=f"{case_keyword(xls_path)}_{qtype}_global_product_transport",
            production_rows=explicit_production_rows,
        )
        if global_transport:
            candidates.append(
                make_transport_candidate(
                    "全局逐日整数运输",
                    global_transport,
                    {**global_status, "skip_sanitize": True},
                )
            )
    else:
        global_status["status"] = "SkippedLongHorizon"

    if run_budget_transport:
        warm_start_rows: list[dict[str, Any]] | None = None
        if is_extreme_mode():
            warm_candidates = [
                item for item in candidates
                if not item[2].get("hard_risks")
                and is_full_satisfaction(item[3].get("market_satisfaction", 0.0), item[3].get("shortage", 0.0))
            ]
            if warm_candidates:
                warm_start_rows = max(warm_candidates, key=lambda item: product_transport_priority(item[3], len(item[1])))[1]
        budget_score_transport, budget_score_status = build_budget_score_product_transport(
            forecasts=forecasts,
            factories=factories,
            routes=routes,
            products=products,
            cargo=product_name,
            carriers=carriers,
            xls_path=xls_path,
            qtype=qtype,
            days=plan_days,
            name=f"{case_keyword(xls_path)}_{qtype}_budget_score_product_transport",
            production_rows=explicit_production_rows,
            warm_start_shipments=warm_start_rows,
        )
        if budget_score_transport:
            candidates.append(
                make_transport_candidate(
                    "平台分数成本预算运输",
                    budget_score_transport,
                    {**budget_score_status, "skip_sanitize": True},
                )
            )
    else:
        budget_score_status["status"] = "SkippedAfterInventoryWindowCandidate" if is_extreme_mode() and qtype == "销售" and plan_days > 45 else "SkippedLongHorizon"

    if has_full_valid_satisfaction():
        score_aware_status["status"] = "SkippedFullSatisfactionFastMode" if not is_extreme_mode() else "SkippedFullSatisfactionAfterCostCompare"
    elif is_extreme_mode() and plan_days <= 45:
        if qtype == "生产" and not has_full_valid_satisfaction():
            score_aware_transport, score_aware_status = build_score_aware_product_transport(
                forecasts=forecasts,
                factories=factories,
                routes=routes,
                products=products,
                cargo=product_name,
                carriers=carriers,
                xls_path=xls_path,
                qtype=qtype,
                days=plan_days,
                name=f"{case_keyword(xls_path)}_{qtype}_score_aware_product_transport",
                production_rows=explicit_production_rows,
            )
            if score_aware_transport:
                candidates.append(
                    make_transport_candidate(
                        "平台分数服务率枚举运输",
                        score_aware_transport,
                        score_aware_status,
                    )
                )
        else:
            score_aware_status["status"] = "SkippedAfterBudgetOrSales"
    else:
        score_aware_status["status"] = "SkippedHighScoreFastMode" if not is_extreme_mode() else "SkippedLongHorizon"

    if is_extreme_mode() and qtype == "生产" and case_keyword(xls_path) == "毛衣":
        for candidate_name, candidate_rows, candidate_status, _candidate_metrics in list(candidates):
            advanced_rows = advance_shortage_shipments_by_faster_route(
                forecasts=forecasts,
                shipments=candidate_rows,
                factories=factories,
                routes=routes,
                products=products,
                cargo=product_name,
                carriers=carriers,
                xls_path=xls_path,
                qtype=qtype,
                days=plan_days,
                production_rows=explicit_production_rows,
            )
            if advanced_rows and advanced_rows != candidate_rows:
                candidates.append(
                    make_transport_candidate(
                        f"{candidate_name}+早期断货快线改道",
                        advanced_rows,
                        {**candidate_status, "method": f"{candidate_status.get('method', candidate_name)}+FastShortageRouteAdvance"},
                    )
                )

    selectable_candidates = [item for item in candidates if not item[2].get("hard_risks")]
    full_satisfaction_candidates = [
        item for item in selectable_candidates
        if is_full_satisfaction(item[3].get("market_satisfaction", 0.0), item[3].get("shortage", 0.0))
    ]
    if full_satisfaction_candidates:
        non_time_limited_full = [
            item for item in full_satisfaction_candidates
            if not is_time_limited_status(item[2])
        ]
        selectable_candidates = non_time_limited_full or full_satisfaction_candidates
    if selectable_candidates:
        chosen_name, product_transport, product_solver_status, chosen_metrics = max(
            selectable_candidates,
            key=lambda item: product_transport_priority(item[3], len(item[1])),
        )
    else:
        product_transport = []
        chosen_name = "无硬风险可行成品运输候选"
        product_solver_status = {
            "method": "NoHardFeasibleProductTransport",
            "status": "RejectedAllHardRiskCandidates",
            "hard_risks": sorted({risk for _name, _rows, status, _metrics in candidates for risk in status.get("hard_risks", [])}),
        }
        chosen_metrics = score_product_transport_candidate(
            forecasts=forecasts,
            shipments=product_transport,
            xls_path=xls_path,
            qtype=qtype,
            days=plan_days,
        )
    final_fixed_reroute_status: dict[str, Any] = {"method": "FixedArrivalRerouteMILP", "status": "Skipped"}
    final_delay_reroute_status: dict[str, Any] = {"method": "DelayToCheaperLanePolish", "status": "Skipped"}
    if (
        is_extreme_mode()
        and qtype == "销售"
        and case_keyword(xls_path) == "热水器"
        and product_transport
    ):
        final_fixed_rows, final_fixed_reroute_status = optimize_fixed_arrival_reroute(
            forecasts=forecasts,
            shipments=product_transport,
            factories=factories,
            routes=routes,
            products=products,
            cargo=product_name,
            carriers=carriers,
            days=plan_days,
            production_rows=explicit_production_rows,
            time_limit_sec=env_int("SUPPLY_CHAIN_FIXED_REROUTE_TIMELIMIT", 30),
        )
        if final_fixed_rows and float(final_fixed_reroute_status.get("delta") or 0.0) > 1e-6:
            fixed_candidate = make_transport_candidate(
                f"{chosen_name}+最终固定到货重路由",
                final_fixed_rows,
                {
                    "method": "FinalFixedArrivalReroute",
                    "fixed_arrival_reroute": final_fixed_reroute_status,
                    "skip_sanitize": True,
                },
            )
            candidates.append(fixed_candidate)
            if not fixed_candidate[2].get("hard_risks") and product_transport_priority(fixed_candidate[3], len(fixed_candidate[1])) > product_transport_priority(chosen_metrics, len(product_transport)):
                chosen_name, product_transport, product_solver_status, chosen_metrics = fixed_candidate
        delay_rows, final_delay_reroute_status = polish_delay_to_cheaper_lanes(
            forecasts=forecasts,
            shipments=product_transport,
            factories=factories,
            routes=routes,
            products=products,
            cargo=product_name,
            carriers=carriers,
            days=plan_days,
            production_rows=explicit_production_rows,
            max_rounds=2,
        )
        if delay_rows and float(final_delay_reroute_status.get("delta") or 0.0) > 1e-6:
            delay_candidate = make_transport_candidate(
                f"{chosen_name}+延后低成本完整路线",
                delay_rows,
                {
                    "method": "FinalDelayToCheaperLanePolish",
                    "delay_to_cheaper_lane": final_delay_reroute_status,
                    "skip_sanitize": True,
                },
            )
            candidates.append(delay_candidate)
            if not delay_candidate[2].get("hard_risks") and product_transport_priority(delay_candidate[3], len(delay_candidate[1])) > product_transport_priority(chosen_metrics, len(product_transport)):
                chosen_name, product_transport, product_solver_status, chosen_metrics = delay_candidate
    destination_subset_status: dict[str, Any] = {"method": "DestinationSubsetMILP", "status": "Skipped"}
    if (
        is_extreme_mode()
        and qtype == "销售"
        and case_keyword(xls_path) == "热水器"
        and product_transport
        and heatwater_subset_audit_enabled()
    ):
        subset_rows, destination_subset_status = optimize_destination_subset_transport(
            forecasts=forecasts,
            shipments=product_transport,
            factories=factories,
            routes=routes,
            products=products,
            cargo=product_name,
            carriers=carriers,
            days=plan_days,
            target_nodes={"北方总代", "武汉总代"},
            production_rows=explicit_production_rows,
            time_limit_sec=env_int("SUPPLY_CHAIN_HEATWATER_SUBSET_TIMELIMIT", 120),
            ship_day_window=env_int("SUPPLY_CHAIN_HEATWATER_SUBSET_WINDOW", 2),
        )
        if subset_rows and float(destination_subset_status.get("delta") or 0.0) > 1e-6:
            subset_candidate = make_transport_candidate(
                f"{chosen_name}+北方武汉子集重算",
                subset_rows,
                {
                    "method": "HeatwaterDestinationSubsetMILP",
                    "destination_subset": destination_subset_status,
                    "skip_sanitize": True,
                },
            )
            candidates.append(subset_candidate)
            if not subset_candidate[2].get("hard_risks") and product_transport_priority(subset_candidate[3], len(subset_candidate[1])) > product_transport_priority(chosen_metrics, len(product_transport)):
                chosen_name, product_transport, product_solver_status, chosen_metrics = subset_candidate
    lower_bound_status: dict[str, Any] = {"status": "Skipped"}
    if is_extreme_mode() and qtype == "销售" and case_keyword(xls_path) == "热水器":
        independent_bound_status = independent_destination_lower_bound(
            forecasts=forecasts,
            factories=factories,
            routes=routes,
            products=products,
            cargo=product_name,
            carriers=carriers,
            days=plan_days,
            production_rows=explicit_production_rows,
            include_shipments=True,
        )
        capacity_bound_status = capacity_aware_transport_lower_bound(
            forecasts=forecasts,
            factories=factories,
            routes=routes,
            products=products,
            cargo=product_name,
            carriers=carriers,
            days=plan_days,
            production_rows=explicit_production_rows,
        )
        bound_candidates = [
            ("independent_destination", independent_bound_status),
            ("capacity_aware_lp", capacity_bound_status),
        ]
        bound_name, lower_bound_status = max(
            bound_candidates,
            key=lambda item: float(item[1].get("freight_lower_bound") or 0.0),
        )
        lower_bound = float(lower_bound_status.get("freight_lower_bound") or 0.0)
        feasible_freight = sum(float(row.get("freight_cost") or 0.0) for row in product_transport)
        if lower_bound > 0:
            gap_diagnostics = heatwater_gap_diagnostics(
                forecasts=forecasts,
                factories=factories,
                final_shipments=product_transport,
                independent_status=independent_bound_status,
                days=plan_days,
                production_rows=explicit_production_rows,
            )
            lower_bound_status = {
                **lower_bound_status,
                "bound_name": bound_name,
                "independent_destination": independent_bound_status,
                "capacity_aware_lp": capacity_bound_status,
                "feasible_freight": round(feasible_freight, 4),
                "gap_to_lower_bound": round(feasible_freight - lower_bound, 4),
                "gap_ratio": round((feasible_freight - lower_bound) / max(feasible_freight, 0.001), 6),
                "gap_diagnostics": gap_diagnostics,
            }
    product_solver_status = {
        **product_solver_status,
        "selected": chosen_name,
        "mode": solver_mode_label(),
        "objective": (
            "高分版：先保证满足率100%，兼顾速度和路线数量；满足率达标后不强求最低成本。"
            if not is_extreme_mode()
            else "极限版：枚举/优化更多运输候选，先保证满足率100%，再选择单位物流成本最低方案。"
        ),
        "candidate_metrics": {
            name: {
                "score": round(metrics["score"], 4),
                "unit_logistics": round(metrics["unit_logistics"], 4),
                "market_satisfaction": round(metrics["market_satisfaction"], 4),
                "shortage": round(metrics["shortage"], 4),
                "hard_risks": int(metrics.get("hard_risk_count", 0)),
                "shipments": len(rows),
            }
            for name, rows, _status, metrics in candidates
        },
        "global_status": global_status,
        "warm_window_global_status": warm_window_status,
        "full_global_audit_status": full_global_audit_status,
        "score_aware_status": score_aware_status,
        "budget_score_status": budget_score_status,
        "inventory_window_status": inventory_window_status,
        "extreme_post_polish_status": extreme_post_polish_status,
        "final_fixed_arrival_reroute_status": final_fixed_reroute_status,
        "final_delay_reroute_status": final_delay_reroute_status,
        "destination_subset_status": destination_subset_status,
        "lower_bound_status": lower_bound_status,
        "stage_timings": stage_timings,
    }
    if explicit_production_rows is not None:
        production_rows = explicit_production_rows
    else:
        production_amounts = production_amounts_from_transport(factories, product_transport)
        production_rows = production_rows_from_amounts(factories, product_name, production_amounts, plan_days)

    material_plan: dict[str, Any] | None = None
    if qtype == "生产" and suppliers and factory_materials:
        material_plan = material_procurement_for_production(
            production_rows=production_rows,
            factories=factories,
            factory_materials=factory_materials,
            products=products,
            suppliers=suppliers,
            routes=routes,
            rates=rates,
            carriers=carriers,
            days=plan_days,
            name_prefix=f"{case_keyword(xls_path)}_{qtype}_material",
            note="生产题逐日原料补货",
        )
    procurement_rows = material_plan["procurement"] if material_plan else []
    material_transport_rows = material_plan["shipments"] if material_plan else []

    material_rows: list[dict[str, Any]] = []
    for prod in production_rows:
        for item in factory_materials:
            if item.factory != prod["factory"]:
                continue
            need = prod["amount"] * material_bom(products, item.material)
            shortage = max(0.0, need - item.init)
            material_rows.append(
                {
                    "material": item.material,
                    "factory": item.factory,
                    "need": need,
                    "init": item.init,
                    "shortage": shortage,
                    "note": "库存覆盖" if shortage <= 0 else "需要补采，否则有断料风险",
                }
            )

    market_replay = daily_market_replay(forecasts, product_transport, plan_days)
    total_need = market_replay["demand"] or sum(row["forecast"] for row in forecasts)
    market_satisfaction = market_replay["market_satisfaction"] if forecasts else 1.0
    if qtype == "生产":
        total_prod_need = sum(max(0.0, float(factory.daily or 0.0) * plan_days) for factory in factories)
        total_prod_supply = sum(row["amount"] for row in production_rows)
    else:
        total_prod_need = sum(demand_by_factory.values())
        total_prod_supply = sum(factory.init for factory in factories) + sum(row["amount"] for row in production_rows)
    production_satisfaction = min(1.0, total_prod_supply / max(total_prod_need, 0.001)) if total_prod_need else 1.0
    if material_plan:
        production_satisfaction = min(production_satisfaction, float(material_plan.get("production_satisfaction", 1.0)))
    logistics_total = sum(row["freight_cost"] for row in product_transport)
    logistics_denominator = (
        market_replay["served"]
        if qtype == "销售" and forecasts
        else sum(row["amount"] for row in product_transport)
    )
    unit_logistics = logistics_total / max(logistics_denominator, 0.001)
    if qtype == "销售":
        targets = score_targets(xls_path, "销售")
        points = score_points(xls_path, "销售")
        prediction_deviation = 0.0
        score_rows = [
            {
                "item": "预测偏差率",
                "target": targets["prediction_deviation"],
                "actual": prediction_deviation,
                "points": deviation_score(prediction_deviation, targets["prediction_deviation"], points["预测偏差率"]),
                "max": points["预测偏差率"],
            },
            {
                "item": "单位物流成本",
                "target": targets["unit_logistics_cost"],
                "actual": unit_logistics,
                "points": cost_score(unit_logistics, targets["unit_logistics_cost"], points["单位物流成本"]),
                "max": points["单位物流成本"],
            },
            {
                "item": "市场满足率",
                "target": targets["market_satisfaction"],
                "actual": market_satisfaction,
                "points": satisfaction_score(market_satisfaction, points["市场满足率"]),
                "max": points["市场满足率"],
            },
        ]
        score_note = "候选分：销售题按平台三项公式估算；成品运输为逐日整数优化，预测偏差率需提交后复原"
    else:
        targets = score_targets(xls_path, "生产")
        points = score_points(xls_path, "生产")
        target_logistics = targets.get("unit_logistics_cost", 0.0)
        logistics_points = cost_score(unit_logistics, target_logistics, points["单位物流成本"]) if target_logistics else 0.0
        score_rows = [
            {"item": "单位物流成本", "target": target_logistics, "actual": unit_logistics, "points": logistics_points, "max": points["单位物流成本"]},
            {"item": "市场满足率", "target": targets["market_satisfaction"], "actual": market_satisfaction, "points": satisfaction_score(market_satisfaction, points["市场满足率"]), "max": points["市场满足率"]},
        ]
        score_note = "候选分：生产题按平台公式估算；成品运输为逐日整数优化，市场满足率来自逐日库存仿真"
    simulation = simulate_plan(
        sales_fc=forecasts,
        factories=factories,
        factory_materials=factory_materials,
        production_rows=production_rows,
        procurement_rows=material_transport_rows,
        product_transport_rows=product_transport,
        products=products,
        days=plan_days,
    )
    if material_plan and material_plan.get("failures"):
        simulation["risks"] = list(dict.fromkeys(list(simulation.get("risks", [])) + list(material_plan.get("failures", []))))
        simulation["ok"] = False
        simulation["summary"] = "逐日复核存在缺口；未找到满满足率候选时按满足率优先保留"
    return {
        "qtype": qtype,
        "products": products,
        "sales_fc": forecasts,
        "production": production_rows,
        "material_need": material_rows,
        "procurement": procurement_rows,
        "material_transport": material_transport_rows,
        "product_transport": product_transport,
        "transport_sections": _compact_transport_rows(material_transport_rows + product_transport),
        "score_rows": score_rows,
        "score": sum(row["points"] for row in score_rows),
        "score_note": score_note,
        "unit_procurement": 0.0,
        "unit_logistics": unit_logistics,
        "production_satisfaction": production_satisfaction,
        "market_satisfaction": market_satisfaction,
        "simulation": simulation,
        "market_replay": market_replay,
        "plan_days": plan_days,
        "solver_status": {
            "product_transport": product_solver_status,
            "material_transport": material_plan.get("status", {}) if material_plan else {},
        },
        "assumptions": ["销售量按移动加权平均或逐日数据汇总", "安全库存按日均波动设置 1-3 天", "销售网点按工厂到网点路线成本分配供货工厂"],
    }


def _solve_day_transport_milp(
    *,
    name: str,
    sources: dict[str, dict[str, float]],
    destinations: dict[str, dict[str, Any]],
    routes: list[Route],
    products: list[Product],
    cargo: str,
    carriers: list[Any] | None = None,
    days: int = PLAN_DAYS,
    ship_day_step: int = 1,
    gap_rel: float = 0.001,
    enforce_destination_limits: bool = False,
    max_total_shortage: float | None = None,
    shortage_penalty: float = 100_000_000.0,
    freight_budget_per_served: float | None = None,
    freight_budget_per_shipped: float | None = None,
    strict_freight_upper_bound: float | None = None,
    allow_fallback: bool = True,
    time_limit_sec: int | None = None,
    warm_start_shipments: list[dict[str, Any]] | None = None,
    allowed_ship_days_by_route: dict[tuple[str, str, str], set[int]] | None = None,
    allowed_ship_days_by_option: dict[tuple[str, str, str, str], set[int]] | None = None,
    relax_integrality: bool = False,
    prune_transport_options: bool = True,
    strict_bound_feasibility_only: bool = False,
) -> dict[str, Any]:
    try:
        import pulp
    except ImportError as exc:
        raise RuntimeError("缺少 PuLP，无法执行整数运输优化") from exc

    ratio = charge_ratio(products, cargo)
    source_total_available = {
        src: max(0.0, float(info.get("initial", 0.0) or 0.0) + sum(float(value or 0.0) for value in info.get("supply", [])))
        for src, info in sources.items()
    }
    destination_total_demand = {
        dst: max(0.0, sum(float(value or 0.0) for value in info.get("demand", [])))
        for dst, info in destinations.items()
    }
    destination_total_receivable: dict[str, float] = {}
    for dst, info in destinations.items():
        demand_total = destination_total_demand.get(dst, 0.0)
        limit = float(info.get("limit", 0.0) or 0.0)
        initial = float(info.get("initial", 0.0) or 0.0)
        if enforce_destination_limits and limit > 0:
            destination_total_receivable[dst] = max(0.0, demand_total + max(0.0, limit - initial))
        else:
            destination_total_receivable[dst] = max(
                demand_total + max(0.0, limit - initial) if limit > 0 else demand_total,
                demand_total,
                1.0,
            )
    max_amount = max(
        sum(source_total_available.values()),
        sum(destination_total_demand.values()),
        1.0,
    )
    lane_options: list[LaneOption] = []
    lane_upper_bounds: dict[int, float] = {}
    raw_lane_option_count = 0
    for route in routes:
        if route.src not in sources or route.dst not in destinations:
            continue
        lane_upper = max(
            1.0,
            min(
                source_total_available.get(route.src, max_amount) or max_amount,
                destination_total_receivable.get(route.dst, max_amount) or max_amount,
            ),
        )
        raw_options = route_transport_options(
            route,
            ratio,
            carriers,
            max_amount=lane_upper,
            prune_dominated=False,
        )
        raw_lane_option_count += len(raw_options)
        options = prune_dominated_transport_options(raw_options, lane_upper) if is_extreme_mode() and prune_transport_options else raw_options
        for option in options:
            lane_id = len(lane_options)
            lane_options.append(LaneOption(lane_id, route, option))
            lane_upper_bounds[lane_id] = lane_upper
    if not lane_options:
        return {"status": "NoRoute", "shipments": [], "freight_cost": 0.0, "failures": [f"{cargo} 无可用运输路线"]}

    model = pulp.LpProblem(name, pulp.LpMinimize)
    zero_shortage_required = max_total_shortage is not None and float(max_total_shortage) <= 0
    route_by_lane = {lane.lane_id: lane.route for lane in lane_options}
    option_by_lane = {lane.lane_id: lane.option for lane in lane_options}
    x: dict[tuple[int, int], Any] = {}
    y: dict[tuple[int, int], Any] = {}
    fcost: dict[tuple[int, int], Any] = {}
    piecewise_cost_constraints = 0
    source_inv: dict[tuple[str, int], Any] = {}
    source_supply_var: dict[tuple[str, int], Any] = {}
    dst_inv: dict[tuple[str, int], Any] = {}
    shortage: dict[tuple[str, int], Any] = {}
    over_limit: dict[tuple[str, int], Any] = {}
    cumulative_inventory_model = bool(is_extreme_mode() and zero_shortage_required and enforce_destination_limits)

    ship_days_by_lane: dict[int, list[int]] = {}
    for lane in lane_options:
        lane_id = lane.lane_id
        option = lane.option
        route = route_by_lane[lane_id]
        latest_ship_day = max(1, days - int(option.lead or 0))
        route_key = (route.route, route.src, route.dst)
        option_key = (route.route, route.src, route.dst, option.carrier)
        if allowed_ship_days_by_option is not None:
            lane_ship_days = sorted(
                int(day)
                for day in allowed_ship_days_by_option.get(option_key, set())
                if 1 <= int(day) <= latest_ship_day
            )
        elif allowed_ship_days_by_route is not None:
            lane_ship_days = sorted(
                int(day)
                for day in allowed_ship_days_by_route.get(route_key, set())
                if 1 <= int(day) <= latest_ship_day
            )
        elif ship_day_step <= 1:
            lane_ship_days = list(range(1, latest_ship_day + 1))
        else:
            lane_ship_days = sorted({1, latest_ship_day, *range(1, latest_ship_day + 1, ship_day_step)})
        ship_days_by_lane[lane_id] = lane_ship_days
        for day in lane_ship_days:
            key = (lane_id, day)
            lane_upper = max(1.0, float(lane_upper_bounds.get(lane_id, max_amount) or max_amount))
            src_info = sources.get(route.src, {})
            src_supply = list(src_info.get("supply", []))
            source_cumulative_available = float(src_info.get("initial", 0.0) or 0.0) + sum(float(value or 0.0) for value in src_supply[:day])
            if source_cumulative_available > 0:
                lane_upper = min(lane_upper, max(1.0, source_cumulative_available))
            arrival_day = day + int(option.lead or 0)
            dst_info = destinations.get(route.dst, {})
            dst_demand = list(dst_info.get("demand", []))
            if enforce_destination_limits and 1 <= arrival_day <= days:
                remaining_dst_demand = sum(float(value or 0.0) for value in dst_demand[arrival_day - 1 :])
                dst_limit = float(dst_info.get("limit", 0.0) or 0.0)
                if dst_limit > 0:
                    lane_upper = min(lane_upper, max(1.0, remaining_dst_demand + dst_limit))
            amount_category = "Continuous" if relax_integrality else "Integer"
            binary_category = "Continuous" if relax_integrality else "Binary"
            x[key] = pulp.LpVariable(f"x_{lane_id}_{day}", lowBound=0, upBound=lane_upper, cat=amount_category)
            y[key] = pulp.LpVariable(f"y_{lane_id}_{day}", lowBound=0, upBound=1, cat=binary_category)
            model += x[key] <= lane_upper * y[key]
            model += x[key] >= y[key]
            option_cost_upper = sum(max(float(slope) * lane_upper, float(minimum)) for slope, minimum in option.segment_costs)
            fcost[key] = pulp.LpVariable(f"freight_{lane_id}_{day}", lowBound=0, upBound=option_cost_upper)
            segment_count = len(option.segment_costs)
            for mask in range(1 << segment_count):
                expr_terms = []
                for segment_idx, (slope, minimum) in enumerate(option.segment_costs):
                    if mask & (1 << segment_idx):
                        expr_terms.append(float(slope) * x[key])
                    else:
                        expr_terms.append(float(minimum) * y[key])
                model += fcost[key] >= pulp.lpSum(expr_terms)
                piecewise_cost_constraints += 1

    warm_start_used = 0
    if warm_start_shipments:
        warm_by_key: dict[tuple[str, str, str, str, int], float] = defaultdict(float)
        for row in warm_start_shipments:
            route_name = sv(row.get("route"))
            source = sv(row.get("source") or row.get("factory"))
            destination = sv(row.get("destination"))
            carrier = sv(row.get("carrier"))
            ship_day = int(row.get("ship_day") or 0)
            amount = float(row.get("amount") or 0.0)
            if not route_name or not source or not destination or ship_day <= 0 or amount <= 0:
                continue
            warm_by_key[(route_name, source, destination, carrier, ship_day)] += amount
        for lane_id, route in route_by_lane.items():
            option = option_by_lane[lane_id]
            for day in ship_days_by_lane.get(lane_id, []):
                amount = warm_by_key.get((route.route, route.src, route.dst, option.carrier, day), 0.0)
                key = (lane_id, day)
                if key not in x:
                    continue
                upper = float(x[key].upBound or max_amount)
                amount = max(0.0, min(float(amount), upper))
                try:
                    x[key].setInitialValue(int(round(amount)))
                    y[key].setInitialValue(1 if amount > 0 else 0)
                    warm_start_used += 1 if amount > 0 else 0
                except Exception:
                    pass

    for src, info in sources.items():
        supply = list(info.get("supply", []))
        supply_is_capacity = bool(info.get("supply_is_capacity", False))
        if cumulative_inventory_model:
            for day in range(1, days + 1):
                cumulative_outbound = [
                    x[(lane_id, ship_day)]
                    for lane_id, route in route_by_lane.items()
                    if route.src == src
                    for ship_day in ship_days_by_lane.get(lane_id, [])
                    if ship_day <= day and (lane_id, ship_day) in x
                ]
                cumulative_supply = sum(float(value or 0.0) for value in supply[:day])
                model += pulp.lpSum(cumulative_outbound) <= float(info.get("initial", 0.0) or 0.0) + cumulative_supply
            max_total = float(info.get("max_total", 0.0) or 0.0)
            if max_total > 0:
                all_outbound = [
                    x[(lane_id, day)]
                    for lane_id, route in route_by_lane.items()
                    if route.src == src
                    for day in ship_days_by_lane.get(lane_id, [])
                    if (lane_id, day) in x
                ]
                model += pulp.lpSum(all_outbound) <= max_total
            continue
        for day in range(1, days + 1):
            source_inv[(src, day)] = pulp.LpVariable(
                f"src_inv_{len(source_inv)}",
                lowBound=0,
                upBound=info.get("limit") if (info.get("limit") or 0) > 0 else None,
            )
            supply_value = supply[day - 1] if day - 1 < len(supply) else 0.0
            if supply_is_capacity:
                source_supply_var[(src, day)] = pulp.LpVariable(
                    f"src_supply_{len(source_supply_var)}",
                    lowBound=0,
                    upBound=supply_value,
                    cat="Continuous" if relax_integrality else "Integer",
                )
                supply_expr = source_supply_var[(src, day)]
            else:
                supply_expr = supply_value
            outbound = [
                x[(lane_id, day)]
                for lane_id, route in route_by_lane.items()
                if route.src == src
                if (lane_id, day) in x
            ]
            prev = info.get("initial", 0.0) if day == 1 else source_inv[(src, day - 1)]
            model += source_inv[(src, day)] == prev + supply_expr - pulp.lpSum(outbound)
            cumulative_outbound = [
                x[(lane_id, ship_day)]
                for lane_id, route in route_by_lane.items()
                if route.src == src
                for ship_day in ship_days_by_lane.get(lane_id, [])
                if ship_day <= day and (lane_id, ship_day) in x
            ]
            cumulative_supply = sum(float(value or 0.0) for value in supply[:day])
            model += pulp.lpSum(cumulative_outbound) <= float(info.get("initial", 0.0) or 0.0) + cumulative_supply
        max_total = float(info.get("max_total", 0.0) or 0.0)
        if max_total > 0:
            all_outbound = [
                x[(lane_id, day)]
                for lane_id, route in route_by_lane.items()
                if route.src == src
                for day in ship_days_by_lane.get(lane_id, [])
                if (lane_id, day) in x
            ]
            model += pulp.lpSum(all_outbound) <= max_total

    for dst, info in destinations.items():
        demand = list(info.get("demand", []))
        inbound_all = [
            x[(lane_id, ship_day)]
            for lane_id, route in route_by_lane.items()
            if route.dst == dst
            for ship_day in ship_days_by_lane.get(lane_id, [])
            if (lane_id, ship_day) in x
        ]
        if cumulative_inventory_model:
            initial = float(info.get("initial", 0.0) or 0.0)
            total_demand = sum(float(value or 0.0) for value in demand)
            total_required = max(0.0, total_demand - initial)
            if inbound_all:
                model += pulp.lpSum(inbound_all) == total_required
            elif total_required > 1e-6:
                model += 0 >= total_required
            limit = float(info.get("limit", 0.0) or 0.0)
            for day in range(1, days + 1):
                cumulative_arrivals = []
                for lane_id, route in route_by_lane.items():
                    if route.dst != dst:
                        continue
                    lead = int(option_by_lane[lane_id].lead or 0)
                    for ship_day in ship_days_by_lane.get(lane_id, []):
                        if ship_day + lead <= day and (lane_id, ship_day) in x:
                            cumulative_arrivals.append(x[(lane_id, ship_day)])
                cumulative_demand = sum(float(value or 0.0) for value in demand[:day])
                model += initial + pulp.lpSum(cumulative_arrivals) >= cumulative_demand
                if limit > 0:
                    model += initial + pulp.lpSum(cumulative_arrivals) <= cumulative_demand + limit
            continue
        if inbound_all:
            initial = float(info.get("initial", 0.0) or 0.0)
            total_demand = sum(float(value or 0.0) for value in demand)
            limit = float(info.get("limit", 0.0) or 0.0)
            if zero_shortage_required:
                model += pulp.lpSum(inbound_all) >= max(0.0, total_demand - initial)
            if enforce_destination_limits and limit > 0:
                model += pulp.lpSum(inbound_all) <= max(0.0, total_demand + limit - initial)
        for day in range(1, days + 1):
            dst_inv[(dst, day)] = pulp.LpVariable(
                f"dst_inv_{len(dst_inv)}",
                lowBound=0,
            )
            if zero_shortage_required:
                shortage_expr = 0.0
            else:
                shortage[(dst, day)] = pulp.LpVariable(f"short_{len(shortage)}", lowBound=0)
                shortage_expr = shortage[(dst, day)]
            over_limit[(dst, day)] = pulp.LpVariable(f"over_{len(over_limit)}", lowBound=0)
            arrivals = []
            for lane_id, route in route_by_lane.items():
                if route.dst != dst:
                    continue
                ship_day = day - int(option_by_lane[lane_id].lead or 0)
                if (lane_id, ship_day) in x:
                    arrivals.append(x[(lane_id, ship_day)])
            prev = info.get("initial", 0.0) if day == 1 else dst_inv[(dst, day - 1)]
            model += dst_inv[(dst, day)] == prev + pulp.lpSum(arrivals) - (demand[day - 1] if day - 1 < len(demand) else 0.0) + shortage_expr
            if zero_shortage_required:
                cumulative_arrivals = []
                for lane_id, route in route_by_lane.items():
                    if route.dst != dst:
                        continue
                    lead = int(option_by_lane[lane_id].lead or 0)
                    for ship_day in ship_days_by_lane.get(lane_id, []):
                        if ship_day + lead <= day and (lane_id, ship_day) in x:
                            cumulative_arrivals.append(x[(lane_id, ship_day)])
                model += float(info.get("initial", 0.0) or 0.0) + pulp.lpSum(cumulative_arrivals) >= sum(float(value or 0.0) for value in demand[:day])
            limit = float(info.get("limit", 0.0) or 0.0)
            if limit > 0:
                model += over_limit[(dst, day)] >= dst_inv[(dst, day)] - limit
                if enforce_destination_limits:
                    model += dst_inv[(dst, day)] <= limit
                    cumulative_arrivals_for_limit = []
                    for lane_id, route in route_by_lane.items():
                        if route.dst != dst:
                            continue
                        lead = int(option_by_lane[lane_id].lead or 0)
                        for ship_day in ship_days_by_lane.get(lane_id, []):
                            if ship_day + lead <= day and (lane_id, ship_day) in x:
                                cumulative_arrivals_for_limit.append(x[(lane_id, ship_day)])
                    model += float(info.get("initial", 0.0) or 0.0) + pulp.lpSum(cumulative_arrivals_for_limit) <= sum(float(value or 0.0) for value in demand[:day]) + limit
    total_shortage_expr = pulp.lpSum(shortage.values())
    total_freight_expr = pulp.lpSum(fcost.values())
    total_shipped_expr = pulp.lpSum(x.values())
    total_demand_value = sum(sum(info.get("demand", [])) for info in destinations.values())
    if max_total_shortage is not None and float(max_total_shortage) > 0:
        model += total_shortage_expr <= max(0.0, float(max_total_shortage))
    if freight_budget_per_served is not None:
        model += total_freight_expr <= float(freight_budget_per_served) * (total_demand_value - total_shortage_expr)
    if freight_budget_per_shipped is not None:
        model += total_freight_expr <= float(freight_budget_per_shipped) * total_shipped_expr
    if strict_freight_upper_bound is not None:
        model += total_freight_expr <= max(0.0, float(strict_freight_upper_bound))

    source_cost_terms = []
    for (lane_id, day), var in x.items():
        unit_cost = float(sources.get(route_by_lane[lane_id].src, {}).get("unit_cost", 0.0) or 0.0)
        if unit_cost:
            source_cost_terms.append(unit_cost * var)
    over_cost_terms = []
    for (dst, _day), var in over_limit.items():
        fee = float(destinations.get(dst, {}).get("excess_fee", 0.0) or 0.0)
        if fee:
            over_cost_terms.append(fee * var)
    source_supply_terms = [0.001 * var for var in source_supply_var.values()]
    if strict_bound_feasibility_only and strict_freight_upper_bound is not None:
        objective = (
            0.001 * total_shipped_expr
            + pulp.lpSum(source_supply_terms)
            + float(shortage_penalty) * total_shortage_expr
        )
    else:
        objective = (
            total_freight_expr
            + pulp.lpSum(source_cost_terms)
            + pulp.lpSum(over_cost_terms)
            + pulp.lpSum(source_supply_terms)
            + float(shortage_penalty) * total_shortage_expr
        )
    model += objective
    solver_kwargs: dict[str, Any] = {"msg": False, "gapRel": gap_rel}
    if time_limit_sec is not None:
        solver_kwargs["timeLimit"] = int(time_limit_sec)
    if warm_start_shipments:
        solver_kwargs["warmStart"] = True
        if os.name == "nt":
            solver_kwargs["keepFiles"] = True
    cbc_log_path: Path | None = None
    if time_limit_sec is not None:
        cbc_log_path = Path(tempfile.gettempdir()) / f"{safe_stem(name)}_cbc.log"
        solver_kwargs["logPath"] = str(cbc_log_path)
    solve_started_at = time.time()
    status_code = model.solve(pulp.PULP_CBC_CMD(**solver_kwargs))
    solve_seconds = time.time() - solve_started_at
    if solver_kwargs.get("keepFiles"):
        for suffix in ("mps", "mst", "sol", "lp"):
            artifact = Path(f"{name}-pulp.{suffix}")
            try:
                if artifact.exists() and artifact.is_file():
                    artifact.unlink()
            except OSError:
                pass
    raw_solver_status = pulp.LpStatus[status_code]
    cbc_log_text = ""
    if cbc_log_path is not None and cbc_log_path.exists():
        try:
            cbc_log_text = cbc_log_path.read_text(encoding="utf-8", errors="ignore")
            cbc_log_path.unlink(missing_ok=True)
        except OSError:
            pass
    cbc_time_limit = "Stopped on time limit" in cbc_log_text or "Result - Stopped on time limit" in cbc_log_text
    cbc_optimal = "Result - Optimal solution found" in cbc_log_text
    time_limit_hit = bool(time_limit_sec is not None and (cbc_time_limit or (solve_seconds >= max(0.0, float(time_limit_sec) - 0.5) and not cbc_optimal)))
    solver_status = "TimeLimitFeasible" if raw_solver_status == "Optimal" and time_limit_hit else raw_solver_status
    acceptable_solution_statuses = {"Optimal", "TimeLimitFeasible"}
    if solver_status not in acceptable_solution_statuses:
        if not allow_fallback:
            return {
                "status": solver_status,
                "raw_status": raw_solver_status,
                "solve_seconds": solve_seconds,
                "time_limit_hit": time_limit_hit,
                "shipments": [],
                "freight_cost": 0.0,
                "objective": 0.0,
                "model_stats": {
                    "routes": len(routes),
                    "raw_lane_options": raw_lane_option_count,
                    "lane_options": len(lane_options),
                    "shipment_vars": len(x),
                    "binary_vars": len(y),
                    "freight_cost_vars": len(fcost),
                    "piecewise_cost_constraints": piecewise_cost_constraints,
                    "source_inventory_vars": len(source_inv),
                    "destination_inventory_vars": len(dst_inv),
                    "shortage_vars": len(shortage),
                    "over_limit_vars": len(over_limit),
                    "max_amount_global": max_amount,
                    "max_lane_upper": max(lane_upper_bounds.values()) if lane_upper_bounds else 0.0,
                    "warm_start_shipments": len(warm_start_shipments or []),
                    "warm_start_used": warm_start_used,
                    "allowed_ship_day_windows": len(allowed_ship_days_by_route or {}),
                    "allowed_ship_day_option_windows": len(allowed_ship_days_by_option or {}),
                    "relax_integrality": bool(relax_integrality),
                    "prune_transport_options": bool(prune_transport_options),
                    "strict_bound_feasibility_only": bool(strict_bound_feasibility_only and strict_freight_upper_bound is not None),
                },
                "gap_rel": gap_rel,
                "shortage": total_demand_value,
                "max_total_shortage": max_total_shortage,
                "shortage_penalty": shortage_penalty,
                "freight_budget_per_served": freight_budget_per_served,
                "freight_budget_per_shipped": freight_budget_per_shipped,
                "strict_freight_upper_bound": strict_freight_upper_bound,
                "failures": [f"{cargo} 求解未得到可采纳整数解: {solver_status}"],
                "source_supply": {src: 0 for src in sources},
            }
        return _transport_greedy_fallback(
            status=solver_status,
            sources=sources,
            destinations=destinations,
            routes=routes,
            products=products,
            cargo=cargo,
            carriers=carriers,
            days=days,
        )

    shipments: list[dict[str, Any]] = []
    for lane_id, route in route_by_lane.items():
        option = option_by_lane[lane_id]
        for day in ship_days_by_lane.get(lane_id, []):
            amount = int(round(pulp.value(x[(lane_id, day)]) or 0))
            if amount <= 0:
                continue
            freight = transport_option_cost(option, amount)
            shipments.append(
                {
                    "cargo": cargo,
                    "source": route.src,
                    "destination": route.dst,
                    "amount": amount,
                    "ship_day": day,
                    "arrival_day": day + int(option.lead or 0),
                    "route": route.route,
                    "mode": route_mode(route),
                    "lead": int(option.lead or 0),
                    "freight_cost": freight,
                    "carrier": option.carrier,
                }
            )
    failures = []
    total_shortage = 0.0
    for (dst, day), var in shortage.items():
        short = float(pulp.value(var) or 0.0)
        total_shortage += short
        if short > 1e-6:
            failures.append(f"{cargo}-{dst} 第{day}天缺口 {short:.0f}")
    if solver_status in {"Infeasible", "Unbounded", "Undefined"} or (solver_status != "Optimal" and not shipments):
        if not allow_fallback:
            return {
                "status": solver_status,
                "raw_status": raw_solver_status,
                "solve_seconds": solve_seconds,
                "time_limit_hit": time_limit_hit,
                "shipments": [],
                "freight_cost": 0.0,
                "objective": 0.0,
                "gap_rel": gap_rel,
                "shortage": total_demand_value,
                "max_total_shortage": max_total_shortage,
                "shortage_penalty": shortage_penalty,
                "freight_budget_per_served": freight_budget_per_served,
                "freight_budget_per_shipped": freight_budget_per_shipped,
                "strict_freight_upper_bound": strict_freight_upper_bound,
                "failures": [f"{cargo} 求解不可行: {solver_status}"],
                "source_supply": {src: 0 for src in sources},
            }
        return _transport_greedy_fallback(
            status=solver_status,
            sources=sources,
            destinations=destinations,
            routes=routes,
            products=products,
            cargo=cargo,
            carriers=carriers,
            days=days,
        )
    return {
        "status": solver_status,
        "raw_status": raw_solver_status,
        "solve_seconds": solve_seconds,
        "time_limit_hit": time_limit_hit,
        "shipments": sorted(shipments, key=lambda row: (row["ship_day"], row["route"], row["amount"])),
        "freight_cost": sum(row["freight_cost"] for row in shipments),
        "objective": float(pulp.value(model.objective) or 0.0),
        "model_stats": {
            "routes": len(routes),
            "raw_lane_options": raw_lane_option_count,
            "lane_options": len(lane_options),
            "shipment_vars": len(x),
            "binary_vars": len(y),
            "freight_cost_vars": len(fcost),
            "piecewise_cost_constraints": piecewise_cost_constraints,
            "source_inventory_vars": len(source_inv),
            "destination_inventory_vars": len(dst_inv),
            "shortage_vars": len(shortage),
            "over_limit_vars": len(over_limit),
            "max_amount_global": max_amount,
            "max_lane_upper": max(lane_upper_bounds.values()) if lane_upper_bounds else 0.0,
            "warm_start_shipments": len(warm_start_shipments or []),
            "warm_start_used": warm_start_used,
            "allowed_ship_day_windows": len(allowed_ship_days_by_route or {}),
            "allowed_ship_day_option_windows": len(allowed_ship_days_by_option or {}),
            "relax_integrality": bool(relax_integrality),
            "prune_transport_options": bool(prune_transport_options),
            "strict_bound_feasibility_only": bool(strict_bound_feasibility_only and strict_freight_upper_bound is not None),
        },
        "gap_rel": gap_rel,
        "shortage": total_shortage,
        "max_total_shortage": max_total_shortage,
        "shortage_penalty": shortage_penalty,
        "freight_budget_per_served": freight_budget_per_served,
        "freight_budget_per_shipped": freight_budget_per_shipped,
        "strict_freight_upper_bound": strict_freight_upper_bound,
        "failures": failures,
        "source_supply": {
            src: sum(int(round(pulp.value(source_supply_var[(src, day)]) or 0)) for day in range(1, days + 1) if (src, day) in source_supply_var)
            for src in sources
        },
    }


def _solve_material_procurement_milp(
    *,
    material: FactoryMaterial,
    production_daily: list[int],
    products: list[Product],
    suppliers: list[Supplier],
    routes: list[Route],
    rates: dict[tuple[str, str], float],
    carriers: list[Any] | None = None,
    days: int = PLAN_DAYS,
    name: str | None = None,
    note: str = "逐日原料补货",
) -> dict[str, Any]:
    demand = [qty * material_bom(products, material.material) for qty in production_daily]
    supplier_sources = {}
    for supplier in suppliers:
        if supplier.material != material.material:
            continue
        supplier_sources[supplier.name] = {
            "initial": supplier.init,
            "supply": supplier_supply_profile(supplier, days),
            "unit_cost": currency_to_cny(supplier.price, supplier.currency, rates),
            "max_total": supplier.available if supplier.available > 0 else 0.0,
        }
    factory_raw_material_destinations = {
        material.factory: {
            "initial": material.init,
            "demand": demand,
            "limit": material.limit,
        }
    }
    usable_routes = [route for route in routes if route.dst == material.factory and route.src in supplier_sources]
    transport = _solve_day_transport_milp(
        name=name or f"material_{material.factory}_{material.material}",
        sources=supplier_sources,
        destinations=factory_raw_material_destinations,
        routes=usable_routes,
        products=products,
        cargo=material.material,
        carriers=carriers,
        days=days,
        ship_day_step=2 if days >= 30 else 1,
        gap_rel=0.001,
    )
    supplier_by_name = {supplier.name: supplier for supplier in suppliers}
    purchase_cost = 0.0
    for row in transport["shipments"]:
        supplier = supplier_by_name.get(row["source"])
        if supplier:
            row["material"] = material.material
            row["supplier"] = supplier.name
            row["factory"] = material.factory
            row["unit_price"] = currency_to_cny(supplier.price, supplier.currency, rates)
            row["purchase_cost"] = row["amount"] * row["unit_price"]
            row["note"] = note
            purchase_cost += row["purchase_cost"]
    transport["purchase_cost"] = purchase_cost
    return transport


def _compact_transport_rows(shipments: list[dict[str, Any]]) -> dict[str, list[list[Any]]]:
    grouped: dict[tuple[str, str, str, int, int], list[int]] = {}
    for row in shipments:
        key = (
            row.get("cargo", ""),
            row["route"],
            row.get("carrier", ""),
            int(row["amount"]),
            int(row.get("lead") or 0),
        )
        grouped.setdefault(key, []).append(int(row["ship_day"]))

    sections: dict[str, list[list[Any]]] = {}
    for (cargo, route, carrier, amount, lead), days in sorted(grouped.items()):
        days = sorted(days)
        if not days:
            continue
        run_start = days[0]
        prev = days[0]
        prev_gap: int | None = None
        runs: list[tuple[int, int, int]] = []
        count = 1
        for day in days[1:]:
            gap = day - prev
            if prev_gap is None:
                prev_gap = gap
            if gap != prev_gap:
                runs.append((run_start, count, prev_gap or 1))
                run_start = day
                count = 1
                prev_gap = None
            else:
                count += 1
            prev = day
        runs.append((run_start, count, prev_gap or 1))
        title = cargo or "运输计划"
        for start_day, trips, interval in runs:
            sections.setdefault(title, []).append([route, carrier, amount, start_day, trips, interval, lead])
    for rows in sections.values():
        rows.sort(key=lambda row: (int(row[3]), str(row[0]), str(row[1]), int(row[2])))
    return sections


def solve_tv_comprehensive(sections: dict[str, list[tuple[Any, ...]]], xls_path: Path) -> dict[str, Any]:
    plan_days = plan_days_for_case(xls_path, "综合")
    products = parse_products(sections)
    factories = parse_factories(sections)
    factory_materials = parse_factory_materials(sections)
    suppliers = parse_suppliers(sections)
    routes = parse_routes(sections)
    rates = parse_rates(sections)
    carriers = parse_carriers_safe(sections, rates)
    sales = parse_sales(sections, plan_days)
    use_bias = forecast_bias_for_case(xls_path, "综合")
    forecasts = [forecast_node(node, plan_days, use_bias=use_bias) for node in sales]
    if not factories or not forecasts:
        return solve_comprehensive(sections, xls_path)

    factory = factories[0]
    product_name = factory.product or next((p.name for p in products if p.kind == "产品"), "")
    sales_init_total = sum(row["init"] for row in forecasts)
    forecast_total = sum(row["forecast"] for row in forecasts)
    product_destinations = {
        row["node"]: {"initial": row["init"], "demand": spread_integer(row["forecast"], plan_days), "limit": row["limit"]}
        for row in forecasts
    }
    product_routes = [route for route in routes if route.src == factory.name and route.dst in product_destinations]
    targets = score_targets(xls_path, "综合")
    prediction_deviation = 0.0

    def build_candidate(production_target: int) -> dict[str, Any]:
        production_daily = spread_integer(production_target, plan_days)
        production_rows = [{
            "factory": factory.name,
            "product": product_name,
            "amount": production_target,
            "capacity": factory.daily * plan_days,
            "init": factory.init,
            "daily_schedule": production_daily,
        }]
        product_sources = {
            factory.name: {
                "initial": factory.init,
                "supply": production_daily,
                "limit": 0.0,
            }
        }
        product_transport = _solve_day_transport_milp(
            name=f"tv_product_transport_{production_target}",
            sources=product_sources,
            destinations=product_destinations,
            routes=product_routes,
            products=products,
            cargo=product_name,
            carriers=carriers,
            days=plan_days,
            ship_day_step=2,
            gap_rel=0.001,
            enforce_destination_limits=True,
        )
        product_shipments = fill_market_shortages_with_fast_routes(
            forecasts=forecasts,
            shipments=product_transport["shipments"],
            assignment={row["node"]: factory for row in forecasts},
            routes=routes,
            products=products,
            cargo=product_name,
            carriers=carriers,
            factories=factories,
            days=plan_days,
            max_rounds=2,
            production_rows=production_rows,
        )
        product_shipments = sanitize_product_transport(
            forecasts=forecasts,
            shipments=product_shipments,
            factories=factories,
            routes=routes,
            products=products,
            cargo=product_name,
            carriers=carriers,
            days=plan_days,
            production_rows=production_rows,
        )
        product_replay = daily_market_replay(forecasts, product_shipments, plan_days)
        product_transport = {
            **product_transport,
            "shipments": product_shipments,
            "freight_cost": sum(float(row.get("freight_cost") or 0.0) for row in product_shipments),
            "shortage": product_replay["shortage"],
            "failures": product_replay["risks"],
        }
        material_transports = [
            _solve_material_procurement_milp(
                material=item,
                production_daily=production_daily,
                products=products,
                suppliers=suppliers,
                routes=routes,
                rates=rates,
                carriers=carriers,
                days=plan_days,
            )
            for item in factory_materials
        ]
        material_shipments = [row for transport in material_transports for row in transport["shipments"]]
        product_shipments = product_transport["shipments"]
        all_shipments = material_shipments + product_shipments

        procurement = procurement_summary_rows(material_shipments, "电视综合专用逐日原料补货")
        total_purchase = sum(row["purchase_cost"] for row in procurement)
        total_material_freight = sum(row["freight_cost"] for row in procurement)
        total_product_freight = product_transport["freight_cost"]
        material_shortage = sum(transport.get("shortage", 0.0) for transport in material_transports)
        product_shortage = product_transport.get("shortage", 0.0)
        production_actual = max(0.0, production_target - material_shortage / max(len(factory_materials), 1))
        production_satisfaction = min(1.0, production_actual / max(production_target, 0.001)) if production_target else 1.0
        market_satisfaction = min(1.0, (forecast_total - product_shortage) / max(forecast_total, 0.001)) if forecast_total else 1.0
        unit_procurement = (total_purchase + total_material_freight) / max(production_actual, 0.001)
        unit_logistics = total_product_freight / max(product_replay["served"], 0.001)
        points = score_points(xls_path, "综合")
        score_rows = [
            {"item": "预测偏差率", "target": targets["prediction_deviation"], "actual": prediction_deviation, "points": deviation_score(prediction_deviation, targets["prediction_deviation"], points["预测偏差率"]), "max": points["预测偏差率"]},
            {"item": "单位物流成本", "target": targets["unit_logistics_cost"], "actual": unit_logistics, "points": cost_score(unit_logistics, targets["unit_logistics_cost"], points["单位物流成本"]), "max": points["单位物流成本"]},
            {"item": "单位采购成本", "target": targets["unit_procurement_cost"], "actual": unit_procurement, "points": cost_score(unit_procurement, targets["unit_procurement_cost"], points["单位采购成本"]), "max": points["单位采购成本"]},
            {"item": "生产满足率", "target": targets["production_satisfaction"], "actual": production_satisfaction, "points": satisfaction_score(production_satisfaction, points["生产满足率"]), "max": points["生产满足率"]},
            {"item": "市场满足率", "target": targets["market_satisfaction"], "actual": market_satisfaction, "points": satisfaction_score(market_satisfaction, points["市场满足率"]), "max": points["市场满足率"]},
        ]
        product_hard_risks = product_transport_hard_risks(
            forecasts=forecasts,
            shipments=product_shipments,
            factories=factories,
            days=plan_days,
            production_rows=production_rows,
        )
        failures = []
        for transport in material_transports:
            failures.extend(transport.get("failures", []))
        failures.extend(product_transport.get("failures", []))
        failures.extend(risk for risk in product_hard_risks if risk not in failures)
        product_transport_rows = [
            {
                "destination": row["destination"],
                "factory": row.get("source") or row.get("factory"),
                "cargo": row["cargo"],
                "amount": row["amount"],
                "route": row["route"],
                "mode": row["mode"],
                "lead": row["lead"],
                "ship_day": row["ship_day"],
                "arrival_day": row["arrival_day"],
                "freight_cost": row["freight_cost"],
                "carrier": row.get("carrier", ""),
                "note": "电视综合专用逐日成品补货",
            }
            for row in product_shipments
        ]
        simulation = simulate_plan(
            sales_fc=forecasts,
            factories=factories,
            factory_materials=factory_materials,
            production_rows=production_rows,
            procurement_rows=material_shipments,
            product_transport_rows=product_transport_rows,
            products=products,
            days=plan_days,
        )
        if failures:
            simulation["risks"] = list(dict.fromkeys(list(simulation.get("risks", [])) + failures))
            simulation["ok"] = False
            simulation["summary"] = "逐日复核存在缺口；未找到满满足率候选时按满足率优先保留"
        return {
            "qtype": "综合",
            "solver_name": "电视综合专用参数化求解器",
            "case_keyword": case_keyword(xls_path),
            "products": products,
            "sales_fc": forecasts,
            "production": production_rows,
            "procurement": procurement,
            "material_transport": material_shipments,
            "product_transport": product_transport_rows,
            "transport_sections": _compact_transport_rows(all_shipments),
            "score_rows": score_rows,
            "score": sum(row["points"] for row in score_rows),
            "score_note": "平台公式预测；预测偏差率需提交后由平台复原，不再冒充真实分",
            "unit_procurement": unit_procurement,
            "unit_logistics": unit_logistics,
            "production_satisfaction": production_satisfaction,
            "market_satisfaction": market_satisfaction,
            "simulation": simulation,
            "assumptions": ["电视专用模型按当前 xls 数字重新求解", "运输数量为整数", "路线/趟次惩罚为 0", f"{solver_mode_label()}按满足率100%优先选择候选"],
            "plan_days": plan_days,
            "solver_status": {
                "product_transport": product_transport["status"],
                "material_transport": {transport["shipments"][0]["cargo"] if transport["shipments"] else "未知": transport["status"] for transport in material_transports},
            },
        }

    cap = ceil_int(factory.daily * plan_days)
    full_net = ceil_int(max(0.0, forecast_total - sales_init_total))
    old_net = ceil_int(max(0.0, forecast_total - sales_init_total - factory.init))
    candidate_targets = {old_net, full_net, ceil_int(full_net * 0.95)}
    for delta in (-250, 0, 250):
        candidate_targets.add(old_net + delta)
    valid_targets = sorted({min(cap, max(0, target)) for target in candidate_targets if target > 0})
    candidates = [build_candidate(target) for target in valid_targets]
    candidate_summaries = [
        {
            "production_target": int(row["production"][0]["amount"]) if row.get("production") else 0,
            "score": float(row.get("score", 0.0) or 0.0),
            "unit_logistics": float(row.get("unit_logistics", 0.0) or 0.0),
            "unit_procurement": float(row.get("unit_procurement", 0.0) or 0.0),
            "market_satisfaction": float(row.get("market_satisfaction", 0.0) or 0.0),
            "production_satisfaction": float(row.get("production_satisfaction", 0.0) or 0.0),
            "product_freight": sum(float(item.get("freight_cost") or 0.0) for item in row.get("product_transport", [])),
            "material_freight": sum(float(item.get("freight_cost") or 0.0) for item in row.get("material_transport", [])),
            "purchase_cost": sum(float(item.get("purchase_cost") or 0.0) for item in row.get("procurement", [])),
            "product_transport_status": (row.get("solver_status") or {}).get("product_transport", ""),
            "material_transport_status": (row.get("solver_status") or {}).get("material_transport", {}),
            "risk_count": len((row.get("simulation") or {}).get("risks", [])),
            "route_count": len(row.get("product_transport", [])) + len(row.get("material_transport", [])),
        }
        for row in candidates
    ]
    feasible_candidates = [row for row in candidates if not (row.get("simulation") or {}).get("risks")]
    selectable_candidates = feasible_candidates or candidates
    best = max(
        selectable_candidates,
        key=lambda row: result_satisfaction_cost_priority(
            row,
            targets,
            len(row.get("product_transport", [])) + len(row.get("material_transport", [])),
        ),
    )
    best["candidate_targets"] = valid_targets
    best["candidate_summaries"] = sorted(
        candidate_summaries,
        key=lambda row: result_satisfaction_cost_priority(row, targets, int(row.get("route_count", 0))),
        reverse=True,
    )
    best["assumptions"].append(f"候选生产目标 {len(valid_targets)} 个，选中 {qty(best['production'][0]['amount'])}")
    return best


def solve_battery_comprehensive(sections: dict[str, list[tuple[Any, ...]]], xls_path: Path) -> dict[str, Any]:
    plan_days = plan_days_for_case(xls_path, "综合")
    products = parse_products(sections)
    factories = parse_factories(sections)
    factory_materials = parse_factory_materials(sections)
    suppliers = parse_suppliers(sections)
    routes = parse_routes(sections)
    rates = parse_rates(sections)
    carriers = parse_carriers_safe(sections, rates)
    sales = parse_sales(sections, plan_days)
    forecasts = [forecast_node(node, plan_days, use_bias=forecast_bias_for_case(xls_path, "综合")) for node in sales]
    if not factories or not forecasts:
        return solve_comprehensive(sections, xls_path)

    product_name = products[0].name if products else factories[0].product
    ratio = charge_ratio(products, product_name)
    finished_goods_capacity_remaining = {
        factory.name: max(0, int(math.floor(float(factory.init or 0.0) + float(factory.daily or 0.0) * plan_days)))
        for factory in factories
    }

    ranked_by_node: dict[str, list[tuple[float, Factory, Route]]] = {}
    for forecast in forecasts:
        net_need = ceil_int(max(0.0, float(forecast["forecast"]) - float(forecast["init"])))
        ranked: list[tuple[float, Factory, Route]] = []
        for factory in factories:
            route = pick_bulk_route(routes, factory.name, forecast["node"], max(net_need, 1), ratio, urgent=False, carriers=carriers)
            if route:
                reference = max(net_need, 1.0)
                ranked.append((route_unit_cost(route, reference, ratio, carriers), factory, route))
        ranked.sort(key=lambda item: (item[0], route_effective_lead(item[2], max(net_need, 1.0), ratio, carriers)))
        ranked_by_node[forecast["node"]] = ranked

    def regret(row: dict[str, Any]) -> float:
        ranked = ranked_by_node.get(row["node"], [])
        if len(ranked) >= 2:
            return ranked[1][0] - ranked[0][0]
        return ranked[0][0] if ranked else 0.0

    allocations: list[tuple[dict[str, Any], Factory, int, float, float]] = []
    assignment: dict[str, Factory] = {}
    for forecast in sorted(forecasts, key=regret, reverse=True):
        remaining = ceil_int(max(0.0, float(forecast["forecast"]) - float(forecast["init"])))
        if remaining <= 0:
            continue
        ranked = ranked_by_node.get(forecast["node"], [])
        first = True
        for _unit, factory, _route in ranked:
            if remaining <= 0:
                break
            available = finished_goods_capacity_remaining.get(factory.name, 0)
            if available <= 0:
                continue
            amount = min(remaining, available)
            init_share = float(forecast.get("init") or 0.0) if first else 0.0
            allocations.append((forecast, factory, int(amount), init_share, amount / max(float(forecast["forecast"]), 1.0)))
            assignment.setdefault(forecast["node"], factory)
            finished_goods_capacity_remaining[factory.name] = available - int(amount)
            remaining -= int(amount)
            first = False
        if remaining > 0 and ranked:
            _unit, factory, _route = ranked[0]
            allocations.append((forecast, factory, int(remaining), 0.0, remaining / max(float(forecast["forecast"]), 1.0)))
            assignment.setdefault(forecast["node"], factory)

    product_transport = build_fast_product_transport(
        allocations=allocations,
        routes=routes,
        products=products,
        cargo=product_name,
        carriers=carriers,
        days=plan_days,
    )
    product_transport = repair_factory_ship_days(product_transport, factories, plan_days)
    product_transport = fill_market_shortages_with_fast_routes(
        forecasts=forecasts,
        shipments=product_transport,
        assignment=assignment,
        routes=routes,
        products=products,
        cargo=product_name,
        carriers=carriers,
        factories=factories,
        days=plan_days,
        max_rounds=3,
    )
    product_transport = sanitize_product_transport(
        forecasts=forecasts,
        shipments=product_transport,
        factories=factories,
        routes=routes,
        products=products,
        cargo=product_name,
        carriers=carriers,
        days=plan_days,
    )
    production_rows, product_transport = align_production_to_transport(
        forecasts=forecasts,
        product_transport=product_transport,
        factories=factories,
        routes=routes,
        products=products,
        cargo=product_name,
        carriers=carriers,
        days=plan_days,
    )
    product_replay = daily_market_replay(forecasts, product_transport, plan_days)

    material_plan = material_procurement_for_production(
        production_rows=production_rows,
        factories=factories,
        factory_materials=factory_materials,
        products=products,
        suppliers=suppliers,
        routes=routes,
        rates=rates,
        carriers=carriers,
        days=plan_days,
        name_prefix=f"{case_keyword(xls_path)}_battery_material",
        note="蓄电池专用逐日原料补货",
    )
    material_transports = material_plan["shipments"]
    material_failures = material_plan["failures"]
    material_status = material_plan["status"]
    procurement = material_plan["procurement"]

    total_purchase = sum(float(row.get("purchase_cost") or 0.0) for row in procurement)
    total_material_freight = sum(float(row.get("freight_cost") or 0.0) for row in procurement)
    total_product_freight = sum(float(row.get("freight_cost") or 0.0) for row in product_transport)
    actual_production = float(material_plan["actual_production"])
    production_satisfaction = float(material_plan["production_satisfaction"])
    market_satisfaction = product_replay["market_satisfaction"]
    unit_procurement = (total_purchase + total_material_freight) / max(actual_production, 0.001)
    unit_logistics = total_product_freight / max(product_replay["served"], 0.001)

    targets = score_targets(xls_path, "综合")
    points = score_points(xls_path, "综合")
    prediction_deviation = 0.0
    score_rows = [
        {"item": "预测偏差率", "target": targets["prediction_deviation"], "actual": prediction_deviation, "points": deviation_score(prediction_deviation, targets["prediction_deviation"], points["预测偏差率"]), "max": points["预测偏差率"]},
        {"item": "单位物流成本", "target": targets["unit_logistics_cost"], "actual": unit_logistics, "points": cost_score(unit_logistics, targets["unit_logistics_cost"], points["单位物流成本"]), "max": points["单位物流成本"]},
        {"item": "单位采购成本", "target": targets["unit_procurement_cost"], "actual": unit_procurement, "points": cost_score(unit_procurement, targets["unit_procurement_cost"], points["单位采购成本"]), "max": points["单位采购成本"]},
        {"item": "生产满足率", "target": targets["production_satisfaction"], "actual": production_satisfaction, "points": satisfaction_score(production_satisfaction, points["生产满足率"]), "max": points["生产满足率"]},
        {"item": "市场满足率", "target": targets["market_satisfaction"], "actual": market_satisfaction, "points": satisfaction_score(market_satisfaction, points["市场满足率"]), "max": points["市场满足率"]},
    ]
    product_hard_risks = product_transport_hard_risks(
        forecasts=forecasts,
        shipments=product_transport,
        factories=factories,
        days=plan_days,
        production_rows=production_rows,
    )
    risks = material_failures + product_replay["risks"]
    risks.extend(risk for risk in product_hard_risks if risk not in risks)
    simulation = simulate_plan(
        sales_fc=forecasts,
        factories=factories,
        factory_materials=factory_materials,
        production_rows=production_rows,
        procurement_rows=material_transports,
        product_transport_rows=product_transport,
        products=products,
        days=plan_days,
    )
    if risks:
        simulation["risks"] = list(dict.fromkeys(list(simulation.get("risks", [])) + risks))
        simulation["ok"] = False
        simulation["summary"] = "逐日复核存在缺口；未找到满满足率候选时按满足率优先保留"
    return {
        "qtype": "综合",
        "solver_name": "蓄电池综合专用参数化求解器",
        "products": products,
        "sales_fc": forecasts,
        "production": production_rows,
        "procurement": procurement,
        "material_transport": material_transports,
        "product_transport": product_transport,
        "transport_sections": _compact_transport_rows(material_transports + product_transport),
        "score_rows": score_rows,
        "score": sum(row["points"] for row in score_rows),
        "score_note": "蓄电池专用公式候选分：单位物流成本按成品配送口径，原料运费进入单位采购成本；预测偏差率提交后复原。",
        "unit_procurement": unit_procurement,
        "unit_logistics": unit_logistics,
        "production_satisfaction": production_satisfaction,
        "market_satisfaction": market_satisfaction,
        "simulation": simulation,
        "market_replay": product_replay,
        "plan_days": plan_days,
        "solver_status": {"product_transport": "BatteryFastBulk", "material_transport": material_status},
        "assumptions": ["蓄电池专用模型按当前 xls 数字重新分配工厂与路线", "路线/趟次惩罚为 0", "成品优先使用铁路/海运大批量路线，必要时用直达路线桥接早期缺口"],
    }


def solve_comprehensive(sections: dict[str, list[tuple[Any, ...]]], xls_path: Path) -> dict[str, Any]:
    plan_days = plan_days_for_case(xls_path, "综合")
    products = parse_products(sections)
    factories = parse_factories(sections)
    factory_materials = parse_factory_materials(sections)
    suppliers = parse_suppliers(sections)
    routes = parse_routes(sections)
    rates = parse_rates(sections)
    carriers = parse_carriers_safe(sections, rates)
    sales = parse_sales(sections, plan_days)
    use_bias = forecast_bias_for_case(xls_path, "综合")
    forecasts = [forecast_node(node, plan_days, use_bias=use_bias) for node in sales]
    product_name = products[0].name if products else (factories[0].product if factories else "")
    product_flow = solve_sales_or_production(sections, xls_path, "销售")
    product_transport = product_flow.get("product_transport", [])
    production_rows, product_transport = align_production_to_transport(
        forecasts=forecasts,
        product_transport=product_transport,
        factories=factories,
        routes=routes,
        products=products,
        cargo=product_name,
        carriers=carriers,
        days=plan_days,
    )
    market_replay = product_flow.get("market_replay") or daily_market_replay(forecasts, product_transport, plan_days)
    market_replay = daily_market_replay(forecasts, product_transport, plan_days)
    market_satisfaction = market_replay["market_satisfaction"] if forecasts else 1.0

    material_plan = material_procurement_for_production(
        production_rows=production_rows,
        factories=factories,
        factory_materials=factory_materials,
        products=products,
        suppliers=suppliers,
        routes=routes,
        rates=rates,
        carriers=carriers,
        days=plan_days,
        name_prefix=f"{case_keyword(xls_path)}_comprehensive_material",
        note="综合通用逐日原料补货",
    )
    procurement = material_plan["procurement"]
    material_transports = material_plan["shipments"]
    total_purchase = float(material_plan["purchase_cost"])
    total_freight = float(material_plan["freight_cost"])
    actual_production = float(material_plan["actual_production"])
    total_product_ship = sum(row["amount"] for row in product_transport)
    total_product_freight = sum(row["freight_cost"] for row in product_transport)
    unit_procurement = (total_purchase + total_freight) / max(actual_production, 0.001)
    unit_logistics = total_product_freight / max(market_replay["served"] if forecasts else total_product_ship, 0.001)
    production_satisfaction = float(material_plan["production_satisfaction"])
    prediction_deviation = 0.0

    targets = score_targets(xls_path, "综合")
    points = score_points(xls_path, "综合")
    score_rows = [
        {
            "item": "预测偏差率",
            "target": targets["prediction_deviation"],
            "actual": prediction_deviation,
            "points": deviation_score(prediction_deviation, targets["prediction_deviation"], points["预测偏差率"]),
            "max": points["预测偏差率"],
        },
        {
            "item": "单位物流成本",
            "target": targets["unit_logistics_cost"],
            "actual": unit_logistics,
            "points": cost_score(unit_logistics, targets["unit_logistics_cost"], points["单位物流成本"]),
            "max": points["单位物流成本"],
        },
        {
            "item": "单位采购成本",
            "target": targets["unit_procurement_cost"],
            "actual": unit_procurement,
            "points": cost_score(unit_procurement, targets["unit_procurement_cost"], points["单位采购成本"]),
            "max": points["单位采购成本"],
        },
        {
            "item": "生产满足率",
            "target": targets["production_satisfaction"],
            "actual": production_satisfaction,
            "points": satisfaction_score(production_satisfaction, points["生产满足率"]),
            "max": points["生产满足率"],
        },
        {
            "item": "市场满足率",
            "target": targets["market_satisfaction"],
            "actual": market_satisfaction,
            "points": satisfaction_score(market_satisfaction, points["市场满足率"]),
            "max": points["市场满足率"],
        },
    ]
    simulation = simulate_plan(
        sales_fc=forecasts,
        factories=factories,
        factory_materials=factory_materials,
        production_rows=production_rows,
        procurement_rows=material_transports,
        product_transport_rows=product_transport,
        products=products,
        days=plan_days,
    )
    extra_risks = list(material_plan.get("failures", []))
    if extra_risks:
        simulation["risks"] = list(dict.fromkeys(list(simulation.get("risks", [])) + extra_risks))
        simulation["ok"] = False
        simulation["summary"] = "逐日复核存在缺口；未找到满满足率候选时按满足率优先保留"
    return {
        "qtype": "综合",
        "products": products,
        "sales_fc": forecasts,
        "production": production_rows,
        "procurement": procurement,
        "material_transport": material_transports,
        "product_transport": product_transport,
        "transport_sections": _compact_transport_rows(material_transports + product_transport),
        "score_rows": score_rows,
        "score": sum(row["points"] for row in score_rows),
        "score_note": "公式估算：综合评分按已知指标计算；预测偏差率按提交预测量为 0 估算，平台回填真实销量后再复原",
        "unit_procurement": unit_procurement,
        "unit_logistics": unit_logistics,
        "production_satisfaction": production_satisfaction,
        "market_satisfaction": market_satisfaction,
        "simulation": simulation,
        "market_replay": market_replay,
        "plan_days": plan_days,
        "solver_status": {
            "product_transport": product_flow.get("solver_status", {}),
            "material_transport": material_plan.get("status", {}),
        },
        "assumptions": ["综合题串联销售预测、生产、采购、原料运输、成品运输", "成品运输复用逐日整数运输工作流，路线/趟次惩罚为 0"],
    }


def solve_file(xls_path: Path) -> dict[str, Any]:
    sections = read_workbook(xls_path)
    qtype = detect_type(sections, xls_path)
    keyword = case_keyword(xls_path)
    log(f"{xls_path.name}: 题型={qtype} | 案例={keyword}")
    if qtype == "采购":
        result = solve_procurement(sections, xls_path)
    elif qtype == "综合" and keyword == "电视":
        result = solve_tv_comprehensive(sections, xls_path)
    elif qtype == "综合" and "蓄电池" in keyword:
        result = solve_battery_comprehensive(sections, xls_path)
    elif qtype == "综合":
        result = solve_comprehensive(sections, xls_path)
    elif qtype == "销售":
        result = solve_sales_or_production(sections, xls_path, "销售")
    else:
        result = solve_sales_or_production(sections, xls_path, "生产")
    result.setdefault("case_keyword", keyword)
    result["solver_mode"] = solver_mode_label()
    result["xls_path"] = str(xls_path)
    result["title"] = xls_path.stem
    if ACTIVE_HAR_CONTEXT:
        result.setdefault("assumptions", []).append(
            f"平台 HAR 已接入：{Path(str(ACTIVE_HAR_CONTEXT.get('path', ''))).name}；承运商 {len(ACTIVE_HAR_CONTEXT.get('carriers') or [])} 个"
        )
    apply_verified_score_override(result, xls_path)
    return result


def fmt_ratio(value: float) -> str:
    return f"{value * 100:.2f}%"


def render_rows(headers: list[str], rows: list[list[Any]], numeric: set[int] | None = None) -> str:
    if not rows:
        return ""
    numeric = numeric or set()
    out = ["<table><thead><tr>"]
    out.extend(f"<th>{html.escape(header)}</th>" for header in headers)
    out.append("</tr></thead><tbody>")
    for row in rows:
        out.append("<tr>")
        for idx, cell in enumerate(row):
            cls = "num" if idx in numeric else ""
            out.append(f'<td class="{cls}">{html.escape(str(cell))}</td>')
        out.append("</tr>")
    out.append("</tbody></table>")
    return "\n".join(out)


def fmt_day(value: Any) -> str:
    try:
        day = int(value)
        month = 1
        for days_in_month in (31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31):
            if day <= days_in_month:
                return f"{month}-{day:02d}"
            day -= days_in_month
            month += 1
        return f"{month}-{day:02d}"
    except (TypeError, ValueError):
        return html.escape(str(value))


def production_plan_rows(production_rows: list[dict[str, Any]], plan_days: int) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in production_rows:
        amount = ceil_int(row.get("amount", 0))
        if amount <= 0:
            continue
        capacity = float(row.get("capacity") or 0.0)
        daily_capacity = int(math.floor(capacity / max(plan_days, 1))) if capacity > 0 else 0
        if daily_capacity <= 0:
            daily_capacity = max(1, ceil_int(amount / max(plan_days, 1)))
        day = 1
        remaining = amount
        while remaining > 0 and day <= plan_days:
            if remaining >= daily_capacity:
                run_days = min(remaining // daily_capacity, plan_days - day + 1)
                daily_amount = daily_capacity
            else:
                run_days = 1
                daily_amount = remaining
            end_day = day + int(run_days) - 1
            rows.append([row.get("factory", ""), fmt_day(day), fmt_day(end_day), int(daily_amount)])
            remaining -= int(run_days) * int(daily_amount)
            day = end_day + 1
        if remaining > 0:
            rows.append([row.get("factory", ""), fmt_day(plan_days), fmt_day(plan_days), int(remaining)])
    return rows


def carrier_class(carrier: str) -> str:
    if "时代物流" in carrier:
        return "carrier-shidai"
    if "中原物流" in carrier:
        return "carrier-zhongyuan"
    if "时新快运" in carrier:
        return "carrier-shixin"
    if "南方速运" in carrier:
        return "carrier-nanfang"
    if "顺风物流" in carrier:
        return "carrier-shunfeng"
    if "国际快运" in carrier or "国际物流" in carrier or "国际" in carrier:
        return "carrier-guoji"
    if "铁路" in carrier or "铁运" in carrier or "西铁" in carrier or "中铁" in carrier:
        return "carrier-rail"
    if "海运" in carrier:
        return "carrier-sea"
    if "易达快运" in carrier:
        return "carrier-yida"
    if "开源陆运" in carrier:
        return "carrier-kaiyuan"
    return "carrier-other"


def split_route_segments(route: str, carrier: str, start_day: int, lead: int = 0) -> list[tuple[str, str, int | None]]:
    points = [part.strip() for part in str(route).split("-->") if part.strip()]
    if len(points) < 2:
        return [(route, carrier, start_day)]
    segments = [f"{src}-->{dst}" for src, dst in zip(points, points[1:])]
    carriers = [part.strip() for part in str(carrier).split("+") if part.strip()]
    if len(carriers) != len(segments):
        carriers = [carrier] * len(segments)
    durations = [verified_segment_duration(segment, segment_carrier) for segment, segment_carrier in zip(segments, carriers)]
    use_known_offsets = all(duration is not None for duration in durations)
    rows = []
    offset = 0
    for idx, segment in enumerate(segments):
        segment_start = start_day + offset if use_known_offsets or idx == 0 else None
        rows.append((segment, carriers[idx], segment_start))
        if use_known_offsets:
            offset += int(durations[idx] or 0)
    return rows


def transport_fill_table(rows: list[list[Any]]) -> str:
    if not rows:
        return ""
    headers = ["序号", "路线", "承运商", "运输数量", "起运日期", "承运趟数", "几天一趟"]
    out = ["<table class=\"transport\"><thead><tr>"]
    out.extend(f"<th>{html.escape(header)}</th>" for header in headers)
    out.append("</tr></thead><tbody>")
    expanded_rows: list[list[Any]] = []
    base_rows = sorted(rows, key=lambda row: (int(row[3]), str(row[0]), str(row[1]), int(row[2])))
    for row in base_rows:
        route, carrier, amount, start_day, trips, interval = row[:6]
        lead = int(row[6]) if len(row) > 6 else 0
        for segment, segment_carrier, segment_start in split_route_segments(str(route), str(carrier), int(start_day), lead):
            if segment_start is not None:
                date_text = fmt_day(segment_start)
            else:
                date_text = ""
            expanded_rows.append([segment, segment_carrier, int(amount), int(segment_start or start_day), date_text, int(trips), int(interval)])
    for seq, row in enumerate(expanded_rows, start=1):
        route, carrier, amount, _start_day, date_text, trips, interval = row
        interval_text = "-" if int(trips) <= 1 else f"{int(interval)}天"
        cells = [seq, route, carrier, int(amount), date_text, int(trips), interval_text]
        out.append("<tr>")
        for idx, cell in enumerate(cells):
            cls = "num" if idx in {0, 3, 5, 6} else ""
            if idx == 2:
                cls = f"carrier {carrier_class(str(cell))}"
            out.append(f'<td class="{cls}">{html.escape(str(cell))}</td>')
        out.append("</tr>")
    out.append("</tbody></table>")
    return "\n".join(out)


def section(title: str, content: str) -> str:
    if not content:
        return ""
    return f"<section><h2>{html.escape(title)}</h2>{content}</section>"


def aggregate_transport_sections(result: dict[str, Any]) -> dict[str, list[list[Any]]]:
    sections: dict[str, list[list[Any]]] = {}
    for row in result.get("material_transport", []):
        route = row.get("route", "")
        if not route or "缺口" in str(route) or "不足" in str(route):
            continue
        cargo = row.get("material") or row.get("cargo") or "原料"
        sections.setdefault(str(cargo), []).append(
            [
                route,
                row.get("carrier") or row.get("mode") or "承运商待确认",
                ceil_int(row.get("amount", 0)),
                int(row.get("ship_day") or 1),
                int(row.get("trips") or 1),
                int(row.get("interval") or 1),
                int(row.get("lead") or 0),
            ]
        )
    for row in result.get("product_transport", []):
        route = row.get("route", "")
        if not route or "缺口" in str(route) or "不足" in str(route):
            continue
        cargo = row.get("cargo") or "成品"
        sections.setdefault(str(cargo), []).append(
            [
                route,
                row.get("carrier") or row.get("mode") or "承运商待确认",
                ceil_int(row.get("amount", 0)),
                int(row.get("ship_day") or 1),
                int(row.get("trips") or 1),
                int(row.get("interval") or 1),
                int(row.get("lead") or 0),
            ]
        )
    for rows in sections.values():
        rows.sort(key=lambda row: (int(row[3]), str(row[0]), str(row[1]), int(row[2])))
    return sections


def concise_solver_status(status: Any) -> str:
    if not isinstance(status, dict):
        return str(status)
    parts: list[str] = []

    def append_model_stats(bits: list[str], source: dict[str, Any]) -> None:
        stats = source.get("model_stats")
        if not isinstance(stats, dict):
            return
        stat_parts = []
        for label, key in (
            ("lane", "lane_options"),
            ("x", "shipment_vars"),
            ("bin", "binary_vars"),
            ("optwin", "allowed_ship_day_option_windows"),
            ("routewin", "allowed_ship_day_windows"),
        ):
            if stats.get(key) is not None:
                stat_parts.append(f"{label}{int(stats.get(key) or 0)}")
        if source.get("time_limit_hit"):
            stat_parts.append("timelimit")
        if stat_parts:
            bits.append("模型" + "/".join(stat_parts))

    selected = status.get("selected")
    method = status.get("method")
    raw_status = status.get("status")
    if selected:
        parts.append(f"选中：{selected}")
    if method:
        parts.append(f"方法：{method}")
    if raw_status:
        parts.append(f"状态：{raw_status}")
    timing_source = status
    product_status = status.get("product_transport")
    if isinstance(product_status, dict):
        timing_source = product_status
    global_status = timing_source.get("global_status") if isinstance(timing_source, dict) else status.get("global_status")
    if isinstance(global_status, dict) and global_status.get("status"):
        parts.append(f"全局：{global_status.get('status')}")
    warm_status = timing_source.get("warm_window_global_status") if isinstance(timing_source, dict) else status.get("warm_window_global_status")
    if isinstance(warm_status, dict) and warm_status.get("status") and warm_status.get("status") != "Skipped":
        warm_bits = [f"暖窗：{warm_status.get('status')}"]
        if warm_status.get("warm_window_freight") is not None:
            warm_bits.append(f"候选{float(warm_status.get('warm_window_freight') or 0.0):,.0f}")
        if warm_status.get("warm_base_freight") is not None:
            warm_bits.append(f"基准{float(warm_status.get('warm_base_freight') or 0.0):,.0f}")
        if warm_status.get("strict_freight_upper_bound") is not None:
            warm_bits.append(f"反例<{float(warm_status.get('strict_freight_upper_bound') or 0.0):,.0f}")
        if warm_status.get("solve_seconds") is not None:
            warm_bits.append(f"{float(warm_status.get('solve_seconds') or 0.0):.1f}s")
        append_model_stats(warm_bits, warm_status)
        parts.append(" ".join(warm_bits))
    full_audit_status = timing_source.get("full_global_audit_status") if isinstance(timing_source, dict) else status.get("full_global_audit_status")
    if isinstance(full_audit_status, dict) and full_audit_status.get("status") and full_audit_status.get("status") != "Skipped":
        audit_bits = [f"全路线审查：{full_audit_status.get('status')}"]
        if full_audit_status.get("full_global_freight") is not None:
            audit_bits.append(f"候选{float(full_audit_status.get('full_global_freight') or 0.0):,.0f}")
        if full_audit_status.get("full_global_base_freight") is not None:
            audit_bits.append(f"基准{float(full_audit_status.get('full_global_base_freight') or 0.0):,.0f}")
        if full_audit_status.get("solve_seconds") is not None:
            audit_bits.append(f"{float(full_audit_status.get('solve_seconds') or 0.0):.1f}s")
        append_model_stats(audit_bits, full_audit_status)
        parts.append(" ".join(audit_bits))
    pattern_status = timing_source.get("destination_pattern_polish") if isinstance(timing_source, dict) else None
    final_fixed_status = timing_source.get("final_fixed_arrival_reroute_status") if isinstance(timing_source, dict) else None
    extreme_audit = timing_source.get("extreme_post_polish_status") if isinstance(timing_source, dict) else None
    if (isinstance(final_fixed_status, dict) and final_fixed_status.get("status")) or (isinstance(extreme_audit, dict) and extreme_audit.get("status")):
        fixed_status = final_fixed_status if isinstance(final_fixed_status, dict) and final_fixed_status.get("status") != "Skipped" else None
        if fixed_status is None and isinstance(extreme_audit, dict):
            fixed_status = extreme_audit.get("fixed_arrival_reroute")
        if isinstance(fixed_status, dict) and fixed_status.get("status"):
            fixed_bits = [f"固定到货重路由：{fixed_status.get('status')}"]
            if fixed_status.get("delta") is not None:
                fixed_bits.append(f"Δ{float(fixed_status.get('delta') or 0.0):,.0f}")
            if fixed_status.get("freight_cost") is not None:
                fixed_bits.append(f"成本{float(fixed_status.get('freight_cost') or 0.0):,.0f}")
            if fixed_status.get("options") is not None:
                fixed_bits.append(f"opt{int(fixed_status.get('options') or 0)}")
            if fixed_status.get("solve_seconds") is not None:
                fixed_bits.append(f"{float(fixed_status.get('solve_seconds') or 0.0):.1f}s")
            parts.append(" ".join(fixed_bits))
        pattern_status = extreme_audit.get("destination_pattern_polish") or pattern_status
    subset_status = timing_source.get("destination_subset_status") if isinstance(timing_source, dict) else None
    if isinstance(subset_status, dict) and subset_status.get("status") and subset_status.get("status") != "Skipped":
        subset_bits = [f"北方武汉子集：{subset_status.get('status')}"]
        if subset_status.get("delta") is not None:
            subset_bits.append(f"Δ{float(subset_status.get('delta') or 0.0):,.0f}")
        if subset_status.get("freight_cost") is not None:
            subset_bits.append(f"成本{float(subset_status.get('freight_cost') or 0.0):,.0f}")
        if subset_status.get("solve_seconds") is not None:
            subset_bits.append(f"{float(subset_status.get('solve_seconds') or 0.0):.1f}s")
        append_model_stats(subset_bits, subset_status)
        parts.append(" ".join(subset_bits))
    delay_status = timing_source.get("final_delay_reroute_status") if isinstance(timing_source, dict) else None
    if isinstance(delay_status, dict) and delay_status.get("status") and delay_status.get("status") != "Skipped":
        delay_bits = [f"延后低成本路线：{delay_status.get('status', 'Computed')}"]
        if delay_status.get("delta") is not None:
            delay_bits.append(f"Δ{float(delay_status.get('delta') or 0.0):,.0f}")
        if delay_status.get("checked") is not None:
            delay_bits.append(f"chk{int(delay_status.get('checked') or 0)}")
        parts.append(" ".join(delay_bits))
    if isinstance(pattern_status, dict) and pattern_status:
        pattern_label = pattern_status.get("status") or (pattern_status.get("pattern_milp") or {}).get("status") or "已记录"
        pattern_bits = [f"模式组合：{pattern_label}"]
        if pattern_status.get("delta") is not None:
            pattern_bits.append(f"Δ{float(pattern_status.get('delta') or 0.0):,.0f}")
        if pattern_status.get("checked") is not None:
            pattern_bits.append(f"chk{int(pattern_status.get('checked') or 0)}")
        if pattern_status.get("combinations") is not None:
            pattern_bits.append(f"comb{int(pattern_status.get('combinations') or 0)}")
        if pattern_status.get("avoidance_patterns") is not None:
            pattern_bits.append(f"avoid{int(pattern_status.get('avoidance_patterns') or 0)}")
        if pattern_status.get("conflict_shadow_patterns") is not None:
            pattern_bits.append(f"conflict{int(pattern_status.get('conflict_shadow_patterns') or 0)}")
        pattern_milp = pattern_status.get("pattern_milp")
        if isinstance(pattern_milp, dict) and pattern_milp.get("status"):
            pattern_bits.append(f"MILP:{pattern_milp.get('status')}")
        parts.append(" ".join(pattern_bits))
    cycle_status = None
    if isinstance(extreme_audit, dict):
        cycle_status = extreme_audit.get("post_polish_cycles")
    if isinstance(cycle_status, dict) and cycle_status.get("steps"):
        slow_cycle_bits: list[str] = []
        cycle_delta_bits: list[str] = []
        for step in list(cycle_status.get("steps") or [])[:3]:
            if not isinstance(step, dict):
                continue
            cycle_delta_bits.append(f"cy{int(step.get('cycle') or 0)} Δ{float(step.get('delta') or 0.0):,.0f}")
            substeps = [
                ("shift", step.get("shift")),
                ("repolish", step.get("repolish")),
                ("single", step.get("single_replacement")),
            ]
            slowest = max(
                (item for item in substeps if isinstance(item[1], dict)),
                key=lambda item: float(item[1].get("seconds") or 0.0),
                default=None,
            )
            if slowest is not None:
                slow_cycle_bits.append(f"cy{int(step.get('cycle') or 0)}:{slowest[0]} {float(slowest[1].get('seconds') or 0.0):.1f}s")
        if cycle_delta_bits:
            parts.append("循环收益：" + ", ".join(cycle_delta_bits))
        shift_audit_bits = []
        for step in list(cycle_status.get("steps") or [])[:3]:
            if not isinstance(step, dict) or not isinstance(step.get("shift"), dict):
                continue
            shift = step["shift"]
            if shift.get("estimated_candidates") is not None or shift.get("checked_candidates") is not None:
                shift_audit_bits.append(
                    f"cy{int(step.get('cycle') or 0)} est{int(shift.get('estimated_candidates') or 0)}/chk{int(shift.get('checked_candidates') or 0)}"
                )
        if shift_audit_bits:
            parts.append("换量审查：" + ", ".join(shift_audit_bits))
        if slow_cycle_bits:
            parts.append("循环瓶颈：" + ", ".join(slow_cycle_bits))
    block_status = extreme_audit.get("destination_block_polish") if isinstance(extreme_audit, dict) else None
    if isinstance(block_status, dict) and block_status:
        block_bits = [f"块重排：Δ{float(block_status.get('delta') or 0.0):,.0f}"]
        if block_status.get("improvements") is not None:
            block_bits.append(f"imp{int(block_status.get('improvements') or 0)}")
        if block_status.get("blocks_checked") is not None:
            block_bits.append(f"blk{int(block_status.get('blocks_checked') or 0)}")
        if block_status.get("orders_checked") is not None:
            block_bits.append(f"ord{int(block_status.get('orders_checked') or 0)}")
        if block_status.get("option_cache_entries") is not None:
            block_bits.append(f"cache{int(block_status.get('option_cache_entries') or 0)}")
        parts.append(" ".join(block_bits))
    timings = list(timing_source.get("stage_timings") or []) if isinstance(timing_source, dict) else []
    timings.extend(list(timing_source.get("post_polish_timings") or []) if isinstance(timing_source, dict) else [])
    if timings:
        slowest = sorted(
            (item for item in timings if isinstance(item, dict)),
            key=lambda item: float(item.get("seconds") or 0.0),
            reverse=True,
        )[:5]
        if slowest:
            def timing_label(item: dict[str, Any]) -> str:
                label = f"{item.get('stage')} {float(item.get('seconds') or 0.0):.1f}s"
                extras = []
                if "delta" in item:
                    extras.append(f"Δ{float(item.get('delta') or 0.0):.1f}")
                if "improvements" in item:
                    extras.append(f"imp{int(item.get('improvements') or 0)}")
                if "cycles" in item:
                    extras.append(f"cy{int(item.get('cycles') or 0)}")
                if "rounds" in item:
                    extras.append(f"rd{int(item.get('rounds') or 0)}")
                if "checked" in item:
                    extras.append(f"chk{int(item.get('checked') or 0)}")
                return label + (f" ({' '.join(extras)})" if extras else "")

            parts.append(
                "耗时：" + "，".join(
                    timing_label(item)
                    for item in slowest
                )
            )
    lower_bound = timing_source.get("lower_bound_status") if isinstance(timing_source, dict) else None
    if isinstance(lower_bound, dict) and lower_bound.get("status") in {"Computed", "Partial"}:
        feasible = float(lower_bound.get("feasible_freight") or 0.0)
        bound = float(lower_bound.get("freight_lower_bound") or 0.0)
        gap_ratio = float(lower_bound.get("gap_ratio") or 0.0)
        if feasible > 0 and bound > 0:
            parts.append(f"下界：{bound:,.0f}，当前：{feasible:,.0f}，gap {gap_ratio * 100:.2f}%")
        diagnostics = lower_bound.get("gap_diagnostics")
        if isinstance(diagnostics, dict) and diagnostics.get("status") == "Computed":
            if diagnostics.get("independent_capacity_feasible"):
                parts.append("独立下界产能审查：可行")
            else:
                overflow = int(diagnostics.get("max_independent_capacity_overflow") or 0)
                node_deltas = diagnostics.get("node_deltas") or {}
                top_nodes = []
                if isinstance(node_deltas, dict):
                    sorted_nodes = sorted(
                        (
                            (node, float(info.get("delta") or 0.0))
                            for node, info in node_deltas.items()
                            if isinstance(info, dict)
                        ),
                        key=lambda item: item[1],
                        reverse=True,
                    )[:2]
                    top_nodes = [f"{node}+{delta:,.0f}" for node, delta in sorted_nodes if delta > 0]
                suffix = f"，主要差额：{', '.join(top_nodes)}" if top_nodes else ""
                parts.append(f"独立下界产能审查：不可行，最大累计超产{overflow:,.0f}{suffix}")
    hard_risks = status.get("hard_risks") or []
    if hard_risks:
        parts.append(f"硬风险候选已拒绝：{len(hard_risks)}项")
    if not parts and status:
        compact = []
        for key, value in status.items():
            if isinstance(value, dict):
                compact.append(f"{key}:{value.get('status') or value.get('method') or '已记录'}")
            else:
                compact.append(f"{key}:{value}")
            if len(compact) >= 3:
                break
        parts.extend(compact)
    return "；".join(parts) if parts else "已记录"


def render_html(result: dict[str, Any]) -> str:
    title = result["title"]
    qtype = result["qtype"]
    raw_score = result.get("score")
    score_known = isinstance(raw_score, (int, float))
    score = float(raw_score) if score_known else 0.0
    score_text = f"{score:.2f}" if score_known else "待正式指标"
    plan_days = int(result.get("plan_days") or PLAN_DAYS)
    mode_label = str(result.get("solver_mode") or solver_mode_label())
    simulation = result["simulation"]
    score_color = "#15803d" if score >= 85 else "#b45309" if score >= 60 else "#b91c1c"

    score_table = render_rows(
        ["评分项", "指标值", "实际/候选", "得分", "满分"],
        [
            [
                row["item"],
                fmt_ratio(row["target"]) if row["target"] <= 1.0 else money(row["target"]),
                fmt_ratio(row["actual"]) if row["target"] <= 1.0 else money(row["actual"]),
                f"{row['points']:.2f}",
                f"{row['max']:.0f}",
            ]
            for row in result["score_rows"]
        ],
        {1, 2, 3, 4},
    )
    production_table = render_rows(
        ["工厂", "计划日期起", "计划日期止", "每天产量"],
        production_plan_rows(result.get("production", []), plan_days),
        {3},
    )

    procurement_table = render_rows(
        ["原料", "供应商", "去向工厂", "采购量"],
        [
            [
                row["material"],
                row["supplier"],
                row["factory"],
                qty(row["amount"]),
            ]
            for row in result.get("procurement", [])
        ],
        {3},
    )
    sales_table = render_rows(
        ["销售网点", "预测销量"],
        [
            [
                row.get("node", ""),
                qty(row.get("forecast", 0)),
            ]
            for row in result.get("sales_fc", [])
        ],
        {1},
    )

    transport_sections = []
    compact_sections = result.get("transport_sections") or aggregate_transport_sections(result)
    if compact_sections:
        for cargo, rows in compact_sections.items():
            transport_sections.append(section(f"{cargo}运输填报表", transport_fill_table(rows)))

    risk_rows = [[risk] for risk in simulation["risks"]] or [["无硬风险"]]
    risk_table = render_rows(["校验结果"], risk_rows)
    status_table = render_rows(
        ["项目", "状态"],
        [
            ["结果状态", "校验通过" if simulation["ok"] else "存在满足率优先候选的风险项"],
            ["分数", score_text],
            ["计分口径", result.get("score_note", "公式估算")],
            ["求解版本", mode_label],
            ["求解状态", concise_solver_status(result.get("solver_status", "未记录"))],
        ],
    )
    if qtype == "销售":
        plan_sections_html = "".join(
            [
                section("销售量", sales_table),
            ]
        )
    else:
        plan_sections_html = "".join(
            [
                section("采购量", procurement_table),
                section("生产量", production_table),
            ]
        )

    solver_name = result.get("solver_name", f"{qtype}通用求解器")
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <link rel="icon" href="data:,">
  <title>{html.escape(title)} - 求解方案</title>
  <style>
    :root {{ color-scheme: light; }}
    body {{ margin: 0; padding: 24px; font-family: "Segoe UI", "Microsoft YaHei", Arial, sans-serif; background: #f6f7f8; color: #202124; }}
    main {{ max-width: 1180px; margin: 0 auto; }}
    header {{ background: #fff; border: 1px solid #d9dee3; border-radius: 8px; padding: 20px 24px; margin-bottom: 18px; }}
    h1 {{ font-size: 22px; margin: 0 0 8px; font-weight: 700; }}
    .meta {{ color: #5f6b76; font-size: 14px; }}
    .score-final strong {{ color: {score_color}; font-size: 32px; line-height: 1; }}
    .pill {{ border: 1px solid #cbd5e1; border-radius: 999px; padding: 6px 10px; font-size: 13px; color: #334155; background: #f8fafc; }}
    section {{ background: #fff; border: 1px solid #d9dee3; border-radius: 8px; padding: 18px 20px; margin: 14px 0; overflow-x: auto; }}
    h2 {{ margin: 0 0 12px; font-size: 17px; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
    th {{ background: #eef2f6; text-align: left; padding: 9px 10px; border-bottom: 1px solid #cbd5e1; white-space: nowrap; }}
    td {{ padding: 9px 10px; border-bottom: 1px solid #e5e7eb; vertical-align: top; }}
    tbody tr:nth-child(even) td {{ background: #f8fafc; }}
    td.num {{ text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }}
    tr:last-child td {{ border-bottom: 0; }}
    .note {{ color: #5f6b76; font-size: 13px; line-height: 1.6; }}
    .carrier {{ font-weight: 600; }}
    .carrier-shidai {{ background: #fee2e2 !important; color: #991b1b; }}
    .carrier-zhongyuan {{ background: #dbeafe !important; color: #1e3a8a; }}
    .carrier-shixin {{ background: #fef3c7 !important; color: #92400e; }}
    .carrier-nanfang {{ background: #dcfce7 !important; color: #166534; }}
    .carrier-shunfeng {{ background: #e0f2fe !important; color: #075985; }}
    .carrier-guoji {{ background: #ede9fe !important; color: #5b21b6; }}
    .carrier-rail {{ background: #fef3c7 !important; color: #92400e; }}
    .carrier-sea {{ background: #dcfce7 !important; color: #166534; }}
    .carrier-yida {{ background: #e0f2fe !important; color: #075985; }}
    .carrier-kaiyuan {{ background: #ede9fe !important; color: #5b21b6; }}
    .carrier-other {{ background: #e5e7eb !important; color: #374151; }}
    ul {{ margin: 8px 0 0; padding-left: 20px; }}
  </style>
</head>
<body>
<main>
  <header>
    <h1>{html.escape(title)}</h1>
    <div class="meta">题型：{html.escape(qtype)} | 版本：{html.escape(mode_label)} | 求解器：{html.escape(solver_name)} | 周期：{plan_days} 天 | 运输量：整数</div>
  </header>
  {plan_sections_html}
  {"".join(transport_sections)}
  <section><h2>结果状态与分数</h2><div class="score-final"><strong>{score_text}</strong></div>{status_table}{score_table}{risk_table}</section>
</main>
</body>
</html>
"""


def output_path_for(xls_path: Path, mode: str | None = None) -> Path:
    label = "极限版" if (mode or ACTIVE_SOLVER_MODE) == "extreme" else "高分版"
    return xls_path.with_name(f"{xls_path.stem}_{label}求解方案.html")


def write_solution(result: dict[str, Any]) -> Path:
    xls_path = Path(result["xls_path"])
    mode = "extreme" if str(result.get("solver_mode")) == "极限版" else "high"
    for suffix in ("_方案.html", "_启发式方案.html", "_验证方案.html", "_求解方案.html", f"_{result.get('solver_mode', solver_mode_label())}求解方案.html"):
        old_path = xls_path.with_name(f"{xls_path.stem}{suffix}")
        if old_path.exists():
            old_path.unlink()
    out_path = output_path_for(xls_path, mode)
    out_path.write_text(render_html(result), encoding="utf-8")
    return out_path


def self_test() -> None:
    procurement_cost = cost_score(423.71, 415.0, 60)
    procurement_satisfaction = satisfaction_score(0.9842, 40)
    assert abs(procurement_cost - 53.70) < 0.05, procurement_cost
    assert abs(procurement_satisfaction - 39.37) < 0.05, procurement_satisfaction

    assert deviation_score(0.1612, 0.05, 10) == 0
    assert abs(cost_score(107.90, 110.0, 20) - 20.0) < 0.01
    assert abs(cost_score(1809.29, 1600.0, 20) - 6.92) < 0.05
    assert abs(satisfaction_score(0.8198, 15) - 12.30) < 0.05
    assert abs(satisfaction_score(0.66, 35) - 23.10) < 0.05
    log("self-test passed")


def discover_xls(root: Path) -> list[Path]:
    skipped_dirs = {"verified_outputs", "__pycache__"}
    return sorted(
        path
        for path in root.rglob("*.xls")
        if not path.name.startswith("~$") and not any(part in skipped_dirs for part in path.parts)
    )


def generation_time_estimate(xls_path: Path, mode: str) -> str:
    keyword = case_keyword(xls_path)
    extreme = mode == "extreme"
    if extreme:
        if keyword == "热水器":
            return "预计 1-3 分钟；库存窗口极限分解，100% 满足后压低物流成本"
        if keyword in {"硫磺", "羽绒服", "蓄电池"}:
            return "预计 10-30 分钟；满满足优先，随后压低成本"
        return "预计 3-15 分钟；满满足优先，随后压低成本"
    if keyword == "热水器":
        return "预计 1-2 分钟；先找 100% 满足率快解"
    if keyword in {"硫磺", "羽绒服", "蓄电池"}:
        return "预计 2-5 分钟；先找 100% 满足率快解"
    return "预计 10-60 秒；先找 100% 满足率快解"


def main() -> None:
    global ACTIVE_HAR_CONTEXT
    parser = argparse.ArgumentParser(description="供应链 .xls 启发式方案生成器")
    parser.add_argument("xls", nargs="?", help="输入 .xls 文件路径")
    parser.add_argument("--har", nargs="+", help="易木平台导出的 HAR/前端 HTML/JS，可同时传运输计划和销售计划")
    parser.add_argument("--all", action="store_true", help="处理当前目录下所有 .xls")
    parser.add_argument("--self-test", action="store_true", help="运行评分公式自检")
    parser.add_argument("--check-env", action="store_true", help="检查 Python 依赖和 LibreOffice")
    parser.add_argument(
        "--mode",
        choices=("high", "extreme", "both"),
        default="both",
        help="求解版本：high=高分版快速少路线，extreme=极限版低成本搜索，both=两个版本都输出",
    )
    args = parser.parse_args()

    if args.check_env:
        if not check_environment():
            raise SystemExit(1)
        return

    if args.self_test:
        self_test()
        return

    if args.har:
        set_har_context(load_frontend_contexts(args.har))
        log(
            f"前端数据接入: {', '.join(args.har)} | 承运商={len((ACTIVE_HAR_CONTEXT or {}).get('carriers') or [])} "
            f"| 路段={len((ACTIVE_HAR_CONTEXT or {}).get('segment_distances') or {})} "
            f"| 业务请求={len((ACTIVE_HAR_CONTEXT or {}).get('business_entries') or [])}"
        )

    targets: list[Path]
    if args.all:
        targets = discover_xls(Path.cwd())
    elif args.xls:
        targets = [Path(args.xls)]
    else:
        parser.print_help()
        return

    if not targets:
        raise SystemExit("未找到 .xls 文件")

    modes = ["high", "extreme"] if args.mode == "both" else [args.mode]
    for xls_path in targets:
        for mode in modes:
            set_solver_mode(mode)
            ROUTE_TRANSPORT_OPTIONS_CACHE.clear()
            log(f"预计生成时间: {xls_path.name} | 版本={solver_mode_label()} | {generation_time_estimate(xls_path, mode)}")
            result = solve_file(xls_path)
            out_path = write_solution(result)
            raw_score = result.get("score")
            score_text = f"{float(raw_score):.1f}" if isinstance(raw_score, (int, float)) else "待正式指标"
            log(f"输出: {out_path} | 版本={result.get('solver_mode')} | 分数={score_text} | 校验={'通过' if result['simulation']['ok'] else '有风险'}")


if __name__ == "__main__":
    main()
